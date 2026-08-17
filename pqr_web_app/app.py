import csv
import hashlib
import io
import json
import os
import re
import tempfile
from datetime import datetime, timedelta

from flask import Flask, Response, flash, redirect, render_template, request, session
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from modules.auth import authenticate, can, current_user, is_effective_duty_officer, login_required, permission_required
from modules.db import (
    CLUSTER_EVENT_REGIONS,
    PQR_STATION_CODES,
    STATION_CLUSTER_BY_CODE,
    STATION_CLUSTERS,
    STATION_REGION_BY_CLUSTER,
    get_db,
    init_db,
)
from modules.google_sheets import append_pqr_reports_to_sheet, fetch_pqr_sheet_import_rows
from modules.imports import (
    MISSING_OFFICER_PLACEHOLDER,
    can_accept_missing_officer_only,
    ensure_event,
    import_data_sheet_csv,
    import_data_sheet_rows,
    import_row_hash,
)
from modules.phivolcs import (
    extract_bulletin_intensities,
    fetch_monthly_archive_events,
    fetch_recent_events,
    infer_region_code,
    location_suffix,
    phivolcs_url_indicates_felt,
)
from modules.timeutils import (
    APP_LOCAL_TZ,
    normalize_timestamp_to_utc_iso,
    parse_datetime_local_as_utc,
    parse_event_key_to_utc,
    parse_utc_iso,
    phivolcs_datetime_to_utc_key,
    to_utc_iso,
    utc_iso_to_pst_display,
    utc_now,
    within_update_window,
)
from modules.validation import optional_float, validate_event_form, validate_pqr_form, validate_pqr_form_strict

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-only-change-me")
if os.environ.get("FLASK_ENV") == "production":
    if app.secret_key == "dev-only-change-me" or len(app.secret_key) < 32:
        raise RuntimeError("Production requires SECRET_KEY with at least 32 characters.")
    app.config.update(
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE="Lax",
        SESSION_COOKIE_SECURE=os.environ.get("SESSION_COOKIE_SECURE", "1") != "0",
    )
init_db()

USER_ROLES = ("admin", "duty_officer", "station_user", "reviewer", "read_only")
PQR_DATE_RANGE_DAYS = 7
LOGIN_MAP_VIEWBOX = {
    "width": 702.39001,
    "height": 1209.4381,
    "west": 116.927573,
    "north": 20.834769,
    "east": 126.606549,
    "south": 4.640292,
}
MONTH_NAMES = (
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
)
QUARTER_MONTHS = {
    "1": (1, 2, 3),
    "2": (4, 5, 6),
    "3": (7, 8, 9),
    "4": (10, 11, 12),
}
SEMESTER_MONTHS = {
    "1": (1, 2, 3, 4, 5, 6),
    "2": (7, 8, 9, 10, 11, 12),
}


def ordinal_label(value):
    return {"1": "1st", "2": "2nd", "3": "3rd", "4": "4th"}.get(str(value), str(value))


@app.context_processor
def inject_user():
    now_pst = utc_now().astimezone(APP_LOCAL_TZ)
    user = current_user()
    assignment_summary = None
    unread_count = 0
    if user and user["role"] == "station_user":
        conn = get_db()
        try:
            assignment_summary = current_user_assignment_summary(conn, user)
        finally:
            conn.close()
    if user:
        conn = get_db()
        try:
            unread_count = unread_message_count(conn, user)
        finally:
            conn.close()
    return {
        "current_user": user,
        "current_user_assignment_summary": assignment_summary,
        "unread_message_count": unread_count,
        "can": can,
        "today_pst_display": f"{now_pst.strftime('%b')} {now_pst.day}, {now_pst.year}",
        "time_pst_display": now_pst.strftime("%I:%M %p"),
    }


@app.route("/healthz")
def healthz():
    return {"status": "ok"}


@app.before_request
def enforce_password_change():
    allowed_endpoints = {"healthz", "login", "logout", "change_password", "static"}
    if request.endpoint in allowed_endpoints:
        return None
    user = current_user()
    if user and user.get("must_change_password"):
        flash("Please change your temporary password before continuing.")
        return redirect("/change-password")
    return None


def pagination_args(default_per_page=50, max_per_page=100):
    try:
        page = int(request.args.get("page", 1))
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get("per_page", default_per_page))
    except ValueError:
        per_page = default_per_page
    page = max(page, 1)
    per_page = min(max(per_page, 1), max_per_page)
    return page, per_page, (page - 1) * per_page


def build_pagination(page, per_page, total_count):
    total_pages = max((total_count + per_page - 1) // per_page, 1)
    return {
        "page": page,
        "per_page": per_page,
        "total_count": total_count,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page": max(page - 1, 1),
        "next_page": min(page + 1, total_pages),
    }


def parse_local_date(value):
    try:
        return datetime.strptime(str(value or "").strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def default_pqr_date_range():
    today = utc_now().astimezone(APP_LOCAL_TZ).date()
    return today - timedelta(days=PQR_DATE_RANGE_DAYS - 1), today


def login_map_point(latitude, longitude):
    try:
        lat = float(latitude)
        lon = float(longitude)
    except (TypeError, ValueError):
        return None
    bounds = LOGIN_MAP_VIEWBOX
    if not (bounds["west"] <= lon <= bounds["east"] and bounds["south"] <= lat <= bounds["north"]):
        return None
    x = ((lon - bounds["west"]) / (bounds["east"] - bounds["west"])) * bounds["width"]
    y = ((bounds["north"] - lat) / (bounds["north"] - bounds["south"])) * bounds["height"]
    return round(x, 1), round(y, 1)


def login_map_marker_radius(magnitude):
    try:
        mag = float(magnitude or 0)
    except (TypeError, ValueError):
        mag = 0
    return round(min(max(3.8 + mag * 1.2, 4.5), 11), 1)


def today_login_earthquakes(conn):
    today = utc_now().astimezone(APP_LOCAL_TZ).date()
    start_local = datetime(today.year, today.month, today.day, tzinfo=APP_LOCAL_TZ)
    end_local = start_local + timedelta(days=1)
    rows = conn.execute(
        """
        SELECT event_key, event_datetime_utc, latitude, longitude, magnitude, reference_location
        FROM earthquake_events
        WHERE latitude IS NOT NULL
          AND longitude IS NOT NULL
          AND event_datetime_utc >= ?
          AND event_datetime_utc < ?
        ORDER BY event_datetime_utc DESC
        LIMIT 120
        """,
        (to_utc_iso(start_local), to_utc_iso(end_local)),
    ).fetchall()
    markers = []
    for index, row in enumerate(rows):
        point = login_map_point(row["latitude"], row["longitude"])
        if not point:
            continue
        event_time = parse_utc_iso(row["event_datetime_utc"])
        markers.append(
            {
                "event_key": row["event_key"],
                "x": point[0],
                "y": point[1],
                "r": login_map_marker_radius(row["magnitude"]),
                "magnitude": row["magnitude"],
                "reference_location": row["reference_location"],
                "time_label": event_time.astimezone(APP_LOCAL_TZ).strftime("%I:%M %p") if event_time else "",
                "latest": index == 0,
            }
        )
    return list(reversed(markers))


def event_local_date(event_datetime_utc):
    event_time = parse_utc_iso(event_datetime_utc)
    if not event_time:
        return None
    return event_time.astimezone(APP_LOCAL_TZ).date()


def event_local_month(event_datetime_utc):
    event_date = event_local_date(event_datetime_utc)
    return event_date.strftime("%Y-%m") if event_date else ""


def parse_month_value(value):
    text = str(value or "").strip()
    try:
        return datetime.strptime(text, "%Y-%m").strftime("%Y-%m")
    except ValueError:
        return ""


def parse_year_value(value, default_year=None):
    try:
        year = int(value)
    except (TypeError, ValueError):
        return default_year
    if 2000 <= year <= 2100:
        return year
    return default_year


def month_display_label(value):
    parsed = parse_month_value(value)
    if not parsed:
        return ""
    return datetime.strptime(parsed, "%Y-%m").strftime("%B %Y")


def build_year_month_options(event_rows, month_value="", year_value=""):
    current_date = utc_now().astimezone(APP_LOCAL_TZ).date()
    requested_month = parse_month_value(month_value)
    requested_year = parse_year_value(year_value)
    if requested_year is None and requested_month:
        requested_year = int(requested_month[:4])
    selected_year = requested_year or current_date.year

    event_counts = {}
    event_years = set()
    for row in event_rows:
        event_month = event_local_month(row["event_datetime_utc"])
        if not event_month:
            continue
        event_counts[event_month] = event_counts.get(event_month, 0) + 1
        event_years.add(int(event_month[:4]))
    year_options = sorted(event_years | {selected_year, current_date.year, 2026}, reverse=True)

    if requested_month and int(requested_month[:4]) == selected_year:
        selected_month = requested_month
    elif selected_year == current_date.year:
        selected_month = f"{selected_year}-{current_date.month:02d}"
    else:
        selected_month = f"{selected_year}-01"

    month_options = []
    for month in range(1, 13):
        value = f"{selected_year}-{month:02d}"
        count = event_counts.get(value, 0)
        event_word = "event" if count == 1 else "events"
        month_options.append(
            {
                "value": value,
                "label": f"{MONTH_NAMES[month - 1]} {selected_year} ({count} {event_word})",
                "event_count": count,
            }
        )
    return {
        "selected_year": selected_year,
        "selected_month": selected_month,
        "month_options": month_options,
        "year_options": year_options,
    }


def no_required_compliance_label(required):
    return "-" if not required else None


def normalize_month_range(start_value, end_value, fallback_month):
    start_month = parse_month_value(start_value) or parse_month_value(fallback_month)
    end_month = parse_month_value(end_value) or start_month
    if start_month and end_month and start_month > end_month:
        start_month, end_month = end_month, start_month
    return start_month, end_month


def month_range_label(start_month, end_month):
    if not start_month:
        return ""
    if start_month == end_month or not end_month:
        return month_display_label(start_month)
    return f"{month_display_label(start_month)} to {month_display_label(end_month)}"


def format_percent(value):
    if value is None:
        return "-"
    return f"{value:.2f}%"


DEFAULT_MONITORING_SIGNATURES = {
    "prepared_name": "JOEL S. OESTAR",
    "prepared_title": "Science Research Analyst",
    "reviewed_name": "JOHNLERY P. DEXIMO",
    "reviewed_title": "Senior Research Specialist",
    "approved_name": "ANGELITO G. LANUZA",
    "approved_title": "Supervising SRS",
    "noted_name": "WINCHELLE IAN G. SEVILLA",
    "noted_title": "Chief, SOEPD",
}


def monitoring_signatures_from_request():
    signatures = {}
    for key, default_value in DEFAULT_MONITORING_SIGNATURES.items():
        signatures[key] = request.args.get(key, default_value).strip() or default_value
    return signatures


def report_timestamp_label():
    local_time = utc_now().astimezone(APP_LOCAL_TZ)
    return f"{local_time.month}/{local_time.day}/{local_time.year} {local_time:%H:%M:%S}"


def parse_monitoring_period(period_type, year_value, period_value):
    today = utc_now().astimezone(APP_LOCAL_TZ).date()
    period_type = (period_type or "quarter").strip().lower()
    if period_type not in {"quarter", "semester"}:
        period_type = "quarter"
    try:
        year = int(year_value or today.year)
    except (TypeError, ValueError):
        year = today.year
    if year < 2000 or year > 2100:
        year = today.year

    if period_type == "semester":
        default_period = "1" if today.month <= 6 else "2"
        period = str(period_value or default_period)
        if period not in SEMESTER_MONTHS:
            period = default_period
        months = SEMESTER_MONTHS[period]
        period_name = f"{ordinal_label(period)} Semester"
        file_label = f"{year}_S{period}"
    else:
        default_period = str(((today.month - 1) // 3) + 1)
        period = str(period_value or default_period)
        if period not in QUARTER_MONTHS:
            period = default_period
        months = QUARTER_MONTHS[period]
        period_name = f"{ordinal_label(period)} Quarter"
        file_label = f"{year}_Q{period}"

    month_labels = [{"number": month, "name": MONTH_NAMES[month - 1], "key": f"{year}-{month:02d}"} for month in months]
    subtitle = f"for the {period_name} ({MONTH_NAMES[months[0] - 1]} to {MONTH_NAMES[months[-1] - 1]} {year})"
    return {
        "type": period_type,
        "year": year,
        "period": period,
        "months": month_labels,
        "subtitle": subtitle,
        "file_label": file_label,
    }


def event_local_datetime_label(event_datetime_utc):
    event_time = parse_utc_iso(event_datetime_utc)
    if not event_time:
        return ""
    return event_time.astimezone(APP_LOCAL_TZ).strftime("%d %B %Y - %I:%M %p")


def in_one_manned_exemption_window(event_datetime_utc):
    event_time = parse_utc_iso(event_datetime_utc)
    if not event_time:
        return False
    local_time = event_time.astimezone(APP_LOCAL_TZ)
    weekday = local_time.weekday()
    hour = local_time.hour + (local_time.minute / 60) + (local_time.second / 3600)
    return (
        (weekday == 4 and hour >= 15)
        or weekday in {5, 6}
        or (weekday == 0 and hour < 10)
    )


def station_is_one_manned(station):
    return bool(dict(station or {}).get("is_one_manned"))


def station_exempt_from_pqr(station, event):
    return (
        station_is_one_manned(station)
        and in_one_manned_exemption_window(event.get("event_datetime_utc"))
        and not within_update_window(event.get("event_datetime_utc"))
    )


def client_ip():
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.remote_addr or ""


def log_login_event(username, event_type, user=None):
    conn = get_db()
    try:
        conn.execute(
            """
            INSERT INTO login_audit_logs (
                user_id, username, role, station_id, event_type, ip_address, user_agent
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user["id"] if user else None,
                (username or "").strip(),
                user["role"] if user else None,
                user["station_id"] if user else None,
                event_type,
                client_ip(),
                request.headers.get("User-Agent", ""),
            ),
        )
        conn.commit()
    finally:
        conn.close()


def validate_user_form(form, require_password=False):
    errors = []
    username = (form.get("username") or "").strip()
    display_name = (form.get("display_name") or "").strip()
    role = form.get("role") or ""
    password = form.get("password") or ""
    station_id = form.get("station_id") or ""
    station_assignment_mode = form.get("station_assignment_mode") or "existing"

    if not username:
        errors.append("Username is required.")
    if not display_name:
        errors.append("Display name is required.")
    if role not in USER_ROLES:
        errors.append("Role is invalid.")
    if require_password and len(password) < 8:
        errors.append("Password must be at least 8 characters.")
    if role == "station_user":
        if form.get("station_status") == "custom" and not (form.get("custom_station_status") or "").strip():
            errors.append("Custom station status is required.")
        if station_assignment_mode == "custom":
            if not (form.get("custom_station_code") or "").strip():
                errors.append("Custom station code is required.")
            if not (form.get("custom_station_name") or "").strip():
                errors.append("Custom station name is required.")
            if form.get("custom_station_cluster") not in STATION_CLUSTERS:
                errors.append("Custom station cluster is required.")
        elif not station_id:
            errors.append("Station user must be assigned to a station.")
    return errors


def station_assignment_ids(conn, user, permission="submit"):
    if not user or user["role"] != "station_user" or is_effective_duty_officer(user):
        return []
    permission_column = "can_submit" if permission == "submit" else "can_view"
    rows = conn.execute(
        f"""
        SELECT station_id
        FROM user_station_assignments
        WHERE user_id = ?
          AND {permission_column} = 1
        ORDER BY is_primary DESC, station_id
        """,
        (user["id"],),
    ).fetchall()
    station_ids = [row["station_id"] for row in rows]
    primary_station_id = user.get("station_id")
    if primary_station_id and primary_station_id not in station_ids:
        station_ids.insert(0, primary_station_id)
    return station_ids


def user_can_access_station(conn, user, station_id, permission="submit"):
    if not user or user["role"] != "station_user" or is_effective_duty_officer(user):
        return True
    return str(station_id) in {str(item) for item in station_assignment_ids(conn, user, permission)}


def assigned_station_rows(conn, user, permission="submit"):
    station_ids = station_assignment_ids(conn, user, permission)
    if not station_ids:
        return []
    placeholders = ", ".join("?" for _ in station_ids)
    rows = conn.execute(
        f"""
        SELECT id, station_code, station_name, region_code, cluster_name,
               station_type, include_in_pqr_compliance, is_one_manned, station_status
        FROM stations
        WHERE is_active = 1
          AND id IN ({placeholders})
          {"AND include_in_pqr_compliance = 1" if permission == "submit" else ""}
        ORDER BY CASE id
            {" ".join(f"WHEN ? THEN {index}" for index, _ in enumerate(station_ids))}
            ELSE 999
        END
        """,
        station_ids + station_ids,
    ).fetchall()
    return rows


def station_id_filter_sql(station_ids, table_alias="stations"):
    if not station_ids:
        return " AND 1 = 0", []
    placeholders = ", ".join("?" for _ in station_ids)
    return f" AND {table_alias}.id IN ({placeholders})", list(station_ids)


def form_station_assignment_ids(form, primary_station_id):
    assigned_ids = [value for value in form.getlist("assigned_station_ids") if str(value).strip()]
    if primary_station_id:
        primary_text = str(primary_station_id)
        assigned_ids = [value for value in assigned_ids if str(value) != primary_text]
        assigned_ids.insert(0, primary_text)
    seen = set()
    normalized = []
    for station_id in assigned_ids:
        text = str(station_id).strip()
        if text and text not in seen:
            normalized.append(text)
            seen.add(text)
    return normalized


def sync_user_station_assignments(conn, user_id, role, primary_station_id, assigned_station_ids):
    conn.execute("DELETE FROM user_station_assignments WHERE user_id = ?", (user_id,))
    if role != "station_user":
        return
    if assigned_station_ids:
        placeholders = ", ".join("?" for _ in assigned_station_ids)
        valid_station_ids = {
            str(row["id"])
            for row in conn.execute(
                f"SELECT id FROM stations WHERE id IN ({placeholders})",
                assigned_station_ids,
            ).fetchall()
        }
        assigned_station_ids = [station_id for station_id in assigned_station_ids if str(station_id) in valid_station_ids]
    for station_id in assigned_station_ids:
        conn.execute(
            """
            INSERT OR IGNORE INTO user_station_assignments (
                user_id, station_id, is_primary, can_submit, can_view
            )
            VALUES (?, ?, ?, 1, 1)
            """,
            (user_id, station_id, 1 if str(station_id) == str(primary_station_id) else 0),
        )


def current_user_assignment_summary(conn, user):
    if not user or user["role"] != "station_user" or is_effective_duty_officer(user):
        return None
    stations = assigned_station_rows(conn, user, "view")
    if not stations:
        return None
    primary = next((station for station in stations if station["id"] == user.get("station_id")), stations[0])
    extra_count = max(len(stations) - 1, 0)
    return {
        "primary": primary,
        "stations": stations,
        "label": f"{primary['station_code']} + {extra_count} station(s)" if extra_count else primary["station_code"],
    }


def validate_password_change(user, current_password, new_password, confirm_password):
    errors = []
    if not check_password_hash(user["password_hash"], current_password or ""):
        errors.append("Current password is incorrect.")
    if len(new_password or "") < 8:
        errors.append("New password must be at least 8 characters.")
    if (new_password or "") != (confirm_password or ""):
        errors.append("New password and confirmation do not match.")
    if (new_password or "").strip().lower() == (user["username"] or "").strip().lower():
        errors.append("New password must not be the same as your username.")
    if check_password_hash(user["password_hash"], new_password or ""):
        errors.append("New password must be different from your current password.")
    return errors


def active_stations(conn):
    return conn.execute(
        """
        SELECT id, station_code, station_name, region_code, cluster_name,
               station_type, include_in_pqr_compliance, is_one_manned, station_status
        FROM stations
        WHERE is_active = 1
        ORDER BY station_code
        """
    ).fetchall()


REGION_CHOICES = (
    ("NL", "North Luzon"),
    ("SL", "South Luzon"),
    ("VIS", "Visayas"),
    ("MIN", "Mindanao"),
)


def normalize_rule_text(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def province_region_from_rules(conn, location):
    location_text = normalize_rule_text(location)
    if not location_text:
        return None
    province_parts = re.findall(r"\(([^()]*)\)", str(location or ""))
    search_texts = [normalize_rule_text(part) for part in reversed(province_parts)]
    search_texts.append(location_text)
    rules = conn.execute(
        """
        SELECT province_name, region_code
        FROM province_region_rules
        WHERE is_active = 1
        ORDER BY priority ASC, province_name ASC
        """
    ).fetchall()
    for rule in rules:
        province = normalize_rule_text(rule["province_name"])
        if not province:
            continue
        for text in search_texts:
            if province in text:
                return rule["region_code"]
    return None


def infer_event_region_code(conn, location):
    return province_region_from_rules(conn, location) or infer_region_code(location)


def apply_event_region_rules(conn, event):
    dynamic_region = infer_event_region_code(conn, getattr(event, "reference_location", ""))
    if dynamic_region and dynamic_region != getattr(event, "region_code", None):
        event.region_code = dynamic_region
    return event


def recompute_event_regions(conn, year=None, month=None):
    where = ["COALESCE(exclude_from_pqr_rating, 0) = 0"]
    params = []
    if year and month:
        start = datetime(year, month, 1)
        end = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
        where.append("event_datetime_utc >= ? AND event_datetime_utc < ?")
        params.extend([start.strftime("%Y-%m-%dT00:00:00Z"), end.strftime("%Y-%m-%dT00:00:00Z")])
    elif year:
        where.append("event_datetime_utc >= ? AND event_datetime_utc < ?")
        params.extend([f"{year:04d}-01-01T00:00:00Z", f"{year + 1:04d}-01-01T00:00:00Z"])

    rows = conn.execute(
        f"""
        SELECT id, reference_location, region_code
        FROM earthquake_events
        WHERE {" AND ".join(where)}
        """,
        params,
    ).fetchall()
    changed = 0
    for row in rows:
        new_region = infer_event_region_code(conn, row["reference_location"])
        if new_region and new_region != row["region_code"]:
            conn.execute(
                "UPDATE earthquake_events SET region_code = ? WHERE id = ?",
                (new_region, row["id"]),
            )
            changed += 1
    reconcile_required_submissions(conn)
    return {"checked": len(rows), "changed": changed}


def cluster_name_for_station(station_code, region_code=None):
    if isinstance(station_code, dict):
        station = station_code
        cluster_name = station.get("cluster_name")
        if cluster_name:
            return cluster_name
        station_code = station.get("station_code")
        region_code = station.get("region_code", region_code)
    cluster_name = STATION_CLUSTER_BY_CODE.get(station_code)
    if cluster_name:
        return cluster_name
    for name, code in STATION_REGION_BY_CLUSTER.items():
        if code == region_code:
            return name
    return ""


def station_cluster_sort_key(station):
    cluster_name = cluster_name_for_station(station) or ""
    station_code = dict(station or {}).get("station_code", "")
    station_order = list(STATION_CLUSTERS.get(cluster_name, ()))
    if station_code in station_order:
        return (0, station_order.index(station_code), station_code)
    return (1, 999, station_code)


def resolve_station_assignment(conn, form, errors):
    role = form.get("role") or ""
    if role != "station_user":
        return None

    station_status_label = resolve_station_status_label(form)
    station_status = 1 if station_status_label == "1M" else 0
    if (form.get("station_assignment_mode") or "existing") != "custom":
        station_id = form.get("station_id") or None
        if station_id:
            conn.execute(
                "UPDATE stations SET is_one_manned = ?, station_status = ? WHERE id = ?",
                (station_status, station_status_label, station_id),
            )
        return station_id

    station_code = (form.get("custom_station_code") or "").strip().upper()
    station_name = (form.get("custom_station_name") or "").strip()
    cluster_name = form.get("custom_station_cluster") or ""
    region_code = STATION_REGION_BY_CLUSTER.get(cluster_name)
    if not station_code or not station_name or not region_code:
        return None

    existing_station = conn.execute(
        "SELECT id FROM stations WHERE station_code = ?",
        (station_code,),
    ).fetchone()
    if existing_station:
        errors.append(f"Station code {station_code} already exists. Select it from the list instead.")
        return None

    cursor = conn.execute(
        """
        INSERT INTO stations (station_code, station_name, region_code, cluster_name, is_one_manned, station_status, is_active)
        VALUES (?, ?, ?, ?, ?, ?, 1)
        """,
        (station_code, station_name, region_code, cluster_name, station_status, station_status_label),
    )
    return cursor.lastrowid


def resolve_station_status_label(form):
    station_status = (form.get("station_status") or "regular").strip()
    if station_status == "1M":
        return "1M"
    if station_status == "custom":
        custom_status = (form.get("custom_station_status") or "").strip()
        return custom_status or "Regular Station"
    return "Regular Station"


def station_time_left(event_datetime_utc):
    event_time = parse_utc_iso(event_datetime_utc)
    if not event_time:
        return {"label": "Unknown", "hours_left": 0, "seconds_left": 0}

    deadline = event_time + timedelta(hours=22)
    seconds_left = max(int((deadline - utc_now()).total_seconds()), 0)
    hours = seconds_left // 3600
    minutes = (seconds_left % 3600) // 60

    if seconds_left <= 0:
        label = "Expired"
    elif hours > 0:
        label = f"{hours}h {minutes}m"
    else:
        label = f"{minutes}m"

    return {"label": label, "hours_left": hours, "seconds_left": seconds_left}


MESSAGE_CATEGORIES = {
    "all": "All",
    "action": "Action Required",
    "announcement": "Announcements",
    "system": "System Alerts",
    "review": "Review",
    "read": "Read",
}
MESSAGE_PRIORITIES = {
    "urgent": "Urgent",
    "due_soon": "Due Soon",
    "info": "Info",
    "resolved": "Resolved",
}


def message_created_label(value):
    parsed = parse_utc_iso(value)
    if not parsed:
        try:
            parsed = datetime.fromisoformat((value or "").replace("Z", "+00:00"))
        except ValueError:
            return value or ""
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=APP_LOCAL_TZ)
    return parsed.astimezone(APP_LOCAL_TZ).strftime("%d %b %Y, %I:%M %p")


def read_message_keys(conn, user):
    if not user:
        return set()
    rows = conn.execute(
        "SELECT message_key FROM message_reads WHERE user_id = ?",
        (user["id"],),
    ).fetchall()
    return {row["message_key"] for row in rows}


def stored_messages_for_user(conn, user):
    if not user:
        return []
    role_values = {user.get("role"), user.get("effective_role")}
    role_values.discard(None)
    target_station_ids = []
    if user["role"] == "station_user":
        target_station_ids = station_assignment_ids(conn, user, "view")

    role_placeholders = ", ".join("?" for _ in role_values) or "NULL"
    station_filter = "messages.target_station_id IS NULL"
    params = [user["id"], *role_values]
    if target_station_ids:
        station_placeholders = ", ".join("?" for _ in target_station_ids)
        station_filter = f"(messages.target_station_id IS NULL OR messages.target_station_id IN ({station_placeholders}))"
        params.extend(target_station_ids)

    rows = conn.execute(
        f"""
        SELECT messages.*, users.display_name AS created_by_name,
               stations.station_code AS target_station_code
        FROM messages
        LEFT JOIN users ON users.id = messages.created_by
        LEFT JOIN stations ON stations.id = messages.target_station_id
        WHERE messages.is_active = 1
          AND (
            messages.target_user_id = ?
            OR messages.created_by = ?
            OR (
              messages.target_user_id IS NULL
              AND (messages.target_role IS NULL OR messages.target_role IN ({role_placeholders}))
              AND {station_filter}
            )
          )
        ORDER BY messages.created_at DESC, messages.id DESC
        """,
        [user["id"], *params],
    ).fetchall()
    return [
        {
            **dict(row),
            "key": f"stored:{row['id']}",
            "source": "stored",
            "created_label": message_created_label(row["created_at"]),
        }
        for row in rows
    ]


def station_pending_messages(conn, user):
    station_ids = station_assignment_ids(conn, user, "submit")
    if not station_ids:
        return []
    placeholders = ", ".join("?" for _ in station_ids)
    rows = conn.execute(
        f"""
        SELECT earthquake_events.id AS event_id, earthquake_events.event_key,
               earthquake_events.event_datetime_utc, earthquake_events.magnitude,
               earthquake_events.reference_location, stations.id AS station_id,
               stations.station_code
        FROM pqr_required_submissions
        JOIN earthquake_events ON earthquake_events.id = pqr_required_submissions.event_id
        JOIN stations ON stations.id = pqr_required_submissions.station_id
        LEFT JOIN pqr_reports
          ON pqr_reports.event_id = earthquake_events.id
         AND pqr_reports.station_id = pqr_required_submissions.station_id
        WHERE pqr_required_submissions.station_id IN ({placeholders})
          AND pqr_required_submissions.status = 'pending'
          AND earthquake_events.status = 'open'
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
          AND pqr_reports.id IS NULL
        ORDER BY earthquake_events.event_datetime_utc DESC
        LIMIT 40
        """,
        station_ids,
    ).fetchall()
    messages = []
    for row in rows:
        if not within_update_window(row["event_datetime_utc"]):
            continue
        time_left = station_time_left(row["event_datetime_utc"])
        due_soon = time_left["seconds_left"] > 0 and time_left["hours_left"] <= 3
        messages.append(
            {
                "id": None,
                "key": f"pending:{row['event_id']}:{row['station_id']}",
                "source": "generated",
                "title": f"PQR needed for {row['event_key']}",
                "body": (
                    f"{row['station_code']} has a pending PQR for M{row['magnitude'] or 'N/A'} "
                    f"near {row['reference_location']}. Time left: {time_left['label']}."
                ),
                "category": "action",
                "priority": "due_soon" if due_soon else "urgent",
                "action_url": f"/pqr/new?station_id={row['station_id']}&event_id={row['event_id']}",
                "created_at": row["event_datetime_utc"],
                "created_label": utc_iso_to_pst_display(row["event_datetime_utc"]),
                "target_station_code": row["station_code"],
            }
        )
    return messages


def operations_pending_messages(conn):
    rows = conn.execute(
        """
        SELECT earthquake_events.id AS event_id, earthquake_events.event_key,
               earthquake_events.event_datetime_utc, earthquake_events.magnitude,
               earthquake_events.reference_location,
               COUNT(pqr_required_submissions.id) AS required_count,
               SUM(CASE WHEN pqr_required_submissions.status = 'pending' THEN 1 ELSE 0 END) AS pending_count
        FROM earthquake_events
        JOIN pqr_required_submissions ON pqr_required_submissions.event_id = earthquake_events.id
        WHERE earthquake_events.status = 'open'
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
        GROUP BY earthquake_events.id
        HAVING SUM(CASE WHEN pqr_required_submissions.status = 'pending' THEN 1 ELSE 0 END) > 0
        ORDER BY earthquake_events.event_datetime_utc DESC
        LIMIT 25
        """
    ).fetchall()
    messages = []
    for row in rows:
        time_left = station_time_left(row["event_datetime_utc"])
        due_soon = time_left["seconds_left"] > 0 and time_left["hours_left"] <= 3
        messages.append(
            {
                "id": None,
                "key": f"event-pending:{row['event_id']}:{row['pending_count']}",
                "source": "generated",
                "title": f"{row['pending_count']} pending submission(s) for {row['event_key']}",
                "body": (
                    f"M{row['magnitude'] or 'N/A'} near {row['reference_location']} has "
                    f"{row['pending_count']} of {row['required_count']} required station response(s) still pending."
                ),
                "category": "action",
                "priority": "due_soon" if due_soon else "info",
                "action_url": "/stations/status",
                "created_at": row["event_datetime_utc"],
                "created_label": utc_iso_to_pst_display(row["event_datetime_utc"]),
                "target_station_code": "",
            }
        )
    return messages


def sync_failure_messages(conn):
    row = conn.execute(
        """
        SELECT COUNT(*) AS failed_count, MAX(pqr_reports.id) AS latest_report_id
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        WHERE pqr_reports.sheet_sync_status = 'failed'
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
        """
    ).fetchone()
    if not row or not row["failed_count"]:
        return []
    return [
        {
            "id": None,
            "key": f"sync-failed:{row['failed_count']}:{row['latest_report_id']}",
            "source": "generated",
            "title": "Google Sheets sync needs attention",
            "body": f"{row['failed_count']} PQR report(s) failed to sync. Retry from the dashboard or reports page.",
            "category": "system",
            "priority": "urgent",
            "action_url": "/reports",
            "created_at": "",
            "created_label": "Current",
            "target_station_code": "",
        }
    ]


def generated_messages_for_user(conn, user):
    if not user:
        return []
    messages = []
    if can(user, "submit") and user["role"] == "station_user" and not is_effective_duty_officer(user):
        messages.extend(station_pending_messages(conn, user))
    if can(user, "review") or can(user, "create_event"):
        messages.extend(operations_pending_messages(conn))
    if can(user, "export"):
        messages.extend(sync_failure_messages(conn))
    return messages


def messages_for_user(conn, user, selected_filter="all"):
    read_keys = read_message_keys(conn, user)
    messages = stored_messages_for_user(conn, user) + generated_messages_for_user(conn, user)
    for message in messages:
        message["is_read"] = message["key"] in read_keys
        message["category_label"] = MESSAGE_CATEGORIES.get(message["category"], message["category"].title())
        message["priority_label"] = MESSAGE_PRIORITIES.get(message["priority"], message["priority"].title())
    if selected_filter == "read":
        messages = [message for message in messages if message["is_read"]]
    elif selected_filter in MESSAGE_CATEGORIES and selected_filter != "all":
        messages = [message for message in messages if message["category"] == selected_filter and not message["is_read"]]
    else:
        messages = [message for message in messages if not message["is_read"]]
    priority_order = {"urgent": 0, "due_soon": 1, "info": 2, "resolved": 3}
    return sorted(
        messages,
        key=lambda item: (item["is_read"], priority_order.get(item["priority"], 9), item.get("created_at") or ""),
    )


def unread_message_count(conn, user):
    return len(messages_for_user(conn, user, "all")) if user else 0


def mark_message_key_read(conn, user, message_key):
    visible_keys = {message["key"] for message in messages_for_user(conn, user, "all")}
    visible_keys.update(message["key"] for message in messages_for_user(conn, user, "read"))
    if message_key not in visible_keys:
        return False
    conn.execute(
        """
        INSERT OR IGNORE INTO message_reads (user_id, message_key)
        VALUES (?, ?)
        """,
        (user["id"], message_key),
    )
    return True


def stored_message_detail(conn, user, message_id):
    messages = stored_messages_for_user(conn, user)
    allowed_keys = {message["key"] for message in messages}
    message = conn.execute(
        """
        SELECT messages.*, users.display_name AS created_by_name,
               stations.station_code AS target_station_code,
               stations.station_name AS target_station_name
        FROM messages
        LEFT JOIN users ON users.id = messages.created_by
        LEFT JOIN stations ON stations.id = messages.target_station_id
        WHERE messages.id = ?
          AND messages.is_active = 1
        """,
        (message_id,),
    ).fetchone()
    if not message or f"stored:{message['id']}" not in allowed_keys:
        return None
    message = dict(message)
    message["key"] = f"stored:{message['id']}"
    message["created_label"] = message_created_label(message["created_at"])
    message["category_label"] = MESSAGE_CATEGORIES.get(message["category"], message["category"].title())
    message["priority_label"] = MESSAGE_PRIORITIES.get(message["priority"], message["priority"].title())
    return message


def message_replies(conn, message_id):
    rows = conn.execute(
        """
        SELECT message_replies.*, users.display_name, users.role,
               stations.station_code
        FROM message_replies
        JOIN users ON users.id = message_replies.created_by
        LEFT JOIN stations ON stations.id = users.station_id
        WHERE message_replies.message_id = ?
        ORDER BY message_replies.created_at ASC, message_replies.id ASC
        """,
        (message_id,),
    ).fetchall()
    return [
        {
            **dict(row),
            "created_label": message_created_label(row["created_at"]),
        }
        for row in rows
    ]


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        user = authenticate(username, request.form.get("password", ""))
        if user:
            session["user_id"] = user["id"]
            conn = get_db()
            conn.execute("UPDATE users SET last_login_at = CURRENT_TIMESTAMP WHERE id = ?", (user["id"],))
            conn.commit()
            conn.close()
            log_login_event(username, "login_success", user)
            flash("Signed in.")
            if user.get("must_change_password"):
                return redirect("/change-password")
            if user["role"] == "reviewer":
                return redirect("/pqr/list")
            return redirect("/dashboard")
        log_login_event(username, "login_failed")
        flash("Invalid username or password.")
    conn = get_db()
    try:
        login_earthquakes = today_login_earthquakes(conn)
    finally:
        conn.close()
    return render_template("login.html", login_earthquakes=login_earthquakes)


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    user = current_user()
    if request.method == "POST":
        errors = validate_password_change(
            user,
            request.form.get("current_password", ""),
            request.form.get("new_password", ""),
            request.form.get("confirm_password", ""),
        )
        if errors:
            for error in errors:
                flash(error)
        else:
            conn = get_db()
            conn.execute(
                """
                UPDATE users
                SET password_hash = ?,
                    must_change_password = 0,
                    password_changed_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (generate_password_hash(request.form.get("new_password", "")), user["id"]),
            )
            conn.commit()
            conn.close()
            flash("Password updated successfully.")
            if user["role"] == "reviewer":
                return redirect("/pqr/list")
            return redirect("/dashboard")
    return render_template("change_password.html", force_change=bool(user.get("must_change_password")))


@app.route("/logout")
def logout():
    user = current_user()
    if user:
        log_login_event(user["username"], "logout", user)
    session.clear()
    flash("Signed out.")
    return redirect("/login")


@app.route("/messages")
@login_required
def messages_inbox():
    selected_filter = request.args.get("filter", "all").strip()
    if selected_filter not in MESSAGE_CATEGORIES:
        selected_filter = "all"
    user = current_user()
    conn = get_db()
    try:
        inbox_messages = messages_for_user(conn, user, selected_filter)
        unread_count = unread_message_count(conn, user)
    finally:
        conn.close()
    return render_template(
        "messages.html",
        messages=inbox_messages,
        categories=MESSAGE_CATEGORIES,
        selected_filter=selected_filter,
        unread_count=unread_count,
        title="Messages",
    )


@app.route("/messages/read", methods=["POST"])
@login_required
def mark_message_read():
    user = current_user()
    message_key = request.form.get("message_key", "").strip()
    conn = get_db()
    try:
        marked = mark_message_key_read(conn, user, message_key)
        conn.commit()
    finally:
        conn.close()
    if not marked:
        flash("Message was not found or is no longer available.")
    return redirect(request.form.get("next") or "/messages")


@app.route("/messages/<int:message_id>", methods=["GET", "POST"])
@login_required
def message_thread(message_id):
    user = current_user()
    conn = get_db()
    try:
        message = stored_message_detail(conn, user, message_id)
        if not message:
            conn.close()
            flash("Message was not found or is no longer available.")
            return redirect("/messages")
        if request.method == "POST":
            body = request.form.get("body", "").strip()
            if not body:
                flash("Reply cannot be empty.")
            else:
                conn.execute(
                    """
                    INSERT INTO message_replies (message_id, body, created_by)
                    VALUES (?, ?, ?)
                    """,
                    (message_id, body, user["id"]),
                )
                conn.execute(
                    """
                    INSERT OR IGNORE INTO message_reads (user_id, message_key)
                    VALUES (?, ?)
                    """,
                    (user["id"], message["key"]),
                )
                conn.commit()
                flash("Reply sent.")
                return redirect(f"/messages/{message_id}")
        mark_message_key_read(conn, user, message["key"])
        conn.commit()
        replies = message_replies(conn, message_id)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return render_template(
        "message_thread.html",
        message=message,
        replies=replies,
        title=message["title"],
    )


@app.route("/messages/read-all", methods=["POST"])
@login_required
def mark_all_messages_read():
    user = current_user()
    selected_filter = request.form.get("filter", "all").strip()
    conn = get_db()
    try:
        visible_messages = messages_for_user(conn, user, selected_filter)
        for message in visible_messages:
            conn.execute(
                """
                INSERT OR IGNORE INTO message_reads (user_id, message_key)
                VALUES (?, ?)
                """,
                (user["id"], message["key"]),
            )
        conn.commit()
    finally:
        conn.close()
    flash("Visible messages marked as read.")
    return redirect(f"/messages?filter={selected_filter}" if selected_filter else "/messages")


@app.route("/admin/messages/new", methods=["GET", "POST"])
@login_required
@permission_required("manage")
def admin_new_message():
    conn = get_db()
    users = conn.execute(
        """
        SELECT users.id, users.display_name, users.username, users.role,
               stations.station_code
        FROM users
        LEFT JOIN stations ON stations.id = users.station_id
        WHERE users.is_active = 1
        ORDER BY users.display_name, users.username
        """
    ).fetchall()
    stations = conn.execute(
        """
        SELECT id, station_code, station_name
        FROM stations
        WHERE is_active = 1
        ORDER BY station_code
        """
    ).fetchall()
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        body = request.form.get("body", "").strip()
        category = request.form.get("category", "announcement").strip()
        priority = request.form.get("priority", "info").strip()
        target_role = request.form.get("target_role", "").strip() or None
        target_user_id = request.form.get("target_user_id", "").strip() or None
        target_station_id = request.form.get("target_station_id", "").strip() or None
        action_url = request.form.get("action_url", "").strip()
        errors = []
        if not title:
            errors.append("Title is required.")
        if not body:
            errors.append("Message body is required.")
        if category not in {"action", "announcement", "system", "review"}:
            errors.append("Choose a valid category.")
        if priority not in MESSAGE_PRIORITIES:
            errors.append("Choose a valid priority.")
        if target_role and target_role not in USER_ROLES:
            errors.append("Choose a valid target role.")
        if errors:
            for error in errors:
                flash(error)
        else:
            conn.execute(
                """
                INSERT INTO messages (
                    title, body, category, priority, target_role, target_user_id,
                    target_station_id, action_url, created_by
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    title,
                    body,
                    category,
                    priority,
                    target_role,
                    target_user_id,
                    target_station_id,
                    action_url,
                    current_user()["id"],
                ),
            )
            conn.commit()
            conn.close()
            flash("Message published.")
            return redirect("/messages")
    conn.close()
    return render_template(
        "admin_message_form.html",
        categories={key: value for key, value in MESSAGE_CATEGORIES.items() if key not in {"all", "read"}},
        priorities=MESSAGE_PRIORITIES,
        roles=USER_ROLES,
        users=users,
        stations=stations,
        title="New Message",
    )


@app.route("/admin/login-logs")
@login_required
@permission_required("manage")
def admin_login_logs():
    page, per_page, offset = pagination_args(default_per_page=50, max_per_page=100)
    event_type = request.args.get("event_type", "").strip()
    username = request.args.get("username", "").strip()
    where = "WHERE 1 = 1"
    params = []
    if event_type:
        where += " AND login_audit_logs.event_type = ?"
        params.append(event_type)
    if username:
        where += " AND login_audit_logs.username LIKE ?"
        params.append(f"%{username}%")

    conn = get_db()
    total_count = conn.execute(
        f"SELECT COUNT(*) AS count FROM login_audit_logs {where}",
        params,
    ).fetchone()["count"]
    logs = conn.execute(
        f"""
        SELECT login_audit_logs.*, stations.station_code, stations.station_name
        FROM login_audit_logs
        LEFT JOIN stations ON stations.id = login_audit_logs.station_id
        {where}
        ORDER BY login_audit_logs.created_at DESC
        LIMIT ? OFFSET ?
        """,
        params + [per_page, offset],
    ).fetchall()
    conn.close()
    return render_template(
        "admin_login_logs.html",
        logs=logs,
        event_type=event_type,
        username=username,
        pagination=build_pagination(page, per_page, total_count),
    )


@app.route("/admin/users")
@login_required
@permission_required("manage")
def admin_users():
    search = request.args.get("search", "").strip()
    role = request.args.get("role", "").strip()
    where = "WHERE 1 = 1"
    params = []
    if search:
        where += """
            AND (
                users.username LIKE ?
                OR users.display_name LIKE ?
                OR users.email LIKE ?
                OR stations.station_code LIKE ?
                OR stations.station_name LIKE ?
                OR EXISTS (
                    SELECT 1
                    FROM user_station_assignments
                    JOIN stations AS assigned_search_station
                      ON assigned_search_station.id = user_station_assignments.station_id
                    WHERE user_station_assignments.user_id = users.id
                      AND (
                        assigned_search_station.station_code LIKE ?
                        OR assigned_search_station.station_name LIKE ?
                      )
                )
            )
        """
        term = f"%{search}%"
        params.extend([term, term, term, term, term, term, term])
    if role:
        if role == "duty_officer":
            where += """
                AND (
                    users.role = ?
                    OR primary_effective_station.station_code = 'QVP'
                    OR EXISTS (
                        SELECT 1
                        FROM user_station_assignments
                        JOIN stations AS assigned_effective_filter_station
                          ON assigned_effective_filter_station.id = user_station_assignments.station_id
                        WHERE user_station_assignments.user_id = users.id
                          AND assigned_effective_filter_station.station_code = 'QVP'
                    )
                )
            """
            params.append(role)
        else:
            where += " AND users.role = ?"
            params.append(role)

    conn = get_db()
    users = conn.execute(
        f"""
        SELECT users.id, users.username, users.display_name, users.email,
               users.role, users.is_active, users.created_at,
               users.must_change_password, users.password_changed_at, users.last_login_at,
               (
                   SELECT COUNT(*)
                   FROM user_station_assignments
                   WHERE user_station_assignments.user_id = users.id
               ) AS assigned_station_count,
               CASE
                   WHEN users.role = 'duty_officer'
                     OR EXISTS (
                        SELECT 1
                        FROM user_station_assignments
                        JOIN stations AS assigned_effective_station
                          ON assigned_effective_station.id = user_station_assignments.station_id
                        WHERE user_station_assignments.user_id = users.id
                          AND assigned_effective_station.station_code = 'QVP'
                     )
                     OR primary_effective_station.station_code = 'QVP'
                   THEN 'duty_officer'
                   ELSE users.role
               END AS effective_role,
               stations.station_code, stations.station_name,
               stations.is_one_manned, stations.station_status
        FROM users
        LEFT JOIN stations ON stations.id = users.station_id
        LEFT JOIN stations AS primary_effective_station
            ON primary_effective_station.id = users.station_id
        {where}
        ORDER BY users.role, users.username
        """,
        params,
    ).fetchall()
    conn.close()
    return render_template(
        "admin_users.html",
        users=users,
        search=search,
        role=role,
        roles=USER_ROLES,
    )


@app.route("/admin/users/new", methods=["GET", "POST"])
@login_required
@permission_required("manage")
def admin_user_new():
    conn = get_db()
    stations = active_stations(conn)
    if request.method == "POST":
        errors = validate_user_form(request.form, require_password=True)
        username = request.form.get("username", "").strip()
        existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
        if existing:
            errors.append("Username already exists.")

        role = request.form["role"]
        station_id = resolve_station_assignment(conn, request.form, errors)
        assigned_station_ids = form_station_assignment_ids(request.form, station_id)
        if errors:
            for error in errors:
                flash(error)
            conn.close()
            return render_template(
                "admin_user_form.html",
                user=None,
                stations=stations,
                roles=USER_ROLES,
                station_clusters=STATION_CLUSTERS,
                mode="new",
                assigned_station_ids=assigned_station_ids,
            )

        cursor = conn.execute(
            """
            INSERT INTO users (
                username, password_hash, display_name, email, role, station_id,
                is_active, must_change_password, password_reset_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, CURRENT_TIMESTAMP)
            """,
            (
                username,
                generate_password_hash(request.form["password"]),
                request.form["display_name"].strip(),
                request.form.get("email", "").strip(),
                role,
                station_id,
                1 if request.form.get("is_active") == "1" else 0,
            ),
        )
        sync_user_station_assignments(conn, cursor.lastrowid, role, station_id, assigned_station_ids)
        reconcile_required_submissions(conn)
        conn.commit()
        conn.close()
        flash("User created successfully.")
        return redirect("/admin/users")

    conn.close()
    return render_template(
        "admin_user_form.html",
        user=None,
        stations=stations,
        roles=USER_ROLES,
        station_clusters=STATION_CLUSTERS,
        mode="new",
        assigned_station_ids=[],
    )


@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
@permission_required("manage")
def admin_user_edit(user_id):
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        conn.close()
        flash("User not found.")
        return redirect("/admin/users")
    stations = active_stations(conn)
    assigned_station_ids = [
        str(row["station_id"])
        for row in conn.execute(
            "SELECT station_id FROM user_station_assignments WHERE user_id = ?",
            (user_id,),
        ).fetchall()
    ]

    if request.method == "POST":
        errors = validate_user_form(request.form, require_password=False)
        username = request.form.get("username", "").strip()
        existing = conn.execute(
            "SELECT id FROM users WHERE username = ? AND id != ?",
            (username, user_id),
        ).fetchone()
        if existing:
            errors.append("Username already exists.")
        password = request.form.get("password", "")
        if password and len(password) < 8:
            errors.append("New password must be at least 8 characters.")

        role = request.form["role"]
        station_id = resolve_station_assignment(conn, request.form, errors)
        assigned_station_ids = form_station_assignment_ids(request.form, station_id)
        if errors:
            for error in errors:
                flash(error)
            conn.close()
            return render_template(
                "admin_user_form.html",
                user=user,
                stations=stations,
                roles=USER_ROLES,
                station_clusters=STATION_CLUSTERS,
                mode="edit",
                assigned_station_ids=assigned_station_ids,
            )

        is_active = 1 if request.form.get("is_active") == "1" else 0
        if current_user()["id"] == user_id and not is_active:
            conn.close()
            flash("You cannot deactivate your own account.")
            return redirect(f"/admin/users/{user_id}/edit")

        if password:
            must_change_password = 0 if current_user()["id"] == user_id else 1
            conn.execute(
                """
                UPDATE users
                SET username = ?, password_hash = ?, display_name = ?, email = ?,
                    role = ?, station_id = ?, is_active = ?,
                    must_change_password = ?, password_reset_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    username,
                    generate_password_hash(password),
                    request.form["display_name"].strip(),
                    request.form.get("email", "").strip(),
                    role,
                    station_id,
                    is_active,
                    must_change_password,
                    user_id,
                ),
            )
        else:
            conn.execute(
                """
                UPDATE users
                SET username = ?, display_name = ?, email = ?,
                    role = ?, station_id = ?, is_active = ?
                WHERE id = ?
                """,
                (
                    username,
                    request.form["display_name"].strip(),
                    request.form.get("email", "").strip(),
                    role,
                    station_id,
                    is_active,
                    user_id,
                ),
            )
        reconcile_required_submissions(conn)
        sync_user_station_assignments(conn, user_id, role, station_id, assigned_station_ids)
        conn.commit()
        conn.close()
        flash("User updated successfully.")
        return redirect("/admin/users")

    conn.close()
    return render_template(
        "admin_user_form.html",
        user=user,
        stations=stations,
        roles=USER_ROLES,
        station_clusters=STATION_CLUSTERS,
        mode="edit",
        assigned_station_ids=assigned_station_ids,
    )


@app.route("/admin/users/<int:user_id>/toggle-active", methods=["POST"])
@login_required
@permission_required("manage")
def admin_user_toggle_active(user_id):
    user = current_user()
    if user and user["id"] == user_id:
        flash("You cannot deactivate your own account.")
        return redirect("/admin/users")

    conn = get_db()
    target_user = conn.execute("SELECT id, is_active FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target_user:
        conn.close()
        flash("User not found.")
        return redirect("/admin/users")
    new_status = 0 if target_user["is_active"] else 1
    conn.execute("UPDATE users SET is_active = ? WHERE id = ?", (new_status, user_id))
    conn.commit()
    conn.close()
    flash("User status updated.")
    return redirect("/admin/users")


@app.route("/admin/event-rules")
@login_required
@permission_required("manage")
def admin_event_rules():
    current_year = utc_now().astimezone(APP_LOCAL_TZ).year
    conn = get_db()
    rules = conn.execute(
        """
        SELECT *
        FROM province_region_rules
        ORDER BY is_active DESC, priority ASC, province_name ASC
        """
    ).fetchall()
    conn.close()
    month_options = [{"value": "", "label": "Whole Year"}] + [
        {"value": str(month), "label": MONTH_NAMES[month - 1]}
        for month in range(1, 13)
    ]
    return render_template(
        "admin_event_rules.html",
        rules=rules,
        region_choices=REGION_CHOICES,
        year_options=sorted({2026, current_year}, reverse=True),
        month_options=month_options,
    )


def event_rule_form_values(form):
    province_name = (form.get("province_name") or "").strip()
    region_code = (form.get("region_code") or "").strip().upper()
    try:
        priority = int(form.get("priority") or 100)
    except (TypeError, ValueError):
        priority = 100
    priority = max(1, min(priority, 9999))
    notes = (form.get("notes") or "").strip()
    is_active = 1 if form.get("is_active") == "1" else 0
    errors = []
    if not province_name:
        errors.append("Province or location keyword is required.")
    if region_code not in {code for code, _ in REGION_CHOICES}:
        errors.append("Select a valid region.")
    return {
        "province_name": province_name,
        "region_code": region_code,
        "priority": priority,
        "notes": notes,
        "is_active": is_active,
        "errors": errors,
    }


@app.route("/admin/event-rules", methods=["POST"])
@login_required
@permission_required("manage")
def admin_event_rule_create():
    values = event_rule_form_values(request.form)
    if values["errors"]:
        for error in values["errors"]:
            flash(error)
        return redirect("/admin/event-rules")
    conn = get_db()
    existing = conn.execute(
        "SELECT id FROM province_region_rules WHERE LOWER(province_name) = LOWER(?)",
        (values["province_name"],),
    ).fetchone()
    if existing:
        conn.close()
        flash(f"Rule for {values['province_name']} already exists. Edit the existing rule instead.")
        return redirect("/admin/event-rules")
    conn.execute(
        """
        INSERT INTO province_region_rules (
            province_name, region_code, priority, is_active, notes, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            values["province_name"],
            values["region_code"],
            values["priority"],
            values["is_active"],
            values["notes"],
            to_utc_iso(utc_now()),
        ),
    )
    conn.commit()
    conn.close()
    flash("Event region rule created.")
    return redirect("/admin/event-rules")


@app.route("/admin/event-rules/<int:rule_id>", methods=["POST"])
@login_required
@permission_required("manage")
def admin_event_rule_update(rule_id):
    values = event_rule_form_values(request.form)
    if values["errors"]:
        for error in values["errors"]:
            flash(error)
        return redirect("/admin/event-rules")
    conn = get_db()
    duplicate = conn.execute(
        """
        SELECT id
        FROM province_region_rules
        WHERE LOWER(province_name) = LOWER(?)
          AND id != ?
        """,
        (values["province_name"], rule_id),
    ).fetchone()
    if duplicate:
        conn.close()
        flash(f"Another rule already uses {values['province_name']}.")
        return redirect("/admin/event-rules")
    conn.execute(
        """
        UPDATE province_region_rules
        SET province_name = ?,
            region_code = ?,
            priority = ?,
            is_active = ?,
            notes = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (
            values["province_name"],
            values["region_code"],
            values["priority"],
            values["is_active"],
            values["notes"],
            to_utc_iso(utc_now()),
            rule_id,
        ),
    )
    conn.commit()
    conn.close()
    flash("Event region rule updated.")
    return redirect("/admin/event-rules")


@app.route("/admin/event-rules/<int:rule_id>/delete", methods=["POST"])
@login_required
@permission_required("manage")
def admin_event_rule_delete(rule_id):
    conn = get_db()
    conn.execute("DELETE FROM province_region_rules WHERE id = ?", (rule_id,))
    conn.commit()
    conn.close()
    flash("Event region rule deleted.")
    return redirect("/admin/event-rules")


@app.route("/admin/event-rules/recompute", methods=["POST"])
@login_required
@permission_required("manage")
def admin_event_rules_recompute():
    current_year = utc_now().astimezone(APP_LOCAL_TZ).year
    year = parse_year_value(request.form.get("year"), current_year) or current_year
    month_value = request.form.get("month") or ""
    try:
        month = int(month_value) if month_value else None
    except (TypeError, ValueError):
        month = None
    if month is not None and month not in range(1, 13):
        month = None
    conn = get_db()
    try:
        result = recompute_event_regions(conn, year=year, month=month)
        conn.commit()
    except Exception as error:
        conn.rollback()
        conn.close()
        flash(f"Event region recompute failed: {error}")
        return redirect("/admin/event-rules")
    conn.close()
    period = f"{MONTH_NAMES[month - 1]} {year}" if month else str(year)
    flash(f"Event regions recomputed for {period}: {result['changed']} changed out of {result['checked']} checked.")
    return redirect("/admin/event-rules")


def validate_station_form(form, station_id=None):
    errors = []
    station_code = (form.get("station_code") or "").strip().upper()
    station_name = (form.get("station_name") or "").strip()
    cluster_name = form.get("cluster_name") or ""
    station_status = resolve_station_status_label(form)
    station_type = (form.get("station_type") or "SCSS").strip()
    if not station_code:
        errors.append("Station code is required.")
    if not station_name:
        errors.append("Station name is required.")
    if cluster_name not in STATION_CLUSTERS:
        errors.append("Station cluster is required.")
    if not station_status:
        errors.append("Station status is required.")
    if station_type not in {"SCSS", "STSS"}:
        errors.append("Station type must be SCSS or STSS.")
    return errors


def station_type_from_form(form):
    station_type = (form.get("station_type") or "SCSS").strip()
    return station_type if station_type in {"SCSS", "STSS"} else "SCSS"


def include_in_pqr_from_form(form):
    return 1 if form.get("include_in_pqr_compliance") == "1" else 0


@app.route("/admin/stations")
@login_required
@permission_required("manage")
def admin_stations():
    search = request.args.get("search", "").strip()
    cluster = request.args.get("cluster", "").strip()
    status = request.args.get("status", "").strip()
    conn = get_db()
    where = ["1 = 1"]
    params = []
    if search:
        where.append("(station_code LIKE ? OR station_name LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])
    if cluster:
        where.append("cluster_name = ?")
        params.append(cluster)
    if status == "active":
        where.append("is_active = 1")
    elif status == "inactive":
        where.append("is_active = 0")
    stations = conn.execute(
        f"""
        SELECT stations.*,
               COUNT(DISTINCT user_station_assignments.user_id) AS assigned_user_count
        FROM stations
        LEFT JOIN user_station_assignments
            ON user_station_assignments.station_id = stations.id
        WHERE {" AND ".join(where)}
        GROUP BY stations.id
        ORDER BY stations.is_active DESC, stations.station_code
        """,
        params,
    ).fetchall()
    conn.close()
    return render_template(
        "admin_stations.html",
        stations=stations,
        station_clusters=STATION_CLUSTERS,
        search=search,
        cluster=cluster,
        status=status,
    )


@app.route("/admin/stations/new", methods=["GET", "POST"])
@login_required
@permission_required("manage")
def admin_station_new():
    conn = get_db()
    station = None
    if request.method == "POST":
        errors = validate_station_form(request.form)
        station_code = (request.form.get("station_code") or "").strip().upper()
        existing = conn.execute("SELECT id FROM stations WHERE station_code = ?", (station_code,)).fetchone()
        if existing:
            errors.append(f"Station code {station_code} already exists.")
        cluster_name = request.form.get("cluster_name") or ""
        region_code = STATION_REGION_BY_CLUSTER.get(cluster_name, "")
        station_status_label = resolve_station_status_label(request.form)
        station = {
            "station_code": station_code,
            "station_name": (request.form.get("station_name") or "").strip(),
            "cluster_name": cluster_name,
            "region_code": region_code,
            "station_type": station_type_from_form(request.form),
            "include_in_pqr_compliance": include_in_pqr_from_form(request.form),
            "station_status": station_status_label,
            "is_one_manned": 1 if station_status_label == "1M" else 0,
            "is_active": 1 if request.form.get("is_active") == "1" else 0,
        }
        if errors:
            for error in errors:
                flash(error)
            conn.close()
            return render_template("admin_station_form.html", station=station, station_clusters=STATION_CLUSTERS, mode="new")
        conn.execute(
            """
            INSERT INTO stations (
                station_code, station_name, region_code, cluster_name,
                station_type, include_in_pqr_compliance,
                is_one_manned, station_status, is_active
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                station["station_code"],
                station["station_name"],
                station["region_code"],
                station["cluster_name"],
                station["station_type"],
                station["include_in_pqr_compliance"],
                station["is_one_manned"],
                station["station_status"],
                station["is_active"],
            ),
        )
        reconcile_required_submissions(conn)
        conn.commit()
        conn.close()
        flash("Station created successfully.")
        return redirect("/admin/stations")
    conn.close()
    return render_template("admin_station_form.html", station=station, station_clusters=STATION_CLUSTERS, mode="new")


@app.route("/admin/stations/<int:station_id>/edit", methods=["GET", "POST"])
@login_required
@permission_required("manage")
def admin_station_edit(station_id):
    conn = get_db()
    station = conn.execute("SELECT * FROM stations WHERE id = ?", (station_id,)).fetchone()
    if not station:
        conn.close()
        flash("Station not found.")
        return redirect("/admin/stations")
    if request.method == "POST":
        errors = validate_station_form(request.form, station_id)
        station_code = (request.form.get("station_code") or "").strip().upper()
        existing = conn.execute(
            "SELECT id FROM stations WHERE station_code = ? AND id != ?",
            (station_code, station_id),
        ).fetchone()
        if existing:
            errors.append(f"Station code {station_code} already exists.")
        cluster_name = request.form.get("cluster_name") or ""
        region_code = STATION_REGION_BY_CLUSTER.get(cluster_name, "")
        station_status_label = resolve_station_status_label(request.form)
        station = {
            **dict(station),
            "station_code": station_code,
            "station_name": (request.form.get("station_name") or "").strip(),
            "cluster_name": cluster_name,
            "region_code": region_code,
            "station_type": station_type_from_form(request.form),
            "include_in_pqr_compliance": include_in_pqr_from_form(request.form),
            "station_status": station_status_label,
            "is_one_manned": 1 if station_status_label == "1M" else 0,
            "is_active": 1 if request.form.get("is_active") == "1" else 0,
        }
        if errors:
            for error in errors:
                flash(error)
            conn.close()
            return render_template("admin_station_form.html", station=station, station_clusters=STATION_CLUSTERS, mode="edit")
        conn.execute(
            """
            UPDATE stations
            SET station_code = ?, station_name = ?, region_code = ?, cluster_name = ?,
                station_type = ?, include_in_pqr_compliance = ?,
                is_one_manned = ?, station_status = ?, is_active = ?
            WHERE id = ?
            """,
            (
                station["station_code"],
                station["station_name"],
                station["region_code"],
                station["cluster_name"],
                station["station_type"],
                station["include_in_pqr_compliance"],
                station["is_one_manned"],
                station["station_status"],
                station["is_active"],
                station_id,
            ),
        )
        reconcile_required_submissions(conn)
        conn.commit()
        conn.close()
        flash("Station updated successfully.")
        return redirect("/admin/stations")
    conn.close()
    return render_template("admin_station_form.html", station=station, station_clusters=STATION_CLUSTERS, mode="edit")


@app.route("/admin/stations/<int:station_id>/toggle-active", methods=["POST"])
@login_required
@permission_required("manage")
def admin_station_toggle_active(station_id):
    conn = get_db()
    station = conn.execute("SELECT id, is_active FROM stations WHERE id = ?", (station_id,)).fetchone()
    if not station:
        conn.close()
        flash("Station not found.")
        return redirect("/admin/stations")
    conn.execute("UPDATE stations SET is_active = ? WHERE id = ?", (0 if station["is_active"] else 1, station_id))
    reconcile_required_submissions(conn)
    conn.commit()
    conn.close()
    flash("Station status updated.")
    return redirect("/admin/stations")


def build_station_submission_status(conn, month_value="", year_value=""):
    station_rows = conn.execute(
        """
        SELECT id, station_code, station_name, region_code, cluster_name,
               station_type, include_in_pqr_compliance, is_one_manned
        FROM stations
        WHERE stations.is_active = 1
          AND stations.include_in_pqr_compliance = 1
        ORDER BY station_code
        """,
    ).fetchall()
    event_rows = conn.execute(
        """
        SELECT id, region_code, event_datetime_utc, magnitude, is_felt
        FROM earthquake_events
        WHERE COALESCE(exclude_from_pqr_rating, 0) = 0
        """
    ).fetchall()
    report_rows = conn.execute(
        """
        SELECT pqr_reports.station_id, pqr_reports.event_id,
               pqr_reports.observed_intensities,
               pqr_reports.instrumental_intensities
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        """
    ).fetchall()

    date_options = build_year_month_options(event_rows, month_value, year_value)
    selected_year = date_options["selected_year"]
    selected_month = date_options["selected_month"]
    month_options = date_options["month_options"]
    year_options = date_options["year_options"]

    station_rows_by_cluster = {cluster_name: [] for cluster_name in STATION_CLUSTERS}
    for row in station_rows:
        station = dict(row)
        cluster_name = cluster_name_for_station(station) or "Unassigned Cluster"
        station_rows_by_cluster.setdefault(cluster_name, []).append(station)
    for cluster_name in station_rows_by_cluster:
        station_rows_by_cluster[cluster_name].sort(key=lambda item: item["station_code"])
    selected_event_ids = {
        row["id"]
        for row in event_rows
        if event_local_month(row["event_datetime_utc"]) == selected_month
    }
    report_event_ids_by_station = {}
    felt_event_ids = {
        row["id"]
        for row in event_rows
        if row["is_felt"]
    }
    for row in report_rows:
        if row["event_id"] not in selected_event_ids:
            continue
        report_event_ids_by_station.setdefault(row["station_id"], set()).add(row["event_id"])

    active_events = [
        dict(row)
        for row in event_rows
        if row["id"] in selected_event_ids
    ]
    station_groups = []
    cluster_rankings = []
    below_threshold = []
    total_required = 0
    total_submitted = 0
    total_felt_required = 0
    total_felt_submitted = 0
    best_cluster = {"name": "No Data", "compliance": 0}

    for cluster_name, cluster_stations in station_rows_by_cluster.items():
        if not cluster_stations:
            continue
        cluster_station_rows = []
        visible_regions = set(CLUSTER_EVENT_REGIONS.get(cluster_name, ()))
        eligible_event_ids = {
            event["id"]
            for event in active_events
            if event["region_code"] in visible_regions or magnitude_triggers_all_stations(event["magnitude"])
        }
        eligible_felt_event_ids = eligible_event_ids & felt_event_ids
        required_per_station = len(eligible_event_ids)
        felt_required_per_station = len(eligible_felt_event_ids)
        cluster_required = 0
        cluster_submitted = 0
        cluster_felt_required = 0
        cluster_felt_submitted = 0
        for row in cluster_stations:
            station_code = row.get("station_code")
            station_id = row.get("id")
            station_visible_event_ids = {
                event["id"]
                for event in active_events
                if event["id"] in eligible_event_ids
            }
            submitted_visible_event_ids = (
                report_event_ids_by_station.get(station_id, set()) & station_visible_event_ids
            )
            missed_one_manned_event_ids = {
                event["id"]
                for event in active_events
                if (
                    event["id"] in station_visible_event_ids
                    and event["id"] not in submitted_visible_event_ids
                    and station_exempt_from_pqr(row, event)
                )
            }
            station_eligible_event_ids = station_visible_event_ids - missed_one_manned_event_ids
            station_eligible_felt_event_ids = station_eligible_event_ids & eligible_felt_event_ids
            submitted_event_ids = submitted_visible_event_ids & station_eligible_event_ids
            felt_submitted_event_ids = submitted_visible_event_ids & station_eligible_felt_event_ids
            required = len(station_eligible_event_ids)
            submitted = len(submitted_event_ids)
            felt_required = len(station_eligible_felt_event_ids)
            felt_submitted = len(felt_submitted_event_ids)
            compliance = round((submitted / required) * 100, 2) if required else 0
            felt_compliance = round((felt_submitted / felt_required) * 100, 2) if felt_required else 0
            station_row = {
                "code": station_code,
                "region": row.get("region_code") or STATION_REGION_BY_CLUSTER.get(cluster_name, ""),
                "required": required,
                "submitted": submitted,
                "pending": max(required - submitted, 0),
                "compliance": compliance,
                "felt_required": felt_required,
                "felt_submitted": felt_submitted,
                "felt_compliance": felt_compliance,
            }
            cluster_station_rows.append(station_row)
            cluster_required += required
            cluster_submitted += submitted
            cluster_felt_required += felt_required
            cluster_felt_submitted += felt_submitted
            if required and compliance < 80:
                below_threshold.append(station_row)

        cluster_compliance = round((cluster_submitted / cluster_required) * 100, 2) if cluster_required else 0
        felt_cluster_compliance = (
            round((cluster_felt_submitted / cluster_felt_required) * 100, 2)
            if cluster_felt_required
            else 0
        )
        station_groups.append(
            {
                "name": cluster_name.replace(" Cluster", ""),
                "station_count": len(cluster_stations),
                "stations": cluster_station_rows,
                "required": cluster_required,
                "submitted": cluster_submitted,
                "compliance": cluster_compliance,
                "felt_required": cluster_felt_required,
                "felt_submitted": cluster_felt_submitted,
                "felt_compliance": felt_cluster_compliance,
            }
        )
        cluster_rankings.append(
            {"name": cluster_name.replace(" Cluster", ""), "compliance": cluster_compliance}
        )
        total_required += cluster_required
        total_submitted += cluster_submitted
        total_felt_required += cluster_felt_required
        total_felt_submitted += cluster_felt_submitted
        if cluster_compliance >= best_cluster["compliance"]:
            best_cluster = {"name": cluster_name.replace(" Cluster", ""), "compliance": cluster_compliance}

    overall_compliance = round((total_submitted / total_required) * 100, 2) if total_required else 0
    felt_overall_compliance = (
        round((total_felt_submitted / total_felt_required) * 100, 2)
        if total_felt_required
        else 0
    )
    below_threshold = sorted(below_threshold, key=lambda item: item["compliance"])[:8]
    attention_label = (
        "All clusters on target"
        if not below_threshold
        else ", ".join(sorted({row["region"] for row in below_threshold})[:2])
    )
    return {
        "station_groups": station_groups,
        "cluster_rankings": sorted(cluster_rankings, key=lambda item: item["compliance"], reverse=True),
        "below_threshold": below_threshold,
        "month_options": month_options,
        "year_options": year_options,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "selected_month_label": month_display_label(selected_month),
        "overall_compliance": overall_compliance,
        "felt_overall_compliance": felt_overall_compliance,
        "best_cluster": best_cluster,
        "attention_label": attention_label,
        "total_required": total_required,
        "total_submitted": total_submitted,
    }


def response_duration_label(seconds):
    if seconds is None:
        return "-"
    seconds = max(int(seconds), 0)
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    if hours:
        return f"{hours}h {minutes:02d}m"
    return f"{minutes}m"


def median_seconds(values):
    if not values:
        return None
    ordered = sorted(values)
    midpoint = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[midpoint]
    return (ordered[midpoint - 1] + ordered[midpoint]) / 2


def response_speed_label(within_3h_rate, median_value, pending_count):
    if pending_count:
        return "Needs Follow-up"
    if within_3h_rate >= 90 and median_value is not None and median_value <= 3 * 3600:
        return "Excellent"
    if within_3h_rate >= 75 and median_value is not None and median_value <= 6 * 3600:
        return "Fast"
    if within_3h_rate >= 50:
        return "Acceptable"
    return "Needs Follow-up"


def build_station_response_speed(conn, selected_month):
    rows = conn.execute(
        """
        SELECT stations.id AS station_id, stations.station_code, stations.station_name,
               earthquake_events.event_datetime_utc,
               pqr_reports.submitted_at
        FROM pqr_required_submissions
        JOIN earthquake_events ON earthquake_events.id = pqr_required_submissions.event_id
        JOIN stations ON stations.id = pqr_required_submissions.station_id
        LEFT JOIN pqr_reports
          ON pqr_reports.event_id = pqr_required_submissions.event_id
         AND pqr_reports.station_id = pqr_required_submissions.station_id
        WHERE stations.is_active = 1
          AND stations.include_in_pqr_compliance = 1
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
        ORDER BY stations.station_code, earthquake_events.event_datetime_utc
        """
    ).fetchall()
    station_stats = {}
    network_durations = []
    for row in rows:
        if event_local_month(row["event_datetime_utc"]) != selected_month:
            continue
        stat = station_stats.setdefault(
            row["station_id"],
            {
                "station_id": row["station_id"],
                "station_code": row["station_code"],
                "station_name": row["station_name"],
                "required_count": 0,
                "submitted_count": 0,
                "pending_count": 0,
                "within_3h_count": 0,
                "within_6h_count": 0,
                "within_22h_count": 0,
                "durations": [],
            },
        )
        stat["required_count"] += 1
        event_time = parse_utc_iso(row["event_datetime_utc"])
        submitted_at = parse_utc_iso(row["submitted_at"])
        if not event_time or not submitted_at:
            stat["pending_count"] += 1
            continue
        duration = max((submitted_at - event_time).total_seconds(), 0)
        stat["submitted_count"] += 1
        stat["durations"].append(duration)
        network_durations.append(duration)
        if duration <= 3 * 3600:
            stat["within_3h_count"] += 1
        if duration <= 6 * 3600:
            stat["within_6h_count"] += 1
        if duration <= 22 * 3600:
            stat["within_22h_count"] += 1

    rankings = []
    for stat in station_stats.values():
        avg_value = sum(stat["durations"]) / len(stat["durations"]) if stat["durations"] else None
        median_value = median_seconds(stat["durations"])
        within_3h_rate = round((stat["within_3h_count"] / stat["required_count"]) * 100, 1)
        within_6h_rate = round((stat["within_6h_count"] / stat["required_count"]) * 100, 1)
        missed_rate = round((stat["pending_count"] / stat["required_count"]) * 100, 1)
        consistency_score = round(
            (within_3h_rate * 0.7)
            + ((100 - missed_rate) * 0.2)
            + (within_6h_rate * 0.1),
            1,
        )
        rankings.append(
            {
                **stat,
                "avg_seconds": avg_value,
                "median_seconds": median_value,
                "avg_label": response_duration_label(avg_value),
                "median_label": response_duration_label(median_value),
                "fastest_label": response_duration_label(min(stat["durations"]) if stat["durations"] else None),
                "within_3h_rate": within_3h_rate,
                "within_6h_rate": within_6h_rate,
                "within_22h_rate": round((stat["within_22h_count"] / stat["required_count"]) * 100, 1),
                "consistency_score": consistency_score,
                "speed_label": response_speed_label(within_3h_rate, median_value, stat["pending_count"]),
            }
        )

    rankings.sort(
        key=lambda item: (
            item["median_seconds"] is None,
            item["median_seconds"] if item["median_seconds"] is not None else 999999999,
            -item["within_3h_rate"],
            item["pending_count"],
            item["station_code"],
        )
    )
    total_required = sum(item["required_count"] for item in rankings)
    within_3h_count = sum(item["within_3h_count"] for item in rankings)
    return {
        "rankings": rankings,
        "top_stations": rankings[:10],
        "fastest_station": rankings[0] if rankings else None,
        "network_median_label": response_duration_label(median_seconds(network_durations)),
        "within_3h_count": within_3h_count,
        "total_required": total_required,
        "total_submitted": sum(item["submitted_count"] for item in rankings),
        "within_3h_rate": round((within_3h_count / total_required) * 100, 1) if total_required else 0,
    }


def dashboard_activity_time_label(value):
    parsed = parse_utc_iso(value)
    if not parsed:
        return ""
    local_time = parsed.astimezone(APP_LOCAL_TZ)
    now_local = utc_now().astimezone(APP_LOCAL_TZ)
    if local_time.date() == now_local.date():
        return local_time.strftime("%I:%M %p")
    return local_time.strftime("%b %d")


def build_dashboard_activity(summary, recent_activity, sheet_sync_counts):
    items = []
    for row in recent_activity[:5]:
        items.append(
            {
                "kind": "success",
                "title": "Station submitted PQR",
                "body": f"{row['station_code']} submitted PQR for {row['event_key']}.",
                "time": dashboard_activity_time_label(row["submitted_at"]),
                "url": f"/pqr/list?event_key={row['event_key']}&station={row['station_code']}",
            }
        )
    for row in summary:
        pending_count = row["pending_count"] or 0
        if pending_count <= 0:
            continue
        items.append(
            {
                "kind": "warning" if pending_count < 5 else "danger",
                "title": "Pending station submissions",
                "body": f"{pending_count} station(s) still pending for {row['event_key']}.",
                "time": dashboard_activity_time_label(row["event_datetime_utc"]),
                "url": "/stations/status",
            }
        )
    failed_sync = sheet_sync_counts.get("failed", 0) if sheet_sync_counts else 0
    if failed_sync:
        items.insert(
            0,
            {
                "kind": "danger",
                "title": "Google Sheets sync failed",
                "body": f"{failed_sync} report(s) need sync attention.",
                "time": "Now",
                "url": "/reports",
            },
        )
    priority = {"danger": 0, "warning": 1, "success": 2, "info": 3}
    return sorted(items, key=lambda item: (priority.get(item["kind"], 9), item["time"]))[:8]


REMARK_STATUS_CODES = {
    "With Phase Reading": "WPR",
    "No Quake Record": "NQR",
    "No Data": "ND",
    "No Operation": "NO",
    "Intensities": "INT",
    "No Staff": "NS",
    "On Meeting or Lecture or Other Official Business": "OB",
    "On Fieldwork": "FW",
    "On Leave": "OL",
}


def remark_status_code(remarks):
    return REMARK_STATUS_CODES.get((remarks or "").strip(), (remarks or "").strip())


EVENT_QUERY_CATEGORIES = (
    "No Quake Record",
    "No Operation",
    "On Fieldwork",
    "No Report",
    "No Data",
    "No Staff",
    "On Leave",
    "On Meeting or Lecture or Other Official Business",
    "With Intensity Report",
    "With Phase Reading",
)
ACTIONABLE_PQR_REMARKS = ("With Phase Reading", "Intensities")


def station_report_category(row):
    if not row.get("report_id"):
        return "No Report"
    remarks = (row.get("remarks") or "").strip()
    if (
        (row.get("observed_intensities") or "").strip()
        or (row.get("instrumental_intensities") or "").strip()
        or remarks == "Intensities"
    ):
        return "With Intensity Report"
    return remarks or "No Report"


def station_report_has_intensity(row):
    return bool(
        (row.get("observed_intensities") or "").strip()
        or (row.get("instrumental_intensities") or "").strip()
    )


def event_query_arrival_display(value):
    text = str(value or "").strip()
    return text[2:] if len(text) > 2 else text


def event_query_duration_display(value):
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return str(int(float(text)))
    except ValueError:
        return text.split(".", 1)[0] if "." in text else text


def event_query_status_rows(conn, event):
    rows = conn.execute(
        """
        SELECT stations.id AS station_id, stations.station_code, stations.station_name,
               stations.region_code, stations.cluster_name, stations.is_one_manned, stations.station_status,
               pqr_required_submissions.status AS required_status,
               pqr_reports.id AS report_id, pqr_reports.p_arrival, pqr_reports.s_arrival,
               pqr_reports.duration, pqr_reports.remarks,
               pqr_reports.observed_intensities, pqr_reports.instrumental_intensities,
               pqr_reports.verified_areas_without_intensities, pqr_reports.submitted_at,
               pqr_reports.updated_at
        FROM pqr_required_submissions
        JOIN stations ON stations.id = pqr_required_submissions.station_id
        LEFT JOIN pqr_reports
            ON pqr_reports.event_id = pqr_required_submissions.event_id
           AND pqr_reports.station_id = pqr_required_submissions.station_id
        WHERE pqr_required_submissions.event_id = ?
        ORDER BY stations.station_code
        """,
        (event["id"],),
    ).fetchall()
    if not rows:
        fallback_stations = stations_required_for_event(
            conn,
            event["region_code"],
            event["magnitude"],
            event["event_datetime_utc"],
        )
        return [
            {
                **dict(station),
                "station_id": station["id"],
                "required_status": "pending",
                "report_id": None,
                "p_arrival": "",
                "s_arrival": "",
                "duration": "",
                "remarks": "",
                "observed_intensities": "",
                "instrumental_intensities": "",
                "verified_areas_without_intensities": "",
                "submitted_at": "",
                "updated_at": "",
            }
            for station in fallback_stations
        ]
    return [dict(row) for row in rows]


def event_query_report_rows(conn, event):
    placeholders = ", ".join("?" for _ in ACTIONABLE_PQR_REMARKS)
    return [
        {
            **dict(row),
            "p_arrival_display": event_query_arrival_display(row["p_arrival"]),
            "s_arrival_display": event_query_arrival_display(row["s_arrival"]),
            "duration_display": event_query_duration_display(row["duration"]),
        }
        for row in conn.execute(
            f"""
            SELECT stations.id AS station_id, stations.station_code, stations.station_name,
                   stations.region_code, stations.cluster_name, stations.is_one_manned, stations.station_status,
                   pqr_reports.id AS report_id, pqr_reports.p_arrival, pqr_reports.s_arrival,
                   pqr_reports.duration, pqr_reports.remarks,
                   pqr_reports.observed_intensities, pqr_reports.instrumental_intensities,
                   pqr_reports.verified_areas_without_intensities, pqr_reports.submitted_at,
                   pqr_reports.updated_at
            FROM pqr_reports
            JOIN stations ON stations.id = pqr_reports.station_id
            WHERE pqr_reports.event_id = ?
              AND pqr_reports.remarks IN ({placeholders})
            ORDER BY stations.station_code
            """,
            (event["id"], *ACTIONABLE_PQR_REMARKS),
        ).fetchall()
    ]


def build_event_query_context(conn, event_key="", station_search="", report_category="", station_cluster="", page=1, per_page=10):
    event_key = (event_key or "").strip()
    recent_cutoff = to_utc_iso(utc_now() - timedelta(hours=24))
    recent_events = [
        {
            **dict(row),
            "event_datetime_pst": utc_iso_to_pst_display(row["event_datetime_utc"]),
        }
        for row in conn.execute(
            """
            SELECT event_key, event_datetime_utc, magnitude, reference_location, region_code
            FROM earthquake_events
            WHERE event_datetime_utc >= ?
            ORDER BY event_datetime_utc DESC
            """,
            (recent_cutoff,),
        ).fetchall()
    ]
    event = None
    if event_key:
        event = conn.execute(
            """
            SELECT *
            FROM earthquake_events
            WHERE event_key = ?
               OR event_key LIKE ?
            ORDER BY event_datetime_utc DESC
            LIMIT 1
            """,
            (event_key, f"{event_key}%"),
        ).fetchone()
    if not event:
        if recent_events:
            event = conn.execute(
                "SELECT * FROM earthquake_events WHERE event_key = ?",
                (recent_events[0]["event_key"],),
            ).fetchone()
        else:
            event = conn.execute(
                """
                SELECT *
                FROM earthquake_events
                ORDER BY event_datetime_utc DESC
                LIMIT 1
                """
            ).fetchone()
    if not event:
        return None

    event = dict(event)
    recent_event_keys = {row["event_key"] for row in recent_events}
    status_rows = []
    for row in event_query_status_rows(conn, event):
        cluster_name = cluster_name_for_station(row["station_code"], row["region_code"])
        item = {
            **row,
            "cluster_name": cluster_name,
            "category": station_report_category(row),
            "has_intensity": station_report_has_intensity(row),
        }
        status_rows.append(item)

    report_rows = []
    for row in event_query_report_rows(conn, event):
        cluster_name = cluster_name_for_station(row["station_code"], row["region_code"])
        report_rows.append(
            {
                **row,
                "cluster_name": cluster_name,
                "category": station_report_category(row),
                "has_intensity": station_report_has_intensity(row),
            }
        )

    station_search = (station_search or "").strip()
    report_category = (report_category or "").strip()
    station_cluster = (station_cluster or "").strip()
    filtered_status_rows = status_rows
    filtered_report_rows = report_rows
    if station_search:
        search_text = station_search.lower()
        filtered_status_rows = [
            row
            for row in filtered_status_rows
            if search_text in (row["station_code"] or "").lower()
            or search_text in (row["station_name"] or "").lower()
        ]
        filtered_report_rows = [
            row
            for row in filtered_report_rows
            if search_text in (row["station_code"] or "").lower()
            or search_text in (row["station_name"] or "").lower()
        ]
    if station_cluster:
        filtered_status_rows = [row for row in filtered_status_rows if row["cluster_name"] == station_cluster]
        filtered_report_rows = [row for row in filtered_report_rows if row["cluster_name"] == station_cluster]
    if report_category:
        filtered_status_rows = [row for row in filtered_status_rows if row["category"] == report_category]
        filtered_report_rows = [row for row in filtered_report_rows if row["category"] == report_category]

    category_rows = []
    for category in EVENT_QUERY_CATEGORIES:
        station_codes = [row["station_code"] for row in filtered_status_rows if row["category"] == category]
        category_rows.append({"category": category, "station_codes": station_codes, "count": len(station_codes)})

    total_count = len(filtered_report_rows)
    page = max(int(page or 1), 1)
    per_page = min(max(int(per_page or 10), 1), 100)
    offset = (page - 1) * per_page
    paged_rows = filtered_report_rows[offset : offset + per_page]

    submitted_count = sum(1 for row in filtered_status_rows if row.get("report_id"))
    intensity_count = sum(1 for row in filtered_status_rows if row["has_intensity"])
    map_url = ""
    if event.get("latitude") is not None and event.get("longitude") is not None:
        map_url = f"https://www.google.com/maps/search/?api=1&query={event['latitude']},{event['longitude']}"

    return {
        "event": {
            **event,
            "event_datetime_pst": utc_iso_to_pst_display(event["event_datetime_utc"]),
            "event_datetime_label": event_local_datetime_label(event["event_datetime_utc"]),
            "status_label": "Reviewing" if event["status"] == "open" else event["status"].title(),
            "map_url": map_url,
        },
        "rows": paged_rows,
        "all_rows": filtered_report_rows,
        "category_rows": category_rows,
        "summary": {
            "total": len(filtered_status_rows),
            "submitted": submitted_count,
            "pending": max(len(filtered_status_rows) - submitted_count, 0),
            "intensity": intensity_count,
        },
        "report_count": len(filtered_report_rows),
        "filters": {
            "event_key": event_key or event["event_key"],
            "recent_event_key": event["event_key"] if event["event_key"] in recent_event_keys else "",
            "manual_event_key": event["event_key"] if event["event_key"] not in recent_event_keys else "",
            "station_search": station_search,
            "report_category": report_category,
            "station_cluster": station_cluster,
        },
        "recent_events": recent_events,
        "report_categories": EVENT_QUERY_CATEGORIES,
        "station_clusters": list(STATION_CLUSTERS.keys()),
        "pagination": build_pagination(page, per_page, total_count),
    }


def update_event_felt_status(conn, event_id):
    event = conn.execute(
        """
        SELECT reported_intensities, instrumental_intensities, felt_override
        FROM earthquake_events
        WHERE id = ?
        """,
        (event_id,),
    ).fetchone()
    if not event:
        return False

    if event["felt_override"] in (0, 1):
        is_felt = int(event["felt_override"])
        source = "manual" if is_felt else "manual_no"
    elif (event["reported_intensities"] or "").strip() or (event["instrumental_intensities"] or "").strip():
        is_felt = 1
        source = "bulletin"
    else:
        has_pqr_intensity = conn.execute(
            """
            SELECT 1
            FROM pqr_reports
            WHERE event_id = ?
              AND (
                TRIM(COALESCE(observed_intensities, '')) != ''
                OR TRIM(COALESCE(instrumental_intensities, '')) != ''
              )
            LIMIT 1
            """,
            (event_id,),
        ).fetchone()
        is_felt = 1 if has_pqr_intensity else 0
        source = "pqr" if has_pqr_intensity else "none"

    conn.execute(
        """
        UPDATE earthquake_events
        SET is_felt = ?,
            felt_source = ?,
            felt_checked_at = ?
        WHERE id = ?
        """,
        (is_felt, source, to_utc_iso(utc_now()), event_id),
    )
    return bool(is_felt)


def refresh_all_event_felt_statuses(conn):
    event_ids = [
        row["id"]
        for row in conn.execute("SELECT id FROM earthquake_events").fetchall()
    ]
    felt_count = 0
    for event_id in event_ids:
        if update_event_felt_status(conn, event_id):
            felt_count += 1
    return {"checked": len(event_ids), "felt": felt_count}


def build_station_event_matrix(conn, selected_month):
    station_rows = conn.execute(
        """
        SELECT id, station_code, region_code, cluster_name,
               station_type, include_in_pqr_compliance, is_one_manned
        FROM stations
        WHERE is_active = 1
          AND include_in_pqr_compliance = 1
        ORDER BY station_code
        """,
    ).fetchall()
    event_rows = conn.execute(
        """
        SELECT id, event_key, event_datetime_utc, magnitude, reference_location, region_code,
               source, source_url, is_felt, exclude_from_pqr_rating, pqr_rating_exclusion_reason
        FROM earthquake_events
        WHERE COALESCE(exclude_from_pqr_rating, 0) = 0
        ORDER BY event_datetime_utc DESC, event_key
        """
    ).fetchall()
    selected_events = [
        dict(row)
        for row in event_rows
        if event_local_month(row["event_datetime_utc"]) == selected_month
    ]
    event_ids = [event["id"] for event in selected_events]

    reports_by_event_station = {}
    if event_ids:
        placeholders = ", ".join("?" for _ in event_ids)
        report_rows = conn.execute(
            f"""
            SELECT event_id, station_id, remarks
            FROM pqr_reports
            WHERE event_id IN ({placeholders})
            """,
            event_ids,
        ).fetchall()
        reports_by_event_station = {
            (row["event_id"], row["station_id"]): remark_status_code(row["remarks"])
            for row in report_rows
        }

    station_groups = []
    station_rows_by_cluster = {cluster_name: [] for cluster_name in STATION_CLUSTERS}
    for row in station_rows:
        station = dict(row)
        cluster_name = cluster_name_for_station(station) or "Unassigned Cluster"
        station_rows_by_cluster.setdefault(cluster_name, []).append(station)
    for cluster_name in station_rows_by_cluster:
        station_rows_by_cluster[cluster_name].sort(key=lambda item: item["station_code"])
    for cluster_name, stations_in_cluster in station_rows_by_cluster.items():
        if not stations_in_cluster:
            continue
        region_code = STATION_REGION_BY_CLUSTER.get(cluster_name, stations_in_cluster[0].get("region_code") or "NL")
        station_groups.append(
            {
                "name": cluster_name.replace(" Cluster", "").upper(),
                "class": region_code.lower(),
                "stations": [
                    {
                        "code": station["station_code"],
                        "id": station["id"],
                        "region_code": station["region_code"],
                        "cluster_name": cluster_name,
                        "is_one_manned": station.get("is_one_manned", 0),
                    }
                    for station in stations_in_cluster
                ],
            }
        )

    matrix_events = []
    for index, event in enumerate(selected_events):
        event["is_import_placeholder"] = is_import_placeholder_event(event)
        event["is_excluded_from_pqr_rating"] = bool(event.get("exclude_from_pqr_rating"))
        event["merge_candidates"] = []
        above_event = selected_events[index - 1] if index > 0 else None
        below_event = selected_events[index + 1] if index + 1 < len(selected_events) else None
        for direction, target_event in (("above", above_event), ("below", below_event)):
            candidate = build_event_merge_candidate(conn, event, target_event, direction, max_minutes=10)
            if candidate:
                event["merge_candidates"].append(candidate)
        event["can_exclude_from_pqr_rating"] = event["is_import_placeholder"]
        cell_groups = []
        for group in station_groups:
            cells = []
            for station in group["stations"]:
                value = reports_by_event_station.get((event["id"], station["id"]))
                if value is None and event_visible_for_station(event, station):
                    if station_exempt_from_pqr(station, event):
                        value = "1M"
                    else:
                        value = "0"
                cells.append({"station": station["code"], "value": value or ""})
            cell_groups.append({"class": group["class"], "cells": cells})
        matrix_events.append(
            {
                **event,
                "event_datetime_label": event_local_datetime_label(event["event_datetime_utc"]),
                "is_felt": bool(event["is_felt"]),
                "cell_groups": cell_groups,
            }
        )

    return {
        "station_matrix_groups": station_groups,
        "station_matrix_events": matrix_events,
        "station_matrix_colspan": 7 + sum(len(group["stations"]) for group in station_groups),
    }


def build_pqr_monitoring_report(conn, period_type=None, year=None, period=None):
    period_info = parse_monitoring_period(period_type, year, period)
    month_keys = [month["key"] for month in period_info["months"]]
    station_rows = conn.execute(
        """
        SELECT id, station_code, station_name, region_code, cluster_name,
               station_type, include_in_pqr_compliance, is_one_manned
        FROM stations
        WHERE is_active = 1
          AND include_in_pqr_compliance = 1
        ORDER BY station_code
        """
    ).fetchall()
    event_rows = conn.execute(
        """
        SELECT id, event_key, event_datetime_utc, magnitude, region_code
        FROM earthquake_events
        WHERE COALESCE(exclude_from_pqr_rating, 0) = 0
        ORDER BY event_datetime_utc
        """
    ).fetchall()
    selected_events = [
        dict(row)
        for row in event_rows
        if event_local_month(row["event_datetime_utc"]) in month_keys
    ]
    event_month_by_id = {
        event["id"]: event_local_month(event["event_datetime_utc"])
        for event in selected_events
    }
    event_ids = [event["id"] for event in selected_events]
    reports_by_station_event = set()
    if event_ids:
        placeholders = ", ".join("?" for _ in event_ids)
        report_rows = conn.execute(
            f"""
            SELECT station_id, event_id
            FROM pqr_reports
            WHERE event_id IN ({placeholders})
            """,
            event_ids,
        ).fetchall()
        reports_by_station_event = {
            (row["station_id"], row["event_id"])
            for row in report_rows
        }

    grouped_stations = {cluster_name: [] for cluster_name in STATION_CLUSTERS}
    for row in station_rows:
        station = dict(row)
        cluster_name = cluster_name_for_station(station) or "Unassigned Cluster"
        grouped_stations.setdefault(cluster_name, []).append(station)
    for stations in grouped_stations.values():
        stations.sort(key=station_cluster_sort_key)

    cluster_reports = []
    month_totals = {month_key: {"required": 0, "submitted": 0} for month_key in month_keys}
    overall_required = 0
    overall_submitted = 0
    below_threshold = []

    for cluster_name, stations_in_cluster in grouped_stations.items():
        if not stations_in_cluster:
            continue
        station_reports = []
        cluster_required = 0
        cluster_submitted = 0
        for station in stations_in_cluster:
            monthly = []
            station_required = 0
            station_submitted = 0
            for month in period_info["months"]:
                month_events = [
                    event
                    for event in selected_events
                    if event_month_by_id.get(event["id"]) == month["key"]
                    and event_visible_for_station(event, station)
                    and not station_exempt_from_pqr(station, event)
                ]
                required = len(month_events)
                submitted = sum(
                    1
                    for event in month_events
                    if (station["id"], event["id"]) in reports_by_station_event
                )
                percent = round((submitted / required) * 100, 2) if required else None
                monthly.append(
                    {
                        "key": month["key"],
                        "required": required,
                        "submitted": submitted,
                        "percent": percent,
                        "display": format_percent(percent),
                        "low": percent is not None and percent < 80,
                    }
                )
                station_required += required
                station_submitted += submitted
                month_totals[month["key"]]["required"] += required
                month_totals[month["key"]]["submitted"] += submitted
            total_percent = round((station_submitted / station_required) * 100, 2) if station_required else None
            station_report = {
                "code": station["station_code"],
                "name": station["station_name"],
                "monthly": monthly,
                "required": station_required,
                "submitted": station_submitted,
                "total_percent": total_percent,
                "total_display": format_percent(total_percent),
                "total_low": total_percent is not None and total_percent < 80,
            }
            if station_report["total_low"]:
                below_threshold.append(station_report)
            station_reports.append(station_report)
            cluster_required += station_required
            cluster_submitted += station_submitted

        cluster_percent = round((cluster_submitted / cluster_required) * 100, 2) if cluster_required else None
        cluster_reports.append(
            {
                "name": cluster_name,
                "display_name": cluster_name.replace(" Cluster", "").upper(),
                "vertical_label": " ".join(cluster_name.replace(" Cluster", "").upper().split()),
                "stations": station_reports,
                "required": cluster_required,
                "submitted": cluster_submitted,
                "percent": cluster_percent,
                "display": format_percent(cluster_percent),
                "low": cluster_percent is not None and cluster_percent < 80,
            }
        )
        overall_required += cluster_required
        overall_submitted += cluster_submitted

    monthly_totals = []
    for month in period_info["months"]:
        totals = month_totals[month["key"]]
        percent = round((totals["submitted"] / totals["required"]) * 100, 2) if totals["required"] else None
        monthly_totals.append(
            {
                "key": month["key"],
                "percent": percent,
                "display": format_percent(percent),
                "low": percent is not None and percent < 80,
            }
        )
    overall_percent = round((overall_submitted / overall_required) * 100, 2) if overall_required else 0
    return {
        "period": period_info,
        "clusters": cluster_reports,
        "monthly_totals": monthly_totals,
        "overall_percent": overall_percent,
        "overall_display": format_percent(overall_percent),
        "overall_low": overall_percent < 80,
        "overall_required": overall_required,
        "overall_submitted": overall_submitted,
        "below_threshold_count": len(below_threshold),
        "as_of": report_timestamp_label(),
    }


@app.route("/")
@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_db()
    close_expired_events(conn)
    reconcile_phivolcs_bulletin_variant_duplicates(conn)
    normalize_duplicate_event_keys(conn)
    reconcile_required_submissions(conn)
    conn.commit()
    summary_rows = conn.execute(
        """
        SELECT
            earthquake_events.id,
            earthquake_events.event_key,
            earthquake_events.event_datetime_utc,
            earthquake_events.magnitude,
            earthquake_events.reference_location,
            earthquake_events.region_code,
            COUNT(pqr_required_submissions.id) AS required_count,
            SUM(CASE WHEN pqr_required_submissions.status = 'submitted' THEN 1 ELSE 0 END) AS submitted_count,
            SUM(CASE WHEN pqr_required_submissions.status = 'pending' THEN 1 ELSE 0 END) AS pending_count
        FROM earthquake_events
        LEFT JOIN pqr_required_submissions
            ON pqr_required_submissions.event_id = earthquake_events.id
        WHERE earthquake_events.status = 'open'
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
        GROUP BY earthquake_events.id
        ORDER BY earthquake_events.event_datetime_utc DESC
        """
    ).fetchall()
    summary = [
        {
            **dict(row),
            "event_datetime_pst": utc_iso_to_pst_display(row["event_datetime_utc"]),
            "status_label": "Complete" if (row["pending_count"] or 0) == 0 else "Open",
        }
        for row in summary_rows
    ]
    recent = conn.execute(
        """
        SELECT pqr_reports.id, earthquake_events.event_key, stations.station_code,
               pqr_reports.remarks, pqr_reports.submitted_at, pqr_reports.updated_at
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        JOIN stations ON stations.id = pqr_reports.station_id
        ORDER BY pqr_reports.submitted_at DESC
        LIMIT 8
        """
    ).fetchall()
    recent_activity = [dict(row) for row in recent]
    sheet_sync_summary = conn.execute(
        """
        SELECT pqr_reports.sheet_sync_status, COUNT(*) AS count
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        WHERE pqr_reports.sheet_sync_status IN ('pending', 'failed')
          AND pqr_reports.created_by IS NOT NULL
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
        GROUP BY pqr_reports.sheet_sync_status
        """
    ).fetchall()
    sheet_sync_counts = {row["sheet_sync_status"]: row["count"] for row in sheet_sync_summary}
    station_status_context = build_station_submission_status(
        conn,
        request.args.get("month"),
        request.args.get("year"),
    )
    response_speed = build_station_response_speed(
        conn,
        station_status_context["selected_month"],
    )
    conn.close()

    open_events = len(summary)
    total_required = sum(row["required_count"] or 0 for row in summary)
    total_submitted = sum(row["submitted_count"] or 0 for row in summary)
    total_pending = sum(row["pending_count"] or 0 for row in summary)
    completion_rate = round((total_submitted / total_required) * 100) if total_required else 0

    today_pst = utc_now().astimezone(APP_LOCAL_TZ).date()
    submitted_today = 0
    for row in recent_activity:
        submitted_at = parse_utc_iso(row["submitted_at"])
        if submitted_at and submitted_at.astimezone(APP_LOCAL_TZ).date() == today_pst:
            submitted_today += 1

    cluster_meta = [
        ("North Luzon (NL)", "NL", "#2877e7"),
        ("South Luzon (SL)", "SL", "#13b9bd"),
        ("Visayas (VIS)", "VIS", "#8654ca"),
        ("Mindanao (MIN)", "MIN", "#74b84b"),
    ]
    cluster_status = []
    for label, region_code, color in cluster_meta:
        region_events = [row for row in summary if row["region_code"] == region_code]
        required = sum(row["required_count"] or 0 for row in region_events)
        submitted = sum(row["submitted_count"] or 0 for row in region_events)
        compliance = round((submitted / required) * 100) if required else 0
        cluster_status.append(
            {
                "label": label,
                "submitted": submitted,
                "required": required,
                "compliance": compliance,
                "color": color,
            }
        )

    kpis = [
        {"label": "Open Earthquake Events", "value": open_events, "accent": "blue", "trend": "Active monitoring"},
        {"label": "Pending PQR Submissions", "value": total_pending, "accent": "orange", "trend": "Across open events"},
        {"label": "Submitted Today", "value": submitted_today, "accent": "green", "trend": "Philippine Standard Time"},
        {"label": "Completion Rate", "value": f"{completion_rate}%", "accent": "indigo", "trend": f"{total_submitted} of {total_required} required"},
        {"label": "Median Response Time", "value": response_speed["network_median_label"], "accent": "purple", "trend": station_status_context["selected_month_label"]},
    ]
    activity_feed = build_dashboard_activity(summary, recent_activity, sheet_sync_counts)

    return render_template(
        "dashboard.html",
        summary=summary[:8],
        recent=recent_activity[:6],
        kpis=kpis,
        station_status=station_status_context,
        cluster_status=cluster_status,
        completion_rate=completion_rate,
        total_pending=total_pending,
        sheet_sync_counts=sheet_sync_counts,
        response_speed=response_speed,
        activity_feed=activity_feed,
    )


@app.route("/station/dashboard")
@login_required
@permission_required("submit")
def station_dashboard():
    user = current_user()
    if user["role"] != "station_user":
        return redirect("/dashboard")

    conn = get_db()
    close_expired_events(conn)
    reconcile_phivolcs_bulletin_variant_duplicates(conn)
    normalize_duplicate_event_keys(conn)
    reconcile_required_submissions(conn)
    conn.commit()

    assigned_stations = assigned_station_rows(conn, user, "submit")
    if not assigned_stations:
        conn.close()
        flash("No station is assigned to your account. Please contact the administrator.")
        return redirect("/logout")
    primary_station = next(
        (row for row in assigned_stations if row["id"] == user.get("station_id")),
        assigned_stations[0],
    )
    station = {
        **dict(primary_station),
        "assignment_count": len(assigned_stations),
        "display_code": (
            f"{primary_station['station_code']} + {len(assigned_stations) - 1} station(s)"
            if len(assigned_stations) > 1
            else primary_station["station_code"]
        ),
    }
    assigned_station_ids = [row["id"] for row in assigned_stations]
    assigned_placeholders = ", ".join("?" for _ in assigned_station_ids)

    pending_rows = conn.execute(
        f"""
        SELECT
            earthquake_events.id,
            earthquake_events.event_key,
            earthquake_events.event_datetime_utc,
            earthquake_events.magnitude,
            earthquake_events.reference_location,
            earthquake_events.region_code,
            pqr_required_submissions.station_id,
            stations.station_code,
            stations.station_name
        FROM pqr_required_submissions
        JOIN earthquake_events
            ON earthquake_events.id = pqr_required_submissions.event_id
        JOIN stations
            ON stations.id = pqr_required_submissions.station_id
        LEFT JOIN pqr_reports
            ON pqr_reports.event_id = earthquake_events.id
           AND pqr_reports.station_id = pqr_required_submissions.station_id
        WHERE pqr_required_submissions.station_id IN ({assigned_placeholders})
          AND pqr_required_submissions.status = 'pending'
          AND earthquake_events.status = 'open'
          AND pqr_reports.id IS NULL
        ORDER BY earthquake_events.event_datetime_utc DESC
        """,
        assigned_station_ids,
    ).fetchall()

    pending_events = []
    near_deadline_count = 0
    for row in pending_rows:
        if not within_update_window(row["event_datetime_utc"]):
            continue
        time_left = station_time_left(row["event_datetime_utc"])
        is_near_deadline = time_left["seconds_left"] > 0 and time_left["hours_left"] <= 3
        if is_near_deadline:
            near_deadline_count += 1
        pending_events.append(
            {
                **dict(row),
                "event_datetime_pst": utc_iso_to_pst_display(row["event_datetime_utc"]),
                "time_left_label": time_left["label"],
                "hours_left": time_left["hours_left"],
                "status_label": "Near Deadline" if is_near_deadline else "Pending",
            }
        )

    recent_rows = conn.execute(
        f"""
        SELECT
            pqr_reports.id,
            pqr_reports.station_id,
            stations.station_code,
            stations.station_name,
            earthquake_events.event_key,
            earthquake_events.event_datetime_utc,
            earthquake_events.magnitude,
            earthquake_events.reference_location,
            pqr_reports.remarks,
            pqr_reports.submitted_at,
            pqr_reports.updated_at,
            pqr_reports.sheet_sync_status,
            pqr_reports.sheet_sync_error
        FROM pqr_reports
        JOIN earthquake_events
            ON earthquake_events.id = pqr_reports.event_id
        JOIN stations
            ON stations.id = pqr_reports.station_id
        WHERE pqr_reports.station_id IN ({assigned_placeholders})
        ORDER BY pqr_reports.submitted_at DESC
        LIMIT 8
        """,
        assigned_station_ids,
    ).fetchall()
    recent_submissions = [
        {
            **dict(row),
            "event_datetime_pst": utc_iso_to_pst_display(row["event_datetime_utc"]),
            "submitted_at_pst": utc_iso_to_pst_display(row["submitted_at"]),
            "can_update": within_update_window(row["event_datetime_utc"]),
        }
        for row in recent_rows
    ]

    compliance_row = conn.execute(
        f"""
        SELECT
            COUNT(pqr_required_submissions.id) AS required_count,
            SUM(CASE WHEN pqr_required_submissions.status = 'submitted' THEN 1 ELSE 0 END) AS submitted_count,
            SUM(CASE WHEN pqr_required_submissions.status = 'pending' THEN 1 ELSE 0 END) AS pending_count
        FROM pqr_required_submissions
        JOIN earthquake_events
            ON earthquake_events.id = pqr_required_submissions.event_id
        WHERE pqr_required_submissions.station_id IN ({assigned_placeholders})
          AND earthquake_events.status = 'open'
        """,
        assigned_station_ids,
    ).fetchone()

    required_count = compliance_row["required_count"] or 0
    submitted_count = compliance_row["submitted_count"] or 0
    pending_count = compliance_row["pending_count"] or 0
    compliance_rate = round((submitted_count / required_count) * 100) if required_count else 100

    today_pst = utc_now().astimezone(APP_LOCAL_TZ).date()
    submitted_today = 0
    for row in recent_submissions:
        submitted_at = parse_utc_iso(row["submitted_at"])
        if submitted_at and submitted_at.astimezone(APP_LOCAL_TZ).date() == today_pst:
            submitted_today += 1

    sync_failed_count = sum(
        1
        for row in recent_submissions
        if row["sheet_sync_status"] == "failed"
    )
    conn.close()

    kpis = [
        {
            "label": "Pending PQRs",
            "value": len(pending_events),
            "accent": "orange",
            "trend": "Events requiring action",
        },
        {
            "label": "Submitted Today",
            "value": submitted_today,
            "accent": "green",
            "trend": "Philippine Standard Time",
        },
        {
            "label": "Near Deadline",
            "value": near_deadline_count,
            "accent": "red",
            "trend": "3 hours or less remaining",
        },
        {
            "label": "Station Compliance",
            "value": f"{compliance_rate}%",
            "accent": "blue",
            "trend": f"{submitted_count} of {required_count} submitted",
        },
    ]

    return render_template(
        "station_dashboard.html",
        station=station,
        kpis=kpis,
        pending_events=pending_events,
        recent_submissions=recent_submissions,
        required_count=required_count,
        submitted_count=submitted_count,
        pending_count=pending_count,
        compliance_rate=compliance_rate,
        sync_failed_count=sync_failed_count,
    )


@app.route("/stations/status")
@login_required
def station_status():
    conn = get_db()
    close_expired_events(conn)
    reconcile_phivolcs_bulletin_variant_duplicates(conn)
    normalize_duplicate_event_keys(conn)
    reconcile_required_submissions(conn)
    conn.commit()
    station_status_context = build_station_submission_status(
        conn,
        request.args.get("month"),
        request.args.get("year"),
    )
    station_status_context.update(
        build_station_event_matrix(conn, station_status_context["selected_month"])
    )
    conn.close()
    return render_template("station_status.html", **station_status_context)


@app.route("/events/new", methods=["GET", "POST"])
@login_required
@permission_required("create_event")
def new_event():
    if request.method == "POST":
        errors = validate_event_form(request.form)
        if errors:
            for error in errors:
                flash(error)
            return render_template("event_form.html")

        event_datetime_utc = parse_datetime_local_as_utc(request.form["event_datetime"])
        event_key = generate_event_key(request.form["event_datetime"], request.form["reference_location"])

        conn = get_db()
        cursor = conn.execute(
            """
            INSERT INTO earthquake_events (
                event_key, event_datetime_utc, latitude, longitude, depth_km,
                magnitude, reference_location, region_code
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_key,
                event_datetime_utc,
                optional_float(request.form.get("latitude")),
                optional_float(request.form.get("longitude")),
                optional_float(request.form.get("depth_km")),
                optional_float(request.form.get("magnitude")),
                request.form["reference_location"].strip(),
                request.form["region_code"],
            ),
        )
        create_required_submissions(
            conn,
            cursor.lastrowid,
            request.form["region_code"],
            optional_float(request.form.get("magnitude")),
            event_datetime_utc,
        )
        conn.commit()
        conn.close()
        flash(f"Earthquake event created: {event_key}")
        return redirect("/dashboard")

    return render_template("event_form.html")


@app.route("/events/phivolcs")
@login_required
@permission_required("create_event")
def phivolcs_import():
    return redirect("/events/list")


@app.route("/events/query")
@login_required
@permission_required("review")
def event_query():
    conn = get_db()
    page, per_page, _offset = pagination_args(default_per_page=10, max_per_page=100)
    manual_event_key = request.args.get("manual_event_key", "").strip()
    recent_event_key = request.args.get("recent_event_key", "").strip()
    selected_event_key = manual_event_key or recent_event_key or request.args.get("event_key", "")
    context = build_event_query_context(
        conn,
        event_key=selected_event_key,
        station_search=request.args.get("station_search", ""),
        report_category=request.args.get("report_category", ""),
        station_cluster=request.args.get("station_cluster", ""),
        page=page,
        per_page=per_page,
    )
    conn.close()
    return render_template("event_query.html", **(context or {}))


@app.route("/events/query/export.csv")
@login_required
@permission_required("export")
def export_event_query_csv():
    conn = get_db()
    manual_event_key = request.args.get("manual_event_key", "").strip()
    recent_event_key = request.args.get("recent_event_key", "").strip()
    selected_event_key = manual_event_key or recent_event_key or request.args.get("event_key", "")
    context = build_event_query_context(
        conn,
        event_key=selected_event_key,
        station_search=request.args.get("station_search", ""),
        report_category=request.args.get("report_category", ""),
        station_cluster=request.args.get("station_cluster", ""),
        page=1,
        per_page=100000,
    )
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Event ID", "Station Code", "Station Name", "Cluster",
        "P-Arrival (s)", "S-Arrival (s)", "Duration (s)",
        "Observed Intensities", "Instrumental Intensities",
        "Verified Areas without Intensities", "Submitted", "Updated",
    ])
    if context:
        for row in context["all_rows"]:
            writer.writerow([
                context["event"]["event_key"],
                row["station_code"],
                row.get("station_name") or "",
                row["cluster_name"],
                row.get("p_arrival") or "",
                row.get("s_arrival") or "",
                row.get("duration") or "",
                row.get("observed_intensities") or "",
                row.get("instrumental_intensities") or "",
                row.get("verified_areas_without_intensities") or "",
                row.get("submitted_at") or "",
                row.get("updated_at") or "",
            ])
    filename = f"event_query_{context['event']['event_key'] if context else 'empty'}.csv"
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.route("/events/list")
@login_required
@permission_required("create_event")
def events_list():
    conn = get_db()
    close_expired_events(conn)
    reconcile_phivolcs_bulletin_variant_duplicates(conn)
    normalize_duplicate_event_keys(conn)
    reconcile_required_submissions(conn)
    conn.commit()
    page, per_page, offset = pagination_args(default_per_page=50, max_per_page=100)
    total_events = conn.execute(
        """
        SELECT COUNT(*) AS count
        FROM earthquake_events
        WHERE status = 'open'
        """
    ).fetchone()["count"]

    db_events = conn.execute(
        """
        SELECT earthquake_events.*,
               COALESCE(required_stats.required_count, 0) AS required_count,
               COALESCE(required_stats.submitted_count, 0) AS submitted_count,
               COALESCE(required_stats.pending_count, 0) AS pending_count
        FROM earthquake_events
        LEFT JOIN (
            SELECT event_id,
                   COUNT(*) AS required_count,
                   SUM(CASE WHEN status = 'submitted' THEN 1 ELSE 0 END) AS submitted_count,
                   SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) AS pending_count
            FROM pqr_required_submissions
            GROUP BY event_id
        ) AS required_stats
            ON required_stats.event_id = earthquake_events.id
        WHERE earthquake_events.status = 'open'
        ORDER BY earthquake_events.event_datetime_utc DESC
        LIMIT ? OFFSET ?
        """,
        (per_page, offset),
    ).fetchall()
    events = [
        {
            **dict(event),
            "event_datetime_pst": utc_iso_to_pst_display(event["event_datetime_utc"]),
        }
        for event in db_events
    ]
    for index, event in enumerate(events):
        event["is_import_placeholder"] = is_import_placeholder_event(event)
        event["is_excluded_from_pqr_rating"] = bool(event.get("exclude_from_pqr_rating"))
        event["merge_candidates"] = []
        if event["is_excluded_from_pqr_rating"]:
            event["can_exclude_from_pqr_rating"] = False
            continue
        above_event = events[index - 1] if index > 0 else None
        below_event = events[index + 1] if index + 1 < len(events) else None
        for direction, target_event in (("above", above_event), ("below", below_event)):
            candidate = build_event_merge_candidate(conn, event, target_event, direction, max_minutes=10)
            if candidate:
                event["merge_candidates"].append(candidate)
        event["can_exclude_from_pqr_rating"] = event["is_import_placeholder"]
    last_sync = conn.execute(
        """
        SELECT *
        FROM sync_runs
        WHERE source = 'PHIVOLCS'
        ORDER BY started_at DESC
        LIMIT 1
        """
    ).fetchone()
    conn.close()
    pagination = build_pagination(page, per_page, total_events)
    return render_template(
        "events_list.html",
        events=events,
        last_sync=last_sync,
        pagination=pagination,
        window_hours=22,
    )


@app.route("/events/sync", methods=["POST"])
@login_required
@permission_required("create_event")
def sync_events():
    result = run_phivolcs_sync()
    if result["status"] == "already_running":
        flash("PHIVOLCS sync is already running. Please check again shortly.")
    elif result["status"] == "success":
        flash(f"PHIVOLCS sync complete. New events added: {result['imported_count']}.")
    else:
        flash(f"PHIVOLCS sync failed: {result['error_message']}")
    return redirect("/events/list")


@app.route("/events/<int:placeholder_id>/merge-pqr/<int:target_id>", methods=["GET"])
@login_required
@permission_required("create_event")
def confirm_event_pqr_merge(placeholder_id, target_id):
    conn = get_db()
    context = get_event_merge_context(conn, placeholder_id, target_id)
    conn.close()
    back_url = safe_return_url(request.args.get("return_url") or request.referrer)
    if not context:
        flash("This event pair is not eligible for PQR merge.")
        return redirect(back_url)
    return render_template("event_merge_confirm.html", back_url=back_url, **context)


@app.route("/events/<int:placeholder_id>/merge-pqr/<int:target_id>", methods=["POST"])
@login_required
@permission_required("create_event")
def submit_event_pqr_merge(placeholder_id, target_id):
    user = current_user()
    back_url = safe_return_url(request.form.get("return_url") or request.referrer)
    conflict_actions = {
        key.removeprefix("conflict_action_"): value
        for key, value in request.form.items()
        if key.startswith("conflict_action_")
    }
    conn = get_db()
    try:
        result = merge_placeholder_pqr_reports(
            conn,
            placeholder_id,
            target_id,
            user["id"] if user else None,
            conflict_actions,
        )
        normalize_duplicate_event_keys(conn)
        reconcile_required_submissions(conn)
        conn.commit()
    except Exception as error:
        conn.rollback()
        conn.close()
        flash(f"PQR merge failed: {error}")
        return redirect(back_url)
    conn.close()
    flash(
        f"Merged {result['moved']} PQR report(s) from {result['source_key']} "
        f"into {result['target_key']}. Conflicts: {result['kept_target']} kept target, "
        f"{result['replaced']} replaced, {result['skipped']} skipped. Historical merge rows were not re-sent to Google Sheets."
    )
    return redirect(back_url)


@app.route("/events/<int:event_id>/pqr-rating-exclusion", methods=["POST"])
@login_required
@permission_required("create_event")
def update_event_pqr_rating_exclusion(event_id):
    user = current_user()
    back_url = safe_return_url(request.form.get("return_url") or request.referrer)
    excluded = request.form.get("excluded") == "1"
    reason = request.form.get("reason", "")
    conn = get_db()
    try:
        event_key = set_event_pqr_rating_exclusion(
            conn,
            event_id,
            excluded,
            reason,
            user["id"] if user else None,
        )
        reconcile_required_submissions(conn)
        normalize_duplicate_event_keys(conn)
        conn.commit()
    except Exception as error:
        conn.rollback()
        conn.close()
        flash(f"Could not update PQR rating exclusion: {error}")
        return redirect(back_url)
    conn.close()
    if excluded:
        flash(f"Event {event_key} excluded from PQR rating.")
    else:
        flash(f"Event {event_key} included in PQR rating.")
    return redirect(back_url)


@app.route("/events/import-phivolcs", methods=["GET"])
@login_required
@permission_required("create_event")
def phivolcs_archive_import_preview():
    current_year = utc_now().astimezone(APP_LOCAL_TZ).year
    selected_year = parse_year_value(request.args.get("year"), current_year) or current_year
    selected_month = request.args.get("month") or str(utc_now().astimezone(APP_LOCAL_TZ).month)
    scan_requested = request.args.get("scan") == "1"
    month_options = available_phivolcs_archive_months(selected_year)
    preview_rows = []
    scan_errors = []
    if scan_requested:
        conn = get_db()
        preview_rows, scan_errors = scan_phivolcs_archive_events(conn, selected_year, selected_month)
        conn.close()
    return render_template(
        "phivolcs_archive_import.html",
        selected_year=selected_year,
        selected_month=selected_month,
        year_options=sorted({2026, current_year, selected_year}, reverse=True),
        month_options=month_options,
        preview_rows=preview_rows,
        scan_errors=scan_errors,
        scan_requested=scan_requested,
    )


@app.route("/events/import-phivolcs", methods=["POST"])
@login_required
@permission_required("create_event")
def phivolcs_archive_import_submit():
    current_year = utc_now().astimezone(APP_LOCAL_TZ).year
    selected_year = parse_year_value(request.form.get("year"), current_year) or current_year
    selected_month = request.form.get("month") or str(utc_now().astimezone(APP_LOCAL_TZ).month)
    selected_tokens = set(request.form.getlist("import_token"))
    legacy_selected_urls = set(request.form.getlist("source_url"))
    if not selected_tokens and not legacy_selected_urls:
        flash("Select at least one new PHIVOLCS event to import.")
        return redirect(f"/events/import-phivolcs?scan=1&year={selected_year}&month={selected_month}")

    conn = get_db()
    imported = 0
    skipped = 0
    errors = []
    try:
        preview_rows, scan_errors = scan_phivolcs_archive_events(conn, selected_year, selected_month)
        errors.extend(scan_errors)
        for row in preview_rows:
            event = row["event"]
            if row["import_token"] not in selected_tokens and event.source_url not in legacy_selected_urls:
                continue
            if row["status"] != "New":
                skipped += 1
                continue
            event_id, created = upsert_phivolcs_event(conn, event)
            mark_phivolcs_archive_felt_hint(conn, event_id, event.source_url)
            create_required_submissions(conn, event_id, event.region_code, event.magnitude, event.event_datetime_utc)
            imported += 1 if created else 0
            skipped += 0 if created else 1
        reconcile_phivolcs_bulletin_variant_duplicates(conn)
        normalize_duplicate_event_keys(conn)
        reconcile_required_submissions(conn)
        conn.commit()
    except Exception as error:
        conn.rollback()
        conn.close()
        flash(f"PHIVOLCS archive import failed: {error}")
        return redirect(f"/events/import-phivolcs?scan=1&year={selected_year}&month={selected_month}")
    conn.close()

    message = f"PHIVOLCS archive import complete: {imported} new event(s) imported, {skipped} skipped."
    if errors:
        message += " Some archive pages could not be scanned: " + "; ".join(errors[:3])
    flash(message)
    return redirect(f"/events/import-phivolcs?scan=1&year={selected_year}&month={selected_month}")


@app.route("/pqr/new", methods=["GET"])
@login_required
@permission_required("submit")
def new_pqr():
    conn = get_db()
    user = current_user()
    close_expired_events(conn)
    reconcile_phivolcs_bulletin_variant_duplicates(conn)
    normalize_duplicate_event_keys(conn)
    reconcile_required_submissions(conn)
    conn.commit()
    station_sql = """
        SELECT id, station_code, station_name, region_code, cluster_name,
               station_type, include_in_pqr_compliance, is_one_manned, station_status
        FROM stations
        WHERE is_active = 1
          AND include_in_pqr_compliance = 1
    """
    params = []
    restrict_to_assigned_stations = user["role"] == "station_user" and not is_effective_duty_officer(user)
    if restrict_to_assigned_stations:
        assigned_station_ids = station_assignment_ids(conn, user, "submit")
        if not assigned_station_ids:
            conn.close()
            flash("No station is assigned to your account. Please contact the administrator.")
            return redirect("/logout")
        station_sql += " AND id IN ({})".format(", ".join("?" for _ in assigned_station_ids))
        params.extend(assigned_station_ids)
    station_sql += " ORDER BY CASE station_code " + " ".join(
        f"WHEN '{code}' THEN {index}" for index, code in enumerate(PQR_STATION_CODES)
    ) + " ELSE 999 END"
    stations = conn.execute(station_sql, params).fetchall()

    selected_station_id = request.args.get("station_id", "")
    selected_event_id = request.args.get("event_id", "").strip()
    if restrict_to_assigned_stations:
        allowed_station_ids = {str(station["id"]) for station in stations}
        primary_station_id = str(user["station_id"] or "")
        if selected_station_id not in allowed_station_ids:
            selected_station_id = primary_station_id if primary_station_id in allowed_station_ids else (next(iter(allowed_station_ids), ""))
    elif selected_station_id and not any(str(station["id"]) == str(selected_station_id) for station in stations):
        selected_station_id = ""

    can_change_event_range = user["role"] == "admin" or is_effective_duty_officer(user)
    default_date_from, default_date_to = default_pqr_date_range()
    selected_date_from = parse_local_date(request.args.get("date_from")) or default_date_from
    selected_date_to = parse_local_date(request.args.get("date_to")) or default_date_to
    if selected_date_from > selected_date_to:
        selected_date_from, selected_date_to = selected_date_to, selected_date_from

    selected_station = next(
        (station for station in stations if str(station["id"]) == str(selected_station_id)),
        None,
    )
    selected_station_cluster = (
        cluster_name_for_station(dict(selected_station))
        if selected_station
        else ""
    )
    selected_station_code = selected_station["station_code"] if selected_station else ""

    events = []
    if selected_station_id and can_change_event_range:
        events = conn.execute(
            """
            SELECT earthquake_events.id, earthquake_events.event_key,
                   earthquake_events.event_datetime_utc, earthquake_events.magnitude,
                   earthquake_events.reference_location, earthquake_events.region_code,
                   earthquake_events.status AS event_status,
                   pqr_required_submissions.status AS required_status,
                   pqr_reports.id AS report_id, pqr_reports.p_polarity,
                   pqr_reports.p_arrival, pqr_reports.s_marker, pqr_reports.s_arrival,
                   pqr_reports.amplitude, pqr_reports.duration, pqr_reports.event_type,
                   pqr_reports.reserved_k, pqr_reports.remarks,
                   pqr_reports.observed_intensities, pqr_reports.instrumental_intensities,
                   pqr_reports.verified_areas_without_intensities,
                   pqr_reports.sheet_sync_status
            FROM earthquake_events
            LEFT JOIN pqr_required_submissions
                ON pqr_required_submissions.event_id = earthquake_events.id
               AND pqr_required_submissions.station_id = ?
            LEFT JOIN pqr_reports
                ON pqr_reports.event_id = earthquake_events.id
               AND pqr_reports.station_id = ?
            WHERE COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
            ORDER BY earthquake_events.event_datetime_utc DESC
            """,
            (selected_station_id, selected_station_id),
        ).fetchall()
    elif selected_station_id:
        events = conn.execute(
            """
            SELECT earthquake_events.id, earthquake_events.event_key,
                   earthquake_events.event_datetime_utc, earthquake_events.magnitude,
                   earthquake_events.reference_location, earthquake_events.region_code,
                   earthquake_events.status AS event_status,
                   pqr_required_submissions.status AS required_status,
                   pqr_reports.id AS report_id, pqr_reports.p_polarity,
                   pqr_reports.p_arrival, pqr_reports.s_marker, pqr_reports.s_arrival,
                   pqr_reports.amplitude, pqr_reports.duration, pqr_reports.event_type,
                   pqr_reports.reserved_k, pqr_reports.remarks,
                   pqr_reports.observed_intensities, pqr_reports.instrumental_intensities,
                   pqr_reports.verified_areas_without_intensities,
                   pqr_reports.sheet_sync_status
            FROM earthquake_events
            LEFT JOIN pqr_required_submissions
                ON pqr_required_submissions.event_id = earthquake_events.id
               AND pqr_required_submissions.station_id = ?
            LEFT JOIN pqr_reports
                ON pqr_reports.event_id = earthquake_events.id
               AND pqr_reports.station_id = ?
            WHERE (
                pqr_required_submissions.status = 'pending'
                OR pqr_reports.id IS NOT NULL
            )
              AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
            ORDER BY earthquake_events.event_datetime_utc DESC
            """,
            (selected_station_id, selected_station_id),
        ).fetchall()
    elif can_change_event_range:
        events = conn.execute(
            """
            SELECT earthquake_events.id, earthquake_events.event_key,
                   earthquake_events.event_datetime_utc, earthquake_events.magnitude,
                   earthquake_events.reference_location, earthquake_events.region_code,
                   earthquake_events.status AS event_status,
                   'pending' AS required_status
            FROM earthquake_events
            WHERE COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
            ORDER BY earthquake_events.event_datetime_utc DESC
            """
        ).fetchall()
    events = [
        {
            **dict(event),
            "event_datetime_pst": utc_iso_to_pst_display(event["event_datetime_utc"]),
            "is_submitted": bool(dict(event).get("report_id")),
            "can_submit_details": (
                bool(selected_station_id)
                and (not selected_station or event_visible_for_station(dict(event), selected_station))
                and (not selected_station or not station_exempt_from_pqr(dict(selected_station), dict(event)))
                and (
                    (
                        can_change_event_range
                        and (
                            not dict(event).get("report_id")
                            or dict(event).get("sheet_sync_status") != "synced"
                        )
                    )
                    or (
                        not dict(event).get("report_id")
                        and
                        dict(event).get("event_status") == "open"
                        and dict(event).get("required_status") == "pending"
                        and within_update_window(event["event_datetime_utc"])
                    )
                )
            ),
        }
        for event in events
        if (
            event_local_date(event["event_datetime_utc"])
            and selected_date_from <= event_local_date(event["event_datetime_utc"]) <= selected_date_to
            and (not selected_station or event_visible_for_station(dict(event), selected_station))
        )
    ]
    region_filters = sorted({event["region_code"] for event in events if event["region_code"]})
    recent = []
    if selected_station_id:
        recent = conn.execute(
            """
            SELECT pqr_reports.id, earthquake_events.event_key,
                   earthquake_events.event_datetime_utc, earthquake_events.magnitude,
                   pqr_reports.remarks, pqr_reports.submitted_at, pqr_reports.status
            FROM pqr_reports
            JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
            WHERE pqr_reports.station_id = ?
            ORDER BY pqr_reports.submitted_at DESC
            LIMIT 5
            """,
            (selected_station_id,),
        ).fetchall()
    conn.close()
    return render_template(
        "pqr_form.html",
        events=events,
        recent=recent,
        stations=stations,
        region_filters=region_filters,
        station_clusters=STATION_CLUSTERS,
        station_cluster_by_code=STATION_CLUSTER_BY_CODE,
        station_cluster_by_id={
            str(station["id"]): cluster_name_for_station(dict(station))
            for station in stations
        },
        selected_station_id=str(selected_station_id or ""),
        selected_station_cluster=selected_station_cluster,
        selected_event_id=selected_event_id,
        can_change_event_range=can_change_event_range,
        selected_date_from=selected_date_from.isoformat(),
        selected_date_to=selected_date_to.isoformat(),
        topbar_date_from_display=f"{selected_date_from.strftime('%b')} {selected_date_from.day}, {selected_date_from.year}",
        topbar_date_to_display=f"{selected_date_to.strftime('%b')} {selected_date_to.day}, {selected_date_to.year}",
    )


@app.route("/pqr/submit", methods=["POST"])
@login_required
@permission_required("submit")
def submit_pqr():
    user = current_user()
    conn = get_db()
    try:
        errors, submitted_count, report_ids = submit_pqr_rows(conn, request.form, user)
        if errors:
            conn.rollback()
            for error in errors:
                flash(error)
            return redirect("/pqr/new")

        conn.commit()
        flash(
            f"{submitted_count} PQR report(s) submitted successfully. "
            "Google Sheets sync queued and will run in the background."
        )
    except Exception as error:
        conn.rollback()
        if "UNIQUE constraint failed" in str(error):
            flash("Duplicate PQR detected. Use Update instead.")
        else:
            flash(f"Submit failed: {error}")
    finally:
        conn.close()
    return redirect("/pqr/list")


@app.route("/sync/google/retry", methods=["POST"])
@login_required
@permission_required("export")
def retry_google_sync():
    result = sync_pending_google_reports(limit=300)
    if result["status"] == "empty":
        flash("No pending or failed Google Sheets sync items.")
        return redirect("/dashboard")
    if result["status"] == "success":
        flash(f"{result['appended_count']} report(s) synced to Google Sheets.")
    else:
        flash(f"Google Sheets retry failed: {result['error_message']}")
    return redirect("/dashboard")


def build_pqr_monitoring_workbook(report, signatures=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    signatures = signatures or DEFAULT_MONITORING_SIGNATURES
    ws = wb.active
    ws.title = "PQR Monitoring"
    months = report["period"]["months"]
    total_cols = 3 + len(months)
    last_col = get_column_letter(total_cols)
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F8FAFC")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = "PQR MONITORING"
    ws["A1"].font = Font(size=22, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"] = report["period"]["subtitle"]
    ws["A2"].font = Font(size=13, bold=True, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    ws["A3"] = report["overall_display"]
    ws["A3"].font = Font(size=14, bold=True)
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["A4"] = "as of"
    ws["B4"] = report["as_of"]
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)

    header_row = 5
    subheader_row = 6
    ws.merge_cells(start_row=header_row, start_column=1, end_row=subheader_row, end_column=1)
    ws.merge_cells(start_row=header_row, start_column=2, end_row=subheader_row, end_column=2)
    ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=2 + len(months))
    ws.merge_cells(start_row=header_row, start_column=total_cols, end_row=subheader_row, end_column=total_cols)
    ws.cell(header_row, 1).value = "Cluster"
    ws.cell(header_row, 2).value = "SCSS"
    ws.cell(header_row, 3).value = "Percentage per Month"
    ws.cell(header_row, total_cols).value = "Total Percentage per SCSS"
    for index, month in enumerate(months, start=3):
        ws.cell(subheader_row, index).value = month["name"]
    for row in ws.iter_rows(min_row=header_row, max_row=subheader_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_index = 7
    for cluster in report["clusters"]:
        start_row = row_index
        for station in cluster["stations"]:
            ws.cell(row_index, 2).value = station["code"]
            ws.cell(row_index, 2).font = Font(bold=True)
            for month_index, monthly in enumerate(station["monthly"], start=3):
                cell = ws.cell(row_index, month_index)
                cell.value = monthly["display"]
                if monthly["low"]:
                    cell.font = Font(bold=True, color="FF0000")
            total_cell = ws.cell(row_index, total_cols)
            total_cell.value = station["total_display"]
            total_cell.font = Font(bold=True, color="FF0000") if station["total_low"] else Font(bold=True)
            for cell in ws[row_index][:total_cols]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_index += 1
        if row_index > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row_index - 1, end_column=1)
            cluster_cell = ws.cell(start_row, 1)
            cluster_cell.value = cluster["display_name"]
            cluster_cell.font = Font(bold=True)
            cluster_cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)
            cluster_cell.border = border
            for col in range(1, total_cols + 1):
                ws.cell(start_row, col).border = Border(
                    left=medium if col == 1 else thin,
                    right=medium if col == total_cols else thin,
                    top=medium,
                    bottom=thin,
                )
                ws.cell(row_index - 1, col).border = Border(
                    left=medium if col == 1 else thin,
                    right=medium if col == total_cols else thin,
                    top=thin,
                    bottom=medium,
                )

    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
    ws.cell(row_index, 1).value = "TOTAL PERCENTAGE"
    ws.cell(row_index, 1).font = Font(bold=True)
    ws.cell(row_index, 1).alignment = Alignment(horizontal="center")
    for month_index, total in enumerate(report["monthly_totals"], start=3):
        cell = ws.cell(row_index, month_index)
        cell.value = total["display"]
        cell.font = Font(bold=True, color="FF0000") if total["low"] else Font(bold=True)
    ws.cell(row_index, total_cols).value = report["overall_display"]
    ws.cell(row_index, total_cols).font = Font(bold=True, color="FF0000") if report["overall_low"] else Font(bold=True)
    for cell in ws[row_index][:total_cols]:
        cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    signature_row = row_index + 5
    left_end = max(2, total_cols // 2)
    right_start = left_end + 1
    signatures = [
        (signature_row, 1, left_end, f"Prepared By: {signatures['prepared_name']}", f"({signatures['prepared_title']})"),
        (signature_row, right_start, total_cols, f"Reviewed By: {signatures['reviewed_name']}", f"({signatures['reviewed_title']})"),
        (signature_row + 4, 1, left_end, f"Approved By: {signatures['approved_name']}", f"({signatures['approved_title']})"),
        (signature_row + 4, right_start, total_cols, f"Noted By: {signatures['noted_name']}", f"({signatures['noted_title']})"),
    ]
    for row, start_col, end_col, name, title in signatures:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        ws.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 1, end_column=end_col)
        ws.cell(row, start_col).value = name
        ws.cell(row + 1, start_col).value = title
        ws.cell(row, start_col).font = Font(bold=True)
        ws.cell(row + 1, start_col).font = Font(bold=True, italic=True)
        ws.cell(row, start_col).alignment = Alignment(horizontal="center")
        ws.cell(row + 1, start_col).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    for col in range(3, total_cols):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions[last_col].width = 22
    ws.freeze_panes = "C7"
    return wb


def build_pqr_monitoring_pdf(report, signatures=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.utils import simpleSplit
    from reportlab.pdfgen import canvas

    signatures = signatures or DEFAULT_MONITORING_SIGNATURES
    months = report["period"]["months"]
    page_size = landscape(A4) if len(months) > 3 else A4
    page_width, page_height = page_size
    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=page_size)
    margin = 10
    table_width = page_width - (margin * 2)

    title_y = page_height - 22
    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawCentredString(page_width / 2, title_y, "PQR MONITORING")
    pdf.setFont("Helvetica-BoldOblique", 13)
    pdf.drawCentredString(page_width / 2, title_y - 24, report["period"]["subtitle"])
    pdf.setFont("Helvetica-BoldOblique", 14)
    pdf.drawCentredString(page_width / 2, title_y - 46, report["overall_display"])
    pdf.setStrokeColor(colors.HexColor("#2563eb"))
    pdf.setLineWidth(1)
    pdf.line(margin, title_y - 58, page_width - margin, title_y - 58)
    pdf.setStrokeColor(colors.black)
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(margin + 2, title_y - 75, "as of")
    pdf.drawString(margin + 48, title_y - 75, report["as_of"])

    table_top = title_y - 86
    signature_height = 96
    table_bottom = margin + signature_height
    row_count = 2 + sum(len(cluster["stations"]) for cluster in report["clusters"]) + 1
    row_height = max(8.8, min(17.5, (table_top - table_bottom) / max(row_count, 1)))
    header_height = row_height * 2
    data_font = max(5.8, min(8.5, row_height - 3.2))
    header_font = max(6.0, min(8.2, row_height - 3.0))

    cluster_w = 43 if len(months) <= 3 else 36
    station_w = 66 if len(months) <= 3 else 52
    total_w = 96 if len(months) <= 3 else 82
    month_w = (table_width - cluster_w - station_w - total_w) / len(months)
    x_positions = [margin, margin + cluster_w, margin + cluster_w + station_w]
    for _ in months:
        x_positions.append(x_positions[-1] + month_w)
    x_positions.append(page_width - margin)

    def rect(x, y, width, height, fill=None, stroke=colors.black, line_width=0.6):
        pdf.setStrokeColor(stroke)
        pdf.setLineWidth(line_width)
        if fill:
            pdf.setFillColor(fill)
            pdf.rect(x, y, width, height, fill=1, stroke=1)
            pdf.setFillColor(colors.black)
        else:
            pdf.rect(x, y, width, height, fill=0, stroke=1)

    def centered(text, x, y, width, height, font="Helvetica", size=8, color=colors.black):
        pdf.setFillColor(color)
        pdf.setFont(font, size)
        lines = simpleSplit(str(text), font, size, max(width - 4, 8)) or [""]
        line_height = size + 1
        start_y = y + (height + (len(lines) - 1) * line_height) / 2 - size
        for index, line in enumerate(lines[:3]):
            pdf.drawCentredString(x + width / 2, start_y - (index * line_height), line)
        pdf.setFillColor(colors.black)

    header_fill = colors.HexColor("#f8fafc")
    y = table_top - header_height
    rect(margin, y, cluster_w, header_height, header_fill, line_width=0.9)
    rect(margin + cluster_w, y, station_w, header_height, header_fill, line_width=0.9)
    rect(margin + cluster_w + station_w, y + row_height, month_w * len(months), row_height, header_fill, line_width=0.9)
    rect(page_width - margin - total_w, y, total_w, header_height, header_fill, line_width=0.9)
    centered("Cluster", margin, y, cluster_w, header_height, "Helvetica-Bold", header_font)
    centered("SCSS", margin + cluster_w, y, station_w, header_height, "Helvetica-Bold", header_font)
    centered("Percentage per Month", margin + cluster_w + station_w, y + row_height, month_w * len(months), row_height, "Helvetica-Bold", header_font)
    centered("Total Percentage per SCSS", page_width - margin - total_w, y, total_w, header_height, "Helvetica-Bold", header_font)
    for index, month in enumerate(months):
        x = margin + cluster_w + station_w + (index * month_w)
        rect(x, y, month_w, row_height, header_fill, line_width=0.6)
        centered(month["name"], x, y, month_w, row_height, "Helvetica-Bold", header_font)

    y -= row_height
    for cluster in report["clusters"]:
        cluster_height = row_height * len(cluster["stations"])
        cluster_top = y + row_height
        cluster_y = cluster_top - cluster_height
        rect(margin, cluster_y, cluster_w, cluster_height, line_width=0.9)
        rect(page_width - margin - total_w, cluster_y, total_w, cluster_height, line_width=0.9)
        label = cluster["display_name"].replace(" ", "")
        pdf.setFont("Helvetica-Bold", max(5.5, min(7.5, cluster_height / max(len(label), 1) - 0.2)))
        label_chars = list(label)
        char_gap = min(8, cluster_height / max(len(label_chars), 1))
        start_y = cluster_y + cluster_height - char_gap
        for index, char in enumerate(label_chars):
            pdf.drawCentredString(margin + cluster_w / 2, start_y - (index * char_gap), char)
        centered(cluster["display"], page_width - margin - total_w, cluster_y, total_w, cluster_height, "Helvetica-Bold", data_font + 0.4, colors.red if cluster["low"] else colors.black)
        for station in cluster["stations"]:
            rect(margin + cluster_w, y, station_w, row_height)
            centered(station["code"], margin + cluster_w, y, station_w, row_height, "Helvetica-Bold", data_font)
            for index, monthly in enumerate(station["monthly"]):
                x = margin + cluster_w + station_w + (index * month_w)
                rect(x, y, month_w, row_height)
                centered(monthly["display"], x, y, month_w, row_height, "Helvetica-Bold" if monthly["low"] else "Helvetica", data_font, colors.red if monthly["low"] else colors.black)
            y -= row_height

    total_y = y
    rect(margin, total_y, cluster_w + station_w, row_height, line_width=0.9)
    centered("TOTAL PERCENTAGE", margin, total_y, cluster_w + station_w, row_height, "Helvetica-Bold", data_font)
    for index, total in enumerate(report["monthly_totals"]):
        x = margin + cluster_w + station_w + (index * month_w)
        rect(x, total_y, month_w, row_height, line_width=0.9)
        centered(total["display"], x, total_y, month_w, row_height, "Helvetica-Bold", data_font, colors.red if total["low"] else colors.black)
    rect(page_width - margin - total_w, total_y, total_w, row_height, line_width=0.9)
    centered(report["overall_display"], page_width - margin - total_w, total_y, total_w, row_height, "Helvetica-Bold", data_font, colors.red if report["overall_low"] else colors.black)

    sig_top = total_y - 44
    left_x = margin + 2
    right_x = page_width / 2 + 8
    block_w = (page_width / 2) - margin - 12
    pdf.setFont("Helvetica-Bold", 8.4)
    pdf.drawCentredString(left_x + block_w / 2, sig_top, f"Prepared By: {signatures['prepared_name']}")
    pdf.drawCentredString(right_x + block_w / 2, sig_top, f"Reviewed By: {signatures['reviewed_name']}")
    pdf.setFont("Helvetica-BoldOblique", 8.2)
    pdf.drawCentredString(left_x + block_w / 2, sig_top - 14, f"({signatures['prepared_title']})")
    pdf.drawCentredString(right_x + block_w / 2, sig_top - 14, f"({signatures['reviewed_title']})")
    pdf.setFont("Helvetica-Bold", 8.4)
    pdf.drawCentredString(left_x + block_w / 2, sig_top - 58, f"Approved By: {signatures['approved_name']}")
    pdf.drawCentredString(right_x + block_w / 2, sig_top - 58, f"Noted By: {signatures['noted_name']}")
    pdf.setFont("Helvetica-BoldOblique", 8.2)
    pdf.drawCentredString(left_x + block_w / 2, sig_top - 72, f"({signatures['approved_title']})")
    pdf.drawCentredString(right_x + block_w / 2, sig_top - 72, f"({signatures['noted_title']})")

    pdf.showPage()
    pdf.save()
    output.seek(0)
    return output


@app.route("/reports/pqr-monitoring")
@login_required
@permission_required("export")
def pqr_monitoring_report():
    conn = get_db()
    report = build_pqr_monitoring_report(
        conn,
        request.args.get("period_type"),
        request.args.get("year"),
        request.args.get("period"),
    )
    signatures = monitoring_signatures_from_request()
    conn.close()
    return render_template(
        "pqr_monitoring_report.html",
        report=report,
        signatures=signatures,
        selected_period_type=report["period"]["type"],
        selected_year=report["period"]["year"],
        selected_period=report["period"]["period"],
        current_year=utc_now().astimezone(APP_LOCAL_TZ).year,
        title="PQR Monitoring Report",
    )


@app.route("/reports/pqr-monitoring.xlsx")
@login_required
@permission_required("export")
def export_pqr_monitoring_xlsx():
    conn = get_db()
    report = build_pqr_monitoring_report(
        conn,
        request.args.get("period_type"),
        request.args.get("year"),
        request.args.get("period"),
    )
    signatures = monitoring_signatures_from_request()
    conn.close()
    workbook = build_pqr_monitoring_workbook(report, signatures)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"pqr_monitoring_{report['period']['file_label']}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/reports/pqr-monitoring.pdf")
@login_required
@permission_required("export")
def export_pqr_monitoring_pdf():
    conn = get_db()
    report = build_pqr_monitoring_report(
        conn,
        request.args.get("period_type"),
        request.args.get("year"),
        request.args.get("period"),
    )
    signatures = monitoring_signatures_from_request()
    conn.close()
    output = build_pqr_monitoring_pdf(report, signatures)
    filename = f"pqr_monitoring_{report['period']['file_label']}.pdf"
    return Response(
        output.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/reports")
@login_required
@permission_required("export")
def reports_home():
    filters = {
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
        "station": request.args.get("station", "").strip(),
        "event_key": request.args.get("event_key", "").strip(),
    }
    conn = get_db()
    stats = conn.execute(
        """
        SELECT
            (SELECT COUNT(*) FROM pqr_reports) AS pqr_count,
            (SELECT COUNT(*) FROM earthquake_events) AS event_count,
            (SELECT COUNT(*) FROM stations WHERE is_active = 1) AS station_count,
            (
                SELECT COUNT(*)
                FROM pqr_reports
                JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
                WHERE pqr_reports.sheet_sync_status = 'pending'
                  AND pqr_reports.created_by IS NOT NULL
                  AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
            ) AS pending_sync,
            (
                SELECT COUNT(*)
                FROM pqr_reports
                JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
                WHERE pqr_reports.sheet_sync_status = 'failed'
                  AND pqr_reports.created_by IS NOT NULL
                  AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
            ) AS failed_sync
        """
    ).fetchone()
    recent_exports = conn.execute(
        """
        SELECT earthquake_events.event_key, stations.station_code, pqr_reports.remarks,
               pqr_reports.submitted_at, pqr_reports.sheet_sync_status
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        JOIN stations ON stations.id = pqr_reports.station_id
        ORDER BY pqr_reports.submitted_at DESC
        LIMIT 5
        """
    ).fetchall()
    last_google_import = conn.execute(
        """
        SELECT *
        FROM sync_runs
        WHERE source = 'GOOGLE_SHEET_IMPORT'
        ORDER BY started_at DESC
        LIMIT 1
        """
    ).fetchone()
    invalid_import_counts = conn.execute(
        """
        SELECT status, COUNT(*) AS count
        FROM google_sheet_import_invalid_rows
        GROUP BY status
        """
    ).fetchall()
    invalid_import_counts = {row["status"]: row["count"] for row in invalid_import_counts}
    conn.close()
    return render_template(
        "reports.html",
        filters=filters,
        stats=stats,
        recent_exports=recent_exports,
        last_google_import=last_google_import,
        invalid_import_counts=invalid_import_counts,
        station_clusters=STATION_CLUSTERS.keys(),
        current_year=utc_now().astimezone(APP_LOCAL_TZ).year,
    )


@app.route("/reports/import-pqr", methods=["POST"])
@login_required
@permission_required("export")
def import_pqr_csv():
    upload = request.files.get("pqr_csv")
    if not upload or not upload.filename:
        flash("Choose a CSV file to import.")
        return redirect("/reports")
    filename = secure_filename(upload.filename)
    if not filename.lower().endswith(".csv"):
        flash("PQR import accepts CSV files only.")
        return redirect("/reports")

    tmp_path = None
    conn = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
            tmp_path = tmp.name
            upload.save(tmp)
        conn = get_db()
        stats = import_data_sheet_csv(conn, tmp_path)
        normalize_duplicate_event_keys(conn)
        reconcile_required_submissions(conn)
        conn.commit()
        conn.close()
        conn = None
        flash(import_pqr_summary_message("PQR CSV import complete", stats))
    except Exception as error:
        if conn:
            try:
                conn.rollback()
                conn.close()
            except Exception:
                pass
        flash(f"PQR import failed: {error}")
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)
    return redirect("/reports")


@app.route("/reports/import-pqr-gsheet", methods=["POST"])
@login_required
@permission_required("export")
def import_pqr_google_sheet():
    result = run_google_sheet_import()
    if result["status"] == "already_running":
        flash("Google Sheets PQR import is already running. Please check the import status shortly.")
    elif result["status"] == "success":
        flash(import_pqr_summary_message("Google Sheets PQR import complete", result["stats"]))
    else:
        flash(f"Google Sheets PQR import failed: {result['error_message']}")
    return redirect("/reports")


def run_google_sheet_import():
    source = "GOOGLE_SHEET_IMPORT"
    started_at = to_utc_iso(utc_now())
    conn = get_db()
    existing_run = conn.execute(
        """
        SELECT id, started_at
        FROM sync_runs
        WHERE source = ?
          AND status = 'running'
        ORDER BY started_at DESC
        LIMIT 1
        """,
        (source,),
    ).fetchone()
    if existing_run:
        started = parse_utc_iso(existing_run["started_at"])
        elapsed = (utc_now() - started).total_seconds() if started else 0
        if elapsed < 15 * 60:
            conn.close()
            return {
                "status": "already_running",
                "stats": empty_import_stats(),
                "error_message": "",
            }
        conn.execute(
            """
            UPDATE sync_runs
            SET status = 'failed',
                finished_at = ?,
                error_message = 'Google Sheets import timed out and was marked stale.'
            WHERE id = ?
            """,
            (to_utc_iso(utc_now()), existing_run["id"]),
        )
        conn.commit()

    cursor = conn.execute(
        """
        INSERT INTO sync_runs (source, status, started_at)
        VALUES (?, 'running', ?)
        """,
        (source, started_at),
    )
    sync_run_id = cursor.lastrowid
    conn.commit()
    conn.close()

    conn = None
    stats = empty_import_stats()
    try:
        sheet_rows = fetch_pqr_sheet_import_rows()
        conn = get_db()
        stats = import_data_sheet_rows(
            conn,
            sheet_rows,
            invalid_recorder=record_invalid_google_sheet_row,
            sync_run_id=sync_run_id,
        )
        normalize_duplicate_event_keys(conn)
        reconcile_required_submissions(conn)
        conn.execute(
            """
            UPDATE sync_runs
            SET status = 'success',
                finished_at = ?,
                imported_count = ?,
                skipped_count = ?,
                invalid_count = ?,
                summary = ?
            WHERE id = ?
            """,
            (
                to_utc_iso(utc_now()),
                stats["imported"],
                stats["skipped"],
                stats["invalid"],
                import_stats_summary(stats),
                sync_run_id,
            ),
        )
        conn.commit()
        conn.close()
        conn = None
        return {"status": "success", "stats": stats, "error_message": ""}
    except Exception as error:
        if conn:
            try:
                conn.rollback()
                conn.close()
            except Exception:
                pass
        failure_conn = get_db()
        try:
            failure_conn.execute(
                """
                UPDATE sync_runs
                SET status = 'failed',
                    finished_at = ?,
                    imported_count = ?,
                    skipped_count = ?,
                    invalid_count = ?,
                    summary = ?,
                    error_message = ?
                WHERE id = ?
                """,
                (
                    to_utc_iso(utc_now()),
                    stats["imported"],
                    stats["skipped"],
                    stats["invalid"],
                    import_stats_summary(stats),
                    str(error),
                    sync_run_id,
                ),
            )
            failure_conn.commit()
        finally:
            failure_conn.close()
        return {"status": "failed", "stats": stats, "error_message": str(error)}


def record_invalid_google_sheet_row(conn, sync_run_id, sheet_row_number, raw_row, normalized_row, error_reason):
    raw_data_json = json.dumps(
        {key: value for key, value in dict(raw_row).items() if not str(key).startswith("__")},
        sort_keys=True,
        ensure_ascii=False,
        default=str,
    )
    row_hash = import_row_hash({key: value for key, value in dict(raw_row).items() if not str(key).startswith("__")})
    conn.execute(
        """
        INSERT INTO google_sheet_import_invalid_rows (
            sync_run_id, sheet_row_number, row_hash, raw_data_json,
            station_code, officer_initials, event_key, remarks, error_reason,
            status, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)
        ON CONFLICT(row_hash) DO UPDATE SET
            sync_run_id = excluded.sync_run_id,
            sheet_row_number = excluded.sheet_row_number,
            raw_data_json = excluded.raw_data_json,
            station_code = excluded.station_code,
            officer_initials = excluded.officer_initials,
            event_key = excluded.event_key,
            remarks = excluded.remarks,
            error_reason = excluded.error_reason,
            updated_at = excluded.updated_at
        """,
        (
            sync_run_id,
            sheet_row_number,
            row_hash,
            raw_data_json,
            normalized_row.get("station_code"),
            normalized_row.get("officer"),
            normalized_row.get("event_key"),
            normalized_row.get("remarks"),
            error_reason,
            to_utc_iso(utc_now()),
        ),
    )


@app.route("/reports/import-invalid")
@login_required
@permission_required("export")
def invalid_import_rows():
    status = request.args.get("status", "pending").strip().lower()
    if status not in {"pending", "resolved", "ignored", "all"}:
        status = "pending"
    params = []
    where = ""
    if status != "all":
        where = "WHERE google_sheet_import_invalid_rows.status = ?"
        params.append(status)
    conn = get_db()
    db_rows = conn.execute(
        f"""
        SELECT google_sheet_import_invalid_rows.*,
               sync_runs.started_at AS import_started_at,
               users.display_name AS resolved_by_name
        FROM google_sheet_import_invalid_rows
        LEFT JOIN sync_runs ON sync_runs.id = google_sheet_import_invalid_rows.sync_run_id
        LEFT JOIN users ON users.id = google_sheet_import_invalid_rows.resolved_by
        {where}
        ORDER BY
            CASE google_sheet_import_invalid_rows.status
              WHEN 'pending' THEN 0
              WHEN 'ignored' THEN 1
              ELSE 2
            END,
            google_sheet_import_invalid_rows.updated_at DESC,
            google_sheet_import_invalid_rows.created_at DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    rows = [invalid_import_row_view(row) for row in db_rows]
    counts = conn.execute(
        """
        SELECT status, COUNT(*) AS count
        FROM google_sheet_import_invalid_rows
        GROUP BY status
        """
    ).fetchall()
    counts = {row["status"]: row["count"] for row in counts}
    problem_counts = build_invalid_import_problem_counts(conn, status)
    conn.close()
    return render_template(
        "invalid_import_rows.html",
        title="Invalid PQR Review",
        rows=rows,
        counts=counts,
        problem_counts=problem_counts,
        status=status,
    )


@app.route("/reports/import-invalid/<int:row_id>/edit", methods=["GET", "POST"])
@login_required
@permission_required("export")
def edit_invalid_import_row(row_id):
    user = current_user()
    conn = get_db()
    invalid_row = conn.execute(
        "SELECT * FROM google_sheet_import_invalid_rows WHERE id = ?",
        (row_id,),
    ).fetchone()
    if not invalid_row:
        conn.close()
        flash("Invalid import row was not found.")
        return redirect("/reports/import-invalid")

    raw_data = json.loads(invalid_row["raw_data_json"] or "{}")
    stations = conn.execute(
        "SELECT id, station_code, station_name, region_code FROM stations WHERE is_active = 1 ORDER BY station_code"
    ).fetchall()
    values = invalid_import_form_values(invalid_row, raw_data, request.form if request.method == "POST" else None)
    errors = []

    if request.method == "POST":
        errors = validate_invalid_import_correction(conn, values)
        if not errors:
            station = conn.execute(
                "SELECT id, region_code FROM stations WHERE id = ?",
                (values["station_id"],),
            ).fetchone()
            event_id, _created_event = ensure_event(conn, values["event_key"], station["region_code"], return_created=True)
            existing_report = conn.execute(
                """
                SELECT id
                FROM pqr_reports
                WHERE event_id = ? AND station_id = ?
                """,
                (event_id, values["station_id"]),
            ).fetchone()
            if existing_report:
                errors.append("A PQR report already exists for this event and station.")
            else:
                submitted_at = values["submitted_at"] or to_utc_iso(utc_now())
                cursor = conn.execute(
                    """
                    INSERT INTO pqr_reports (
                        event_id, station_id, officer_initials, p_polarity, p_arrival,
                        s_marker, s_arrival, amplitude, duration, event_type, reserved_k,
                        remarks, observed_intensities, instrumental_intensities,
                        verified_areas_without_intensities, submitted_at, updated_at,
                        created_by, sheet_sync_status
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'synced')
                    """,
                    (
                        event_id,
                        values["station_id"],
                        values["officer_initials"],
                        values["p_polarity"],
                        values["p_arrival"],
                        values["s_marker"],
                        values["s_arrival"],
                        optional_float(values["amplitude"]),
                        optional_float(values["duration"]),
                        values["event_type"],
                        values["reserved_k"],
                        values["remarks"],
                        values["observed_intensities"],
                        values["instrumental_intensities"],
                        values["verified_areas_without_intensities"],
                        submitted_at,
                        values["updated_at"] or None,
                        user["id"],
                    ),
                )
                conn.execute(
                    """
                    UPDATE pqr_required_submissions
                    SET status = 'submitted'
                    WHERE event_id = ? AND station_id = ?
                    """,
                    (event_id, values["station_id"]),
                )
                conn.execute(
                    """
                    UPDATE google_sheet_import_invalid_rows
                    SET status = 'resolved',
                        resolved_report_id = ?,
                        resolved_by = ?,
                        resolved_at = ?,
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        cursor.lastrowid,
                        user["id"],
                        to_utc_iso(utc_now()),
                        to_utc_iso(utc_now()),
                        row_id,
                    ),
                )
                update_event_felt_status(conn, event_id)
                conn.commit()
                conn.close()
                flash("Invalid Google Sheet row corrected and submitted to the database.")
                return redirect("/reports/import-invalid")
        conn.rollback()

    conn.close()
    return render_template(
        "invalid_import_edit.html",
        title="Correct Invalid PQR",
        invalid_row=invalid_row,
        raw_data=raw_data,
        stations=stations,
        values=values,
        errors=errors,
    )


@app.route("/reports/import-invalid/<int:row_id>/ignore", methods=["POST"])
@login_required
@permission_required("export")
def ignore_invalid_import_row(row_id):
    user = current_user()
    reason = request.form.get("ignore_reason", "").strip()
    conn = get_db()
    conn.execute(
        """
        UPDATE google_sheet_import_invalid_rows
        SET status = 'ignored',
            ignore_reason = ?,
            resolved_by = ?,
            resolved_at = ?,
            updated_at = ?
        WHERE id = ?
          AND status = 'pending'
        """,
        (reason, user["id"], to_utc_iso(utc_now()), to_utc_iso(utc_now()), row_id),
    )
    conn.commit()
    conn.close()
    flash("Invalid Google Sheet row ignored.")
    return redirect("/reports/import-invalid")


@app.route("/reports/import-invalid/accept-missing-officer", methods=["POST"])
@login_required
@permission_required("export")
def accept_missing_officer_invalid_rows():
    user = current_user()
    conn = get_db()
    rows = conn.execute(
        """
        SELECT *
        FROM google_sheet_import_invalid_rows
        WHERE status = 'pending'
        ORDER BY created_at
        """
    ).fetchall()
    accepted = 0
    skipped = 0
    for row in rows:
        problems = invalid_import_problem_items(row["error_reason"])
        if not can_accept_missing_officer_only(problems):
            skipped += 1
            continue
        result = submit_missing_officer_invalid_row(conn, row, user["id"])
        if result:
            accepted += 1
        else:
            skipped += 1
    conn.commit()
    conn.close()
    flash(
        f"Accepted {accepted} row(s) with only missing station officer initials. "
        f"{skipped} row(s) still need review."
    )
    return redirect("/reports/import-invalid")


@app.route("/reports/import-invalid.csv")
@login_required
@permission_required("export")
def export_invalid_import_rows():
    status = request.args.get("status", "pending").strip().lower()
    params = []
    where = ""
    if status in {"pending", "resolved", "ignored"}:
        where = "WHERE status = ?"
        params.append(status)
    conn = get_db()
    rows = conn.execute(
        f"""
        SELECT id, sync_run_id, sheet_row_number, station_code, officer_initials,
               event_key, remarks, error_reason, status, created_at, updated_at,
               resolved_report_id, resolved_at, ignore_reason
        FROM google_sheet_import_invalid_rows
        {where}
        ORDER BY created_at DESC
        """,
        params,
    ).fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id", "sync_run_id", "sheet_row_number", "station_code", "officer_initials",
        "event_key", "remarks", "error_reason", "suggested_fix", "status", "created_at", "updated_at",
        "resolved_report_id", "resolved_at", "ignore_reason",
    ])
    for row in rows:
        row_view = invalid_import_row_view(row)
        writer.writerow([
            row["id"],
            row["sync_run_id"],
            row["sheet_row_number"],
            row["station_code"],
            row["officer_initials"],
            row["event_key"],
            row["remarks"],
            row["error_reason"],
            row_view["suggested_fix"],
            row["status"],
            row["created_at"],
            row["updated_at"],
            row["resolved_report_id"],
            row["resolved_at"],
            row["ignore_reason"],
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=invalid_google_sheet_pqr_rows.csv"},
    )


def submit_missing_officer_invalid_row(conn, invalid_row, user_id):
    raw_data = json.loads(invalid_row["raw_data_json"] or "{}")
    values = invalid_import_form_values(invalid_row, raw_data)
    station = conn.execute(
        """
        SELECT id, region_code
        FROM stations
        WHERE UPPER(station_code) = UPPER(?)
          AND is_active = 1
        """,
        (values["station_code"],),
    ).fetchone()
    if not station or not parse_event_key_to_utc(values["event_key"]) or not values["remarks"]:
        return False
    event_id, _created_event = ensure_event(conn, values["event_key"], station["region_code"], return_created=True)
    existing_report = conn.execute(
        """
        SELECT id
        FROM pqr_reports
        WHERE event_id = ? AND station_id = ?
        """,
        (event_id, station["id"]),
    ).fetchone()
    now = to_utc_iso(utc_now())
    if existing_report:
        conn.execute(
            """
            UPDATE google_sheet_import_invalid_rows
            SET status = 'resolved',
                resolved_report_id = ?,
                resolved_by = ?,
                resolved_at = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (existing_report["id"], user_id, now, now, invalid_row["id"]),
        )
        return True
    submitted_at = normalize_timestamp_to_utc_iso(values["submitted_at"], fallback=utc_now())
    updated_at = normalize_timestamp_to_utc_iso(values["updated_at"])
    cursor = conn.execute(
        """
        INSERT INTO pqr_reports (
            event_id, station_id, officer_initials, p_polarity, p_arrival,
            s_marker, s_arrival, amplitude, duration, event_type, reserved_k,
            remarks, observed_intensities, instrumental_intensities,
            verified_areas_without_intensities, submitted_at, updated_at,
            created_by, sheet_sync_status
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'synced')
        """,
        (
            event_id,
            station["id"],
            MISSING_OFFICER_PLACEHOLDER,
            values["p_polarity"],
            values["p_arrival"],
            values["s_marker"],
            values["s_arrival"],
            optional_float(values["amplitude"]),
            optional_float(values["duration"]),
            values["event_type"],
            values["reserved_k"],
            values["remarks"],
            values["observed_intensities"],
            values["instrumental_intensities"],
            values["verified_areas_without_intensities"],
            submitted_at,
            updated_at,
            user_id,
        ),
    )
    conn.execute(
        """
        UPDATE pqr_required_submissions
        SET status = 'submitted'
        WHERE event_id = ? AND station_id = ?
        """,
        (event_id, station["id"]),
    )
    conn.execute(
        """
        UPDATE google_sheet_import_invalid_rows
        SET status = 'resolved',
            resolved_report_id = ?,
            resolved_by = ?,
            resolved_at = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (cursor.lastrowid, user_id, now, now, invalid_row["id"]),
    )
    update_event_felt_status(conn, event_id)
    return True


INVALID_IMPORT_FIXES = {
    "Missing station code": "Select the correct station from the dropdown before submitting.",
    "Missing station officer initials": f"Accepted automatically only when this is the only problem; officer initials become {MISSING_OFFICER_PLACEHOLDER}.",
    "Missing remarks": "Choose the correct PQR remarks value.",
    "Missing event time": "Enter the event time in UTC as YYYYMMDDHHMM.",
    "Invalid event time (expected YYYYMMDDHHMM)": "Correct the event time to a valid UTC timestamp, for example 202608130131.",
}


def invalid_import_row_view(row):
    item = dict(row)
    problems = invalid_import_problem_items(item.get("error_reason"))
    item["problem_items"] = problems
    item["suggested_fix"] = invalid_import_suggested_fix(problems)
    return item


def invalid_import_problem_items(error_reason):
    normalized_reason = str(error_reason or "").replace(
        "Invalid event time; expected YYYYMMDDHHMM",
        "Invalid event time (expected YYYYMMDDHHMM)",
    )
    problems = [
        part.strip()
        for part in normalized_reason.split(";")
        if part.strip()
    ]
    return problems or ["Unknown validation problem"]


def invalid_import_suggested_fix(problems):
    fixes = [INVALID_IMPORT_FIXES.get(problem) for problem in problems]
    fixes = [fix for fix in fixes if fix]
    if fixes:
        return " ".join(fixes)
    return "Open Edit & Submit, compare with the original Sheet row, then correct the highlighted fields."


def build_invalid_import_problem_counts(conn, status):
    where = ""
    params = []
    if status != "all":
        where = "WHERE status = ?"
        params.append(status)
    rows = conn.execute(
        f"""
        SELECT error_reason
        FROM google_sheet_import_invalid_rows
        {where}
        """,
        params,
    ).fetchall()
    counts = {}
    for row in rows:
        for problem in invalid_import_problem_items(row["error_reason"]):
            counts[problem] = counts.get(problem, 0) + 1
    return [
        {"problem": problem, "count": count, "suggested_fix": invalid_import_suggested_fix([problem])}
        for problem, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    ]


def invalid_import_form_values(invalid_row, raw_data, form=None):
    station_code = (form.get("station_code") if form else invalid_row["station_code"]) or ""
    return {
        "station_code": station_code.strip(),
        "station_id": (form.get("station_id") if form else ""),
        "event_key": ((form.get("event_key") if form else invalid_row["event_key"]) or "").strip()[:12],
        "officer_initials": ((form.get("officer_initials") if form else invalid_row["officer_initials"]) or "").strip(),
        "p_polarity": ((form.get("p_polarity") if form else raw_data.get("P-Polarity")) or "").strip(),
        "reserved_k": ((form.get("reserved_k") if form else raw_data.get("Reserved")) or "").strip(),
        "p_arrival": ((form.get("p_arrival") if form else raw_data.get("P-Arrival")) or "").strip(),
        "s_marker": ((form.get("s_marker") if form else raw_data.get("S")) or "").strip(),
        "s_arrival": ((form.get("s_arrival") if form else raw_data.get("S-Arrival")) or "").strip(),
        "amplitude": ((form.get("amplitude") if form else raw_data.get("Amplitude")) or "").strip(),
        "duration": ((form.get("duration") if form else raw_data.get("Duration")) or "").strip(),
        "event_type": ((form.get("event_type") if form else raw_data.get("Type")) or "").strip(),
        "remarks": ((form.get("remarks") if form else invalid_row["remarks"]) or "").strip(),
        "observed_intensities": ((form.get("observed_intensities") if form else raw_data.get("OBSERVED INTENSITIES")) or "").strip(),
        "instrumental_intensities": ((form.get("instrumental_intensities") if form else raw_data.get("INSTRUMENTAL INTENSITIES")) or "").strip(),
        "verified_areas_without_intensities": ((form.get("verified_areas_without_intensities") if form else raw_data.get("Verified Areas without Intensities")) or "").strip(),
        "submitted_at": ((form.get("submitted_at") if form else raw_data.get("Submitted")) or "").strip(),
        "updated_at": ((form.get("updated_at") if form else raw_data.get("Updated")) or "").strip(),
    }


def validate_invalid_import_correction(conn, values):
    errors = []
    station = conn.execute("SELECT id FROM stations WHERE id = ?", (values["station_id"],)).fetchone()
    if not station:
        errors.append("Select a valid station.")
    pqr_errors = validate_pqr_form_strict(
        {
            "event_id": values["event_key"] or "event",
            "station_id": values["station_id"],
            "officer_initials": values["officer_initials"],
            "p_polarity": values["p_polarity"],
            "p_arrival": values["p_arrival"],
            "s_marker": values["s_marker"],
            "s_arrival": values["s_arrival"],
            "amplitude": values["amplitude"],
            "duration": values["duration"],
            "event_type": values["event_type"],
            "remarks": values["remarks"],
        }
    )
    errors.extend(pqr_errors)
    if not parse_event_key_to_utc(values["event_key"]):
        errors.append("Event Time UTC must be a valid YYYYMMDDHHMM value.")
    for timestamp_key, label in [("submitted_at", "Submitted"), ("updated_at", "Updated")]:
        if values[timestamp_key] and not parse_utc_iso(values[timestamp_key]):
            errors.append(f"{label} timestamp is not supported. Use YYYY-MM-DD HH:MM or leave it blank.")
    return errors


def empty_import_stats():
    return {
        "imported": 0,
        "skipped": 0,
        "skipped_duplicates": 0,
        "skipped_app_submitted": 0,
        "invalid": 0,
        "created_events": 0,
        "created_stations": 0,
    }


def import_stats_summary(stats):
    return (
        f"{stats['imported']} imported, {stats['skipped']} skipped "
        f"({stats['skipped_duplicates']} duplicate, {stats['skipped_app_submitted']} app-submitted, "
        f"{stats['invalid']} invalid). Created {stats['created_events']} event(s), "
        f"{stats['created_stations']} station(s)."
    )


@app.route("/reports/user-analytics")
@login_required
@permission_required("export")
def user_analytics_report():
    filters = user_analytics_filters_from_request()
    conn = get_db()
    analytics = build_user_analytics(conn, filters)
    conn.close()
    return render_template(
        "user_analytics.html",
        title="User Analytics",
        filters=filters,
        **analytics,
    )


def user_analytics_filters_from_request():
    today = utc_now().astimezone(APP_LOCAL_TZ).date()
    default_from = today - timedelta(days=30)
    date_from = request.args.get("date_from", default_from.isoformat()).strip()
    date_to = request.args.get("date_to", today.isoformat()).strip()
    cluster = request.args.get("cluster", "").strip()
    role = request.args.get("role", "").strip()
    if role not in USER_ROLES:
        role = ""
    return {
        "date_from": date_from,
        "date_to": date_to,
        "cluster": cluster,
        "role": role,
    }


def user_analytics_date_bounds(filters):
    try:
        start_date = datetime.strptime(filters["date_from"], "%Y-%m-%d").date()
    except ValueError:
        start_date = utc_now().astimezone(APP_LOCAL_TZ).date() - timedelta(days=30)
    try:
        end_date = datetime.strptime(filters["date_to"], "%Y-%m-%d").date()
    except ValueError:
        end_date = utc_now().astimezone(APP_LOCAL_TZ).date()
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    start_utc = f"{start_date.isoformat()}T00:00:00Z"
    end_exclusive = f"{(end_date + timedelta(days=1)).isoformat()}T00:00:00Z"
    return start_date, end_date, start_utc, end_exclusive


def build_user_analytics(conn, filters):
    start_date, end_date, start_utc, end_exclusive = user_analytics_date_bounds(filters)
    users = conn.execute(
        """
        SELECT users.id, users.username, users.display_name, users.role,
               users.station_id, users.last_login_at,
               stations.station_code, stations.station_name, stations.cluster_name
        FROM users
        LEFT JOIN stations ON stations.id = users.station_id
        WHERE users.is_active = 1
          AND (? = '' OR users.role = ?)
          AND (? = '' OR stations.cluster_name = ?)
        ORDER BY users.role, users.display_name
        """,
        (filters["role"], filters["role"], filters["cluster"], filters["cluster"]),
    ).fetchall()

    rows = []
    for user in users:
        user_dict = dict(user)
        station_ids = assigned_station_ids_for_analytics(conn, user_dict["id"], user_dict.get("station_id"))
        station_codes = assigned_station_codes_for_analytics(conn, station_ids, user_dict.get("station_code"))
        metrics = station_metrics_for_user(conn, station_ids, user_dict["id"], start_utc, end_exclusive)
        submitted = metrics["submitted"]
        required = metrics["required"]
        missing = max(required - submitted, 0)
        completion = round((submitted / required) * 100, 1) if required else (100.0 if submitted else 0.0)
        avg_response_seconds = metrics["avg_response_seconds"]
        rows.append(
            {
                "user_id": user_dict["id"],
                "name": user_dict["display_name"],
                "username": user_dict["username"],
                "role": user_dict["role"].replace("_", " ").title(),
                "station": ", ".join(station_codes) if station_codes else "-",
                "cluster": user_dict.get("cluster_name") or "-",
                "submitted": submitted,
                "required": required,
                "completion": completion,
                "avg_response_seconds": avg_response_seconds,
                "avg_response_label": duration_label(avg_response_seconds),
                "late": metrics["late"],
                "missing": missing,
                "last_activity": metrics["last_activity"] or user_dict.get("last_login_at") or "-",
            }
        )

    reporting_rows = [row for row in rows if row["required"] or row["submitted"]]
    total_required = sum(row["required"] for row in rows)
    total_submitted = sum(row["submitted"] for row in rows)
    total_late = sum(row["late"] for row in rows)
    total_missing = sum(row["missing"] for row in rows)
    weighted_completion = round((total_submitted / total_required) * 100, 1) if total_required else 0
    response_values = [row["avg_response_seconds"] for row in rows if row["avg_response_seconds"] is not None]
    avg_response = round(sum(response_values) / len(response_values)) if response_values else None

    completion_chart_rows = sorted(reporting_rows, key=lambda row: (-row["completion"], row["name"]))[:10]
    response_chart_rows = sorted(
        [row for row in reporting_rows if row["avg_response_seconds"] is not None],
        key=lambda row: row["avg_response_seconds"],
    )[:10]
    late_chart_rows = sorted(reporting_rows, key=lambda row: (-row["late"], row["name"]))[:10]

    return {
        "kpis": {
            "reporting_users": len(reporting_rows),
            "total_users": len(rows),
            "avg_completion": weighted_completion,
            "avg_response_label": duration_label(avg_response),
            "late_reports": total_late,
            "missing_reports": total_missing,
        },
        "users": rows,
        "charts": {
            "completion_labels": [row["station"] if row["station"] != "-" else row["name"] for row in completion_chart_rows],
            "completion_values": [row["completion"] for row in completion_chart_rows],
            "response_labels": [row["station"] if row["station"] != "-" else row["name"] for row in response_chart_rows],
            "response_values": [round(row["avg_response_seconds"] / 60, 1) for row in response_chart_rows],
            "late_labels": [row["station"] if row["station"] != "-" else row["name"] for row in late_chart_rows],
            "late_values": [row["late"] for row in late_chart_rows],
            **monthly_submission_trend(conn, start_date, end_date, start_utc, end_exclusive, filters),
        },
        "station_clusters": STATION_CLUSTERS.keys(),
        "roles": USER_ROLES,
    }


def assigned_station_ids_for_analytics(conn, user_id, primary_station_id=None):
    station_ids = set()
    if primary_station_id:
        station_ids.add(primary_station_id)
    rows = conn.execute(
        """
        SELECT station_id
        FROM user_station_assignments
        WHERE user_id = ?
          AND can_view = 1
        """,
        (user_id,),
    ).fetchall()
    for row in rows:
        if row["station_id"]:
            station_ids.add(row["station_id"])
    return sorted(station_ids)


def assigned_station_codes_for_analytics(conn, station_ids, fallback_station_code=None):
    if not station_ids:
        return [fallback_station_code] if fallback_station_code else []
    placeholders = ",".join("?" for _ in station_ids)
    rows = conn.execute(
        f"""
        SELECT station_code
        FROM stations
        WHERE id IN ({placeholders})
        ORDER BY station_code
        """,
        station_ids,
    ).fetchall()
    return [row["station_code"] for row in rows]


def station_metrics_for_user(conn, station_ids, user_id, start_utc, end_exclusive):
    if station_ids:
        placeholders = ",".join("?" for _ in station_ids)
        required = conn.execute(
            f"""
            SELECT COUNT(*) AS count
            FROM pqr_required_submissions
            JOIN earthquake_events ON earthquake_events.id = pqr_required_submissions.event_id
            WHERE pqr_required_submissions.station_id IN ({placeholders})
              AND earthquake_events.event_datetime_utc >= ?
              AND earthquake_events.event_datetime_utc < ?
              AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
            """,
            [*station_ids, start_utc, end_exclusive],
        ).fetchone()["count"]
        report_filter = f"pqr_reports.station_id IN ({placeholders})"
        params = [*station_ids, start_utc, end_exclusive]
    else:
        required = 0
        report_filter = "(pqr_reports.created_by = ? OR pqr_reports.updated_by = ?)"
        params = [user_id, user_id, start_utc, end_exclusive]

    report_row = conn.execute(
        f"""
        SELECT COUNT(*) AS submitted,
               SUM(
                 CASE
                   WHEN (julianday(pqr_reports.submitted_at) - julianday(earthquake_events.event_datetime_utc)) * 24 > 22
                   THEN 1 ELSE 0
                 END
               ) AS late,
               AVG(
                 CASE
                   WHEN pqr_reports.submitted_at IS NOT NULL
                   THEN MAX((julianday(pqr_reports.submitted_at) - julianday(earthquake_events.event_datetime_utc)) * 86400, 0)
                 END
               ) AS avg_response_seconds,
               MAX(COALESCE(pqr_reports.updated_at, pqr_reports.submitted_at)) AS last_activity
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        WHERE {report_filter}
          AND earthquake_events.event_datetime_utc >= ?
          AND earthquake_events.event_datetime_utc < ?
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
        """,
        params,
    ).fetchone()
    return {
        "required": required or 0,
        "submitted": report_row["submitted"] or 0,
        "late": report_row["late"] or 0,
        "avg_response_seconds": round(report_row["avg_response_seconds"]) if report_row["avg_response_seconds"] is not None else None,
        "last_activity": report_row["last_activity"],
    }


def monthly_submission_trend(conn, start_date, end_date, start_utc, end_exclusive, filters):
    params = [start_utc, end_exclusive]
    cluster_clause = ""
    if filters["cluster"]:
        cluster_clause = "AND stations.cluster_name = ?"
        params.append(filters["cluster"])
    rows = conn.execute(
        f"""
        SELECT substr(pqr_reports.submitted_at, 1, 7) AS month_key,
               COUNT(*) AS submitted,
               SUM(
                 CASE
                   WHEN (julianday(pqr_reports.submitted_at) - julianday(earthquake_events.event_datetime_utc)) * 24 > 22
                   THEN 1 ELSE 0
                 END
               ) AS late
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        JOIN stations ON stations.id = pqr_reports.station_id
        WHERE earthquake_events.event_datetime_utc >= ?
          AND earthquake_events.event_datetime_utc < ?
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
          {cluster_clause}
        GROUP BY substr(pqr_reports.submitted_at, 1, 7)
        ORDER BY month_key
        """,
        params,
    ).fetchall()
    by_month = {row["month_key"]: row for row in rows}
    labels = []
    submitted_values = []
    late_values = []
    current = start_date.replace(day=1)
    last = end_date.replace(day=1)
    while current <= last:
        key = current.strftime("%Y-%m")
        labels.append(current.strftime("%b %Y"))
        row = by_month.get(key)
        submitted_values.append(row["submitted"] if row else 0)
        late_values.append(row["late"] if row and row["late"] else 0)
        next_month = (current.replace(day=28) + timedelta(days=4)).replace(day=1)
        current = next_month
    return {
        "trend_labels": labels,
        "trend_submitted": submitted_values,
        "trend_late": late_values,
    }


def duration_label(seconds):
    if seconds is None:
        return "-"
    try:
        total_seconds = max(0, int(seconds))
    except (TypeError, ValueError):
        return "-"
    minutes, remaining_seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f"{hours}h {minutes:02d}m"
    if minutes:
        return f"{minutes}m {remaining_seconds:02d}s"
    return f"{remaining_seconds}s"


def import_pqr_summary_message(prefix, stats):
    return (
        f"{prefix}: {stats['imported']} row(s) imported, {stats['skipped']} skipped "
        f"({stats['skipped_duplicates']} duplicate, {stats['skipped_app_submitted']} app-submitted, "
        f"{stats['invalid']} invalid). Created {stats['created_events']} event(s), "
        f"{stats['created_stations']} station(s)."
    )


@app.route("/pqr/list")
@login_required
def list_pqr():
    event_key = request.args.get("event_key", "").strip()
    station = request.args.get("station", "").strip()
    user = current_user()
    page, per_page, offset = pagination_args(default_per_page=50, max_per_page=100)
    conn = get_db()
    from_clause = """
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        JOIN stations ON stations.id = pqr_reports.station_id
        WHERE 1 = 1
    """
    params = []
    if user["role"] == "station_user":
        station_filter, station_params = station_id_filter_sql(station_assignment_ids(conn, user, "view"))
        from_clause += station_filter
        params.extend(station_params)
    if event_key:
        from_clause += " AND earthquake_events.event_key LIKE ?"
        params.append(f"%{event_key}%")
    if station:
        from_clause += " AND (stations.station_code LIKE ? OR stations.station_name LIKE ?)"
        params.extend([f"%{station}%", f"%{station}%"])
    total_reports = conn.execute(
        f"SELECT COUNT(*) AS count {from_clause}",
        params,
    ).fetchone()["count"]
    reports = conn.execute(
        f"""
        SELECT pqr_reports.id, earthquake_events.event_key,
               earthquake_events.event_datetime_utc, earthquake_events.magnitude,
               earthquake_events.reference_location, stations.station_code,
               stations.station_name, stations.region_code, pqr_reports.officer_initials,
               pqr_reports.remarks, pqr_reports.submitted_at, pqr_reports.updated_at,
               pqr_reports.sheet_sync_status
        {from_clause}
        ORDER BY earthquake_events.event_datetime_utc DESC
        LIMIT ? OFFSET ?
        """,
        params + [per_page, offset],
    ).fetchall()
    conn.close()
    return render_template(
        "pqr_list.html",
        reports=reports,
        event_key=event_key,
        station=station,
        pagination=build_pagination(page, per_page, total_reports),
    )


@app.route("/pqr/edit/<int:report_id>", methods=["GET", "POST"])
@login_required
@permission_required("update")
def edit_pqr(report_id):
    user = current_user()
    conn = get_db()
    report = fetch_report(conn, report_id)
    if not report:
        conn.close()
        flash("PQR report not found.")
        return redirect("/pqr/list")
    if user["role"] == "station_user" and not is_effective_duty_officer(user) and not user_can_access_station(conn, user, report["station_id"], "submit"):
        conn.close()
        flash("Station users can update only assigned stations.")
        return redirect("/pqr/list")
    if not within_update_window(report["event_datetime_utc"]):
        conn.close()
        flash("Update period expired. PQRs can only be edited within 22 hours from event time.")
        return redirect("/pqr/list")

    if request.method == "POST":
        errors = validate_pqr_form(request.form)
        if errors:
            for error in errors:
                flash(error)
            conn.close()
            return render_template("pqr_edit.html", report=report)

        fields = pqr_update_fields_from_form(request.form)
        changes = build_changes(report, fields)
        if changes:
            now = to_utc_iso(utc_now())
            conn.execute(
                """
                UPDATE pqr_reports
                SET officer_initials = ?, p_polarity = ?, p_arrival = ?, s_marker = ?,
                    s_arrival = ?, amplitude = ?, duration = ?, event_type = ?,
                    reserved_k = ?, remarks = ?, observed_intensities = ?,
                    instrumental_intensities = ?, verified_areas_without_intensities = ?,
                    updated_at = ?, updated_by = ?,
                    sheet_sync_status = 'pending',
                    sheet_sync_error = NULL
                WHERE id = ?
                """,
                tuple(fields.values()) + (now, user["id"], report_id),
            )
            for field, old_value, new_value in changes:
                conn.execute(
                    """
                    INSERT INTO pqr_audit_logs
                        (report_id, updated_by, field_changed, old_value, new_value, reason, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (report_id, user["id"], field, old_value, new_value, request.form.get("reason"), now),
                )
            update_event_felt_status(conn, report["event_id"])
            conn.commit()
            flash("PQR updated successfully. Google Sheets sync queued.")
        else:
            flash("No changes detected.")
        conn.close()
        return redirect("/pqr/list")

    conn.close()
    return render_template("pqr_edit.html", report=report)


@app.route("/export/pqr.csv")
@login_required
@permission_required("export")
def export_pqr_csv():
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    station = request.args.get("station", "").strip()
    event_key = request.args.get("event_key", "").strip()
    conn = get_db()
    where = ["1 = 1"]
    params = []
    if date_from:
        where.append("earthquake_events.event_datetime_utc >= ?")
        params.append(f"{date_from}T00:00:00Z")
    if date_to:
        where.append("earthquake_events.event_datetime_utc <= ?")
        params.append(f"{date_to}T23:59:59Z")
    if station:
        where.append("(stations.station_code LIKE ? OR stations.station_name LIKE ?)")
        params.extend([f"%{station}%", f"%{station}%"])
    if event_key:
        where.append("earthquake_events.event_key LIKE ?")
        params.append(f"%{event_key}%")
    rows = conn.execute(
        f"""
        SELECT stations.station_code AS name, pqr_reports.officer_initials,
               earthquake_events.event_key, pqr_reports.p_polarity, pqr_reports.p_arrival,
               pqr_reports.s_marker, pqr_reports.s_arrival, pqr_reports.amplitude,
               pqr_reports.duration, pqr_reports.event_type, pqr_reports.reserved_k,
               pqr_reports.remarks, pqr_reports.observed_intensities,
               pqr_reports.instrumental_intensities,
               pqr_reports.verified_areas_without_intensities,
               pqr_reports.submitted_at, pqr_reports.updated_at
        FROM pqr_reports
        JOIN stations ON stations.id = pqr_reports.station_id
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        WHERE {" AND ".join(where)}
        ORDER BY earthquake_events.event_datetime_utc DESC
        """,
        params,
    ).fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Name", "Station officers Initial", "Event Time (UTC)", "P-Polarity",
        "P-Arrival", "S", "S-Arrival", "Amplitude", "Duration", "Type",
        "Reserved", "Remarks", "OBSERVED INTENSITIES", "INSTRUMENTAL INTENSITIES",
        "Verified Areas without Intensities", "Submitted", "Updated",
    ])
    for row in rows:
        writer.writerow(row)
    filename_parts = ["pqr_reports"]
    if date_from or date_to:
        filename_parts.append(f"{date_from or 'start'}_to_{date_to or 'end'}")
    if station:
        filename_parts.append(secure_filename(station)[:40])
    if event_key:
        filename_parts.append(secure_filename(event_key)[:40])
    filename = "_".join(part for part in filename_parts if part) + ".csv"
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})


def generate_event_key(phivolcs_datetime_ph, reference_location):
    compact = phivolcs_datetime_to_utc_key(phivolcs_datetime_ph)
    suffix = location_suffix(reference_location)
    location_key = f"{compact}_{suffix}"
    conn = get_db()
    same_minute_rows = conn.execute(
        """
        SELECT id, event_key, reference_location
        FROM earthquake_events
        WHERE event_key = ?
           OR event_key LIKE ?
        """,
        (compact, f"{compact}_%"),
    ).fetchall()
    same_minute_keys = {row["event_key"] for row in same_minute_rows}
    if not same_minute_keys:
        conn.close()
        return compact

    same_location_rows = [
        row
        for row in same_minute_rows
        if location_suffix(row["reference_location"]) == suffix
    ]
    if not same_location_rows:
        conn.close()
        return location_key

    renamed = False
    for row in same_location_rows:
        if row["event_key"] in {compact, location_key}:
            first_key = f"{location_key}_1"
            if first_key not in same_minute_keys:
                conn.execute(
                    "UPDATE earthquake_events SET event_key = ? WHERE id = ?",
                    (first_key, row["id"]),
                )
                renamed = True
                same_minute_keys.discard(row["event_key"])
                same_minute_keys.add(first_key)
            break
    if renamed:
        conn.commit()

    matching_keys = [
        key
        for key in same_minute_keys
        if key == location_key or key.startswith(f"{location_key}_")
    ]
    for index in range(1, len(matching_keys) + 2):
        numbered_key = f"{location_key}_{index}"
        if numbered_key not in same_minute_keys:
            conn.close()
            return numbered_key
    conn.close()
    return f"{location_key}_{len(matching_keys) + 1}"


def upsert_phivolcs_event(conn, event):
    apply_event_region_rules(conn, event)
    if getattr(event, "source_url", ""):
        existing = conn.execute(
            "SELECT id FROM earthquake_events WHERE source_url = ?",
            (event.source_url,),
        ).fetchone()
        if existing:
            if phivolcs_url_indicates_felt(event.source_url):
                base_existing = find_primary_phivolcs_base_event(
                    conn,
                    event.event_key,
                    exclude_id=existing["id"],
                    source_url=event.source_url,
                )
                if base_existing:
                    return base_existing["id"], False
            return existing["id"], False
        if phivolcs_url_indicates_felt(event.source_url):
            base_existing = find_primary_phivolcs_base_event(conn, event.event_key, source_url=event.source_url)
            if base_existing:
                if not (base_existing["source_url"] or "").strip():
                    conn.execute(
                        "UPDATE earthquake_events SET source_url = ? WHERE id = ?",
                        (event.source_url, base_existing["id"]),
                    )
                return base_existing["id"], False

    existing = conn.execute(
        """
        SELECT earthquake_events.id
        FROM earthquake_events
        WHERE event_datetime_utc = ?
          AND reference_location = ?
          AND COALESCE(magnitude, -999) = COALESCE(?, -999)
          AND COALESCE(depth_km, -999) = COALESCE(?, -999)
          AND COALESCE(latitude, -999) = COALESCE(?, -999)
          AND COALESCE(longitude, -999) = COALESCE(?, -999)
        ORDER BY earthquake_events.id
        LIMIT 1
        """,
        (
            event.event_datetime_utc,
            event.reference_location,
            event.magnitude,
            event.depth_km,
            event.latitude,
            event.longitude,
        ),
    ).fetchone()
    if existing:
        if getattr(event, "source_url", ""):
            conn.execute(
                """
                UPDATE earthquake_events
                SET source_url = COALESCE(NULLIF(source_url, ''), ?)
                WHERE id = ?
                """,
                (event.source_url, existing["id"]),
            )
        return existing["id"], False

    existing = conn.execute(
        "SELECT id FROM earthquake_events WHERE event_key = ?",
        (event.event_key,),
    ).fetchone()
    if existing:
        return existing["id"], False

    cursor = conn.execute(
        """
        INSERT INTO earthquake_events (
            event_key, event_datetime_utc, latitude, longitude, depth_km,
            magnitude, reference_location, region_code, source, source_url
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            event.event_key,
            event.event_datetime_utc,
            event.latitude,
            event.longitude,
            event.depth_km,
            event.magnitude,
            event.reference_location,
            event.region_code,
            "PHIVOLCS",
            event.source_url,
        ),
    )
    return cursor.lastrowid, True


def event_base_key(event_key):
    text = str(event_key or "").strip()
    compact = text[:12]
    return compact if len(compact) == 12 and compact.isdigit() else ""


def phivolcs_url_time_token(url):
    match = re.search(r"/(\d{4})_(\d{4})_(\d{4,6})_B", url or "", flags=re.IGNORECASE)
    if not match:
        return None
    return {"year": match.group(1), "date": match.group(2), "time": match.group(3)}


def phivolcs_urls_are_bulletin_variants(left_url, right_url):
    if not left_url or not right_url:
        return False
    if not (phivolcs_url_indicates_felt(left_url) or phivolcs_url_indicates_felt(right_url)):
        return False
    left = phivolcs_url_time_token(left_url)
    right = phivolcs_url_time_token(right_url)
    if not left or not right:
        return False
    if left["year"] != right["year"] or left["date"] != right["date"]:
        return False
    left_time = left["time"]
    right_time = right["time"]
    if left_time == right_time:
        return True
    if len(left_time) == 4 and right_time.startswith(left_time):
        return True
    if len(right_time) == 4 and left_time.startswith(right_time):
        return True
    return False


def find_primary_phivolcs_base_event(conn, event_key, exclude_id=None, source_url=None):
    base_key = event_base_key(event_key)
    if not base_key:
        return None
    params = [base_key]
    exclude_clause = ""
    if exclude_id is not None:
        exclude_clause = "AND id != ?"
        params.append(exclude_id)
    rows = conn.execute(
        f"""
        SELECT *
        FROM earthquake_events
        WHERE substr(event_key, 1, 12) = ?
          AND COALESCE(reference_location, '') != 'Imported from Google Sheet'
          AND (source = 'PHIVOLCS' OR TRIM(COALESCE(source_url, '')) != '')
          {exclude_clause}
        ORDER BY id
        """,
        params,
    ).fetchall()
    if not rows:
        return None
    if source_url:
        variant_rows = [
            row for row in rows
            if not (row["source_url"] or "").strip()
            or phivolcs_urls_are_bulletin_variants(source_url, row["source_url"])
        ]
        if variant_rows:
            rows = variant_rows

    def rank(row):
        source_url = row["source_url"] or ""
        return (
            1 if phivolcs_url_indicates_felt(source_url) else 0,
            1 if row["exclude_from_pqr_rating"] else 0,
            -float(row["magnitude"] or 0),
            row["id"],
        )

    return sorted(rows, key=rank)[0]


def mark_phivolcs_archive_felt_hint(conn, event_id, source_url):
    if not phivolcs_url_indicates_felt(source_url):
        return False
    conn.execute(
        """
        UPDATE earthquake_events
        SET is_felt = 1,
            felt_source = 'phivolcs_f_bulletin',
            felt_checked_at = ?
        WHERE id = ?
          AND felt_override IS NULL
          AND TRIM(COALESCE(reported_intensities, '')) = ''
          AND TRIM(COALESCE(instrumental_intensities, '')) = ''
        """,
        (to_utc_iso(utc_now()), event_id),
    )
    return True


def is_import_placeholder_event(event):
    row = dict(event or {})
    return (
        (row.get("reference_location") or "").strip() == "Imported from Google Sheet"
        and row.get("magnitude") is None
        and row.get("depth_km") is None
        and row.get("latitude") is None
        and row.get("longitude") is None
    )


def is_merge_target_event(event):
    row = dict(event or {})
    return (
        not is_import_placeholder_event(row)
        and (
            row.get("source") == "PHIVOLCS"
            or bool(row.get("source_url"))
            or row.get("magnitude") is not None
            or (row.get("reference_location") or "").strip() not in {"", "Imported from Google Sheet"}
        )
    )


def event_time_diff_seconds(left_event, right_event):
    left_time = parse_utc_iso(dict(left_event or {}).get("event_datetime_utc"))
    right_time = parse_utc_iso(dict(right_event or {}).get("event_datetime_utc"))
    if not left_time or not right_time:
        return None
    return int(abs((left_time - right_time).total_seconds()))


def normalized_merge_location_suffix(event):
    suffix = location_suffix(dict(event or {}).get("reference_location") or "")
    return re.sub(r"[^a-z0-9]+", " ", suffix.lower()).strip()


def same_event_minute_key(event):
    key = str(dict(event or {}).get("event_key") or "")
    match = re.match(r"^(\d{12})", key)
    return match.group(1) if match else ""


def is_near_duplicate_event(source_event, target_event, max_minutes=2):
    source = dict(source_event or {})
    target = dict(target_event or {})
    if not source or not target:
        return False
    if source.get("id") == target.get("id"):
        return False
    if is_import_placeholder_event(source) or is_import_placeholder_event(target):
        return False
    if not is_merge_target_event(target):
        return False
    if source.get("region_code") and target.get("region_code") and source.get("region_code") != target.get("region_code"):
        return False

    source_time = parse_utc_iso(source.get("event_datetime_utc"))
    target_time = parse_utc_iso(target.get("event_datetime_utc"))
    if not source_time or not target_time or target_time <= source_time:
        return False
    diff_seconds = int((target_time - source_time).total_seconds())
    if diff_seconds <= 0 or diff_seconds > max_minutes * 60:
        return False

    source_key = same_event_minute_key(source)
    target_key = same_event_minute_key(target)
    if source_key and target_key and target_key <= source_key:
        return False

    source_mag = source.get("magnitude")
    target_mag = target.get("magnitude")
    if source_mag is not None and target_mag is not None:
        try:
            if abs(float(source_mag) - float(target_mag)) > 0.3:
                return False
        except (TypeError, ValueError):
            return False

    source_suffix = normalized_merge_location_suffix(source)
    target_suffix = normalized_merge_location_suffix(target)
    if not source_suffix or source_suffix != target_suffix:
        return False

    return True


PQR_MERGE_COPY_FIELDS = [
    "officer_initials",
    "p_polarity",
    "p_arrival",
    "s_marker",
    "s_arrival",
    "amplitude",
    "duration",
    "event_type",
    "reserved_k",
    "remarks",
    "observed_intensities",
    "instrumental_intensities",
    "verified_areas_without_intensities",
]


def pqr_report_quality_score(report):
    row = dict(report or {})
    score = 0
    remarks = (row.get("remarks") or "").strip()
    if remarks == "No Quake Record":
        score -= 1
    if remarks == "With Phase Reading":
        score += 5
    if remarks == "Intensities":
        score += 6
    if (row.get("p_arrival") or "").strip() or (row.get("s_arrival") or "").strip():
        score += 4
    if row.get("duration") not in (None, ""):
        score += 1
    if (row.get("observed_intensities") or "").strip():
        score += 6
    if (row.get("instrumental_intensities") or "").strip():
        score += 6
    if (row.get("verified_areas_without_intensities") or "").strip():
        score += 2
    return score


def replace_target_pqr_from_source(conn, source_report, target_report_id, user_id, reason, now):
    values = [source_report[field] for field in PQR_MERGE_COPY_FIELDS]
    conn.execute(
        """
        UPDATE pqr_reports
        SET officer_initials = ?,
            p_polarity = ?,
            p_arrival = ?,
            s_marker = ?,
            s_arrival = ?,
            amplitude = ?,
            duration = ?,
            event_type = ?,
            reserved_k = ?,
            remarks = ?,
            observed_intensities = ?,
            instrumental_intensities = ?,
            verified_areas_without_intensities = ?,
            updated_at = ?,
            updated_by = ?,
            sheet_sync_status = 'synced',
            sheet_synced_at = COALESCE(sheet_synced_at, ?),
            sheet_sync_error = NULL
        WHERE id = ?
        """,
        values + [now, user_id, now, target_report_id],
    )
    conn.execute(
        """
        INSERT INTO pqr_audit_logs (
            report_id, updated_by, field_changed, old_value, new_value, reason, updated_at
        )
        VALUES (?, ?, 'event_key', ?, ?, ?, ?)
        """,
        (
            target_report_id,
            user_id,
            source_report["event_key"],
            source_report["target_event_key"],
            reason,
            now,
        ),
    )


def merge_phivolcs_variant_duplicate(conn, duplicate_id, primary_id, user_id=None):
    if duplicate_id == primary_id:
        return {"moved": 0, "replaced": 0, "kept": 0}
    now = to_utc_iso(utc_now())
    primary = conn.execute("SELECT * FROM earthquake_events WHERE id = ?", (primary_id,)).fetchone()
    duplicate = conn.execute("SELECT * FROM earthquake_events WHERE id = ?", (duplicate_id,)).fetchone()
    if not primary or not duplicate:
        return {"moved": 0, "replaced": 0, "kept": 0}

    moved = 0
    replaced = 0
    kept = 0
    reports = conn.execute(
        """
        SELECT pqr_reports.*, ? AS event_key, ? AS target_event_key
        FROM pqr_reports
        WHERE event_id = ?
        ORDER BY id
        """,
        (duplicate["event_key"], primary["event_key"], duplicate_id),
    ).fetchall()
    for report in reports:
        target_report = conn.execute(
            "SELECT * FROM pqr_reports WHERE event_id = ? AND station_id = ?",
            (primary_id, report["station_id"]),
        ).fetchone()
        if not target_report:
            conn.execute(
                """
                UPDATE pqr_reports
                SET event_id = ?,
                    updated_at = ?,
                    updated_by = ?,
                    sheet_sync_status = 'synced',
                    sheet_synced_at = COALESCE(sheet_synced_at, ?),
                    sheet_sync_error = NULL
                WHERE id = ?
                """,
                (primary_id, now, user_id, now, report["id"]),
            )
            moved += 1
            continue
        if pqr_report_quality_score(report) > pqr_report_quality_score(target_report):
            replace_target_pqr_from_source(
                conn,
                report,
                target_report["id"],
                user_id,
                "Replaced weaker target station PQR while collapsing PHIVOLCS bulletin variant.",
                now,
            )
            replaced += 1
        else:
            kept += 1

    conn.execute(
        """
        UPDATE earthquake_events
        SET reported_intensities = COALESCE(NULLIF(reported_intensities, ''), ?),
            instrumental_intensities = COALESCE(NULLIF(instrumental_intensities, ''), ?),
            intensity_note = COALESCE(NULLIF(intensity_note, ''), ?),
            intensity_checked_at = COALESCE(intensity_checked_at, ?)
        WHERE id = ?
        """,
        (
            duplicate["reported_intensities"],
            duplicate["instrumental_intensities"],
            duplicate["intensity_note"],
            duplicate["intensity_checked_at"],
            primary_id,
        ),
    )
    conn.execute("DELETE FROM pqr_required_submissions WHERE event_id = ?", (duplicate_id,))
    conn.execute(
        """
        UPDATE earthquake_events
        SET status = 'closed',
            exclude_from_pqr_rating = 1,
            pqr_rating_exclusion_reason = ?,
            pqr_rating_excluded_at = ?,
            pqr_rating_excluded_by = ?,
            reference_location = ?
        WHERE id = ?
        """,
        (
            f"Collapsed into PHIVOLCS base event {primary['event_key']} as bulletin variant.",
            now,
            user_id,
            f"Collapsed into {primary['event_key']} bulletin variant",
            duplicate_id,
        ),
    )
    update_event_felt_status(conn, primary_id)
    return {"moved": moved, "replaced": replaced, "kept": kept}


def reconcile_phivolcs_bulletin_variant_duplicates(conn, user_id=None):
    rows = conn.execute(
        """
        SELECT *
        FROM earthquake_events
        WHERE COALESCE(reference_location, '') != 'Imported from Google Sheet'
          AND (source = 'PHIVOLCS' OR TRIM(COALESCE(source_url, '')) != '')
        ORDER BY event_datetime_utc, id
        """
    ).fetchall()
    groups = {}
    for row in rows:
        base_key = event_base_key(row["event_key"])
        if not base_key:
            continue
        suffix = location_suffix(row["reference_location"])
        groups.setdefault((base_key, suffix), []).append(row)

    reconciled = 0
    moved = 0
    replaced = 0
    kept = 0
    for (_base_key, _suffix), group_rows in groups.items():
        if len(group_rows) <= 1:
            continue
        if not any(phivolcs_url_indicates_felt(row["source_url"] or "") for row in group_rows):
            continue
        felt_source_url = next(
            (row["source_url"] for row in group_rows if phivolcs_url_indicates_felt(row["source_url"] or "")),
            None,
        )
        primary = find_primary_phivolcs_base_event(conn, group_rows[0]["event_key"], source_url=felt_source_url)
        if not primary:
            continue
        for row in group_rows:
            if row["id"] == primary["id"]:
                continue
            if not phivolcs_url_indicates_felt(row["source_url"] or "") and not phivolcs_url_indicates_felt(primary["source_url"] or ""):
                continue
            if not phivolcs_urls_are_bulletin_variants(row["source_url"] or "", primary["source_url"] or ""):
                continue
            result = merge_phivolcs_variant_duplicate(conn, row["id"], primary["id"], user_id)
            reconciled += 1
            moved += result["moved"]
            replaced += result["replaced"]
            kept += result["kept"]
    return {"reconciled": reconciled, "moved": moved, "replaced": replaced, "kept": kept}


def merge_conflicting_station_reports(conn, placeholder_id, target_id):
    rows = conn.execute(
        """
        SELECT stations.station_code
        FROM pqr_reports AS placeholder_reports
        JOIN pqr_reports AS target_reports
          ON target_reports.event_id = ?
         AND target_reports.station_id = placeholder_reports.station_id
        JOIN stations ON stations.id = placeholder_reports.station_id
        WHERE placeholder_reports.event_id = ?
        ORDER BY stations.station_code
        """,
        (target_id, placeholder_id),
    ).fetchall()
    return [row["station_code"] for row in rows]


def pqr_report_brief(row):
    report = dict(row or {})
    phase_parts = []
    if report.get("p_arrival"):
        phase_parts.append(f"P {report['p_arrival']}")
    if report.get("s_arrival"):
        phase_parts.append(f"S {report['s_arrival']}")
    if report.get("duration") not in (None, ""):
        phase_parts.append(f"D {report['duration']}")
    intensity_parts = []
    if (report.get("observed_intensities") or "").strip():
        intensity_parts.append(f"Obs: {report['observed_intensities']}")
    if (report.get("instrumental_intensities") or "").strip():
        intensity_parts.append(f"Inst: {report['instrumental_intensities']}")
    if (report.get("verified_areas_without_intensities") or "").strip():
        intensity_parts.append(f"Verified: {report['verified_areas_without_intensities']}")
    return {
        "remarks": report.get("remarks") or "",
        "phase": ", ".join(phase_parts) or "-",
        "intensity": " | ".join(intensity_parts) or "-",
        "updated": report.get("updated_at") or report.get("submitted_at") or "",
    }


def pqr_merge_conflict_rows(conn, placeholder_id, target_id):
    rows = conn.execute(
        """
        SELECT stations.station_code,
               stations.station_name,
               placeholder_reports.id AS placeholder_report_id,
               placeholder_reports.officer_initials AS placeholder_officer_initials,
               placeholder_reports.p_polarity AS placeholder_p_polarity,
               placeholder_reports.p_arrival AS placeholder_p_arrival,
               placeholder_reports.s_marker AS placeholder_s_marker,
               placeholder_reports.s_arrival AS placeholder_s_arrival,
               placeholder_reports.amplitude AS placeholder_amplitude,
               placeholder_reports.duration AS placeholder_duration,
               placeholder_reports.event_type AS placeholder_event_type,
               placeholder_reports.reserved_k AS placeholder_reserved_k,
               placeholder_reports.remarks AS placeholder_remarks,
               placeholder_reports.observed_intensities AS placeholder_observed_intensities,
               placeholder_reports.instrumental_intensities AS placeholder_instrumental_intensities,
               placeholder_reports.verified_areas_without_intensities AS placeholder_verified_areas_without_intensities,
               placeholder_reports.submitted_at AS placeholder_submitted_at,
               placeholder_reports.updated_at AS placeholder_updated_at,
               target_reports.id AS target_report_id,
               target_reports.officer_initials AS target_officer_initials,
               target_reports.p_polarity AS target_p_polarity,
               target_reports.p_arrival AS target_p_arrival,
               target_reports.s_marker AS target_s_marker,
               target_reports.s_arrival AS target_s_arrival,
               target_reports.amplitude AS target_amplitude,
               target_reports.duration AS target_duration,
               target_reports.event_type AS target_event_type,
               target_reports.reserved_k AS target_reserved_k,
               target_reports.remarks AS target_remarks,
               target_reports.observed_intensities AS target_observed_intensities,
               target_reports.instrumental_intensities AS target_instrumental_intensities,
               target_reports.verified_areas_without_intensities AS target_verified_areas_without_intensities,
               target_reports.submitted_at AS target_submitted_at,
               target_reports.updated_at AS target_updated_at
        FROM pqr_reports AS placeholder_reports
        JOIN pqr_reports AS target_reports
          ON target_reports.event_id = ?
         AND target_reports.station_id = placeholder_reports.station_id
        JOIN stations ON stations.id = placeholder_reports.station_id
        WHERE placeholder_reports.event_id = ?
        ORDER BY stations.station_code
        """,
        (target_id, placeholder_id),
    ).fetchall()
    conflicts = []
    for row in rows:
        row = dict(row)
        placeholder_report = {
            field: row.get(f"placeholder_{field}")
            for field in PQR_MERGE_COPY_FIELDS
        }
        placeholder_report["submitted_at"] = row.get("placeholder_submitted_at")
        placeholder_report["updated_at"] = row.get("placeholder_updated_at")
        target_report = {
            field: row.get(f"target_{field}")
            for field in PQR_MERGE_COPY_FIELDS
        }
        target_report["submitted_at"] = row.get("target_submitted_at")
        target_report["updated_at"] = row.get("target_updated_at")
        conflicts.append(
            {
                "station_code": row["station_code"],
                "station_name": row["station_name"],
                "placeholder_report_id": row["placeholder_report_id"],
                "target_report_id": row["target_report_id"],
                "placeholder": pqr_report_brief(placeholder_report),
                "target": pqr_report_brief(target_report),
            }
        )
    return conflicts


def build_event_merge_candidate(conn, placeholder_event, target_event, direction, max_minutes=10):
    if not placeholder_event or not target_event:
        return None
    placeholder = dict(placeholder_event)
    target = dict(target_event)
    is_placeholder_merge = is_import_placeholder_event(placeholder) and is_merge_target_event(target)
    is_near_duplicate_merge = is_near_duplicate_event(placeholder, target, max_minutes=2)
    if not is_placeholder_merge and not is_near_duplicate_merge:
        return None
    diff_seconds = event_time_diff_seconds(placeholder, target)
    if diff_seconds is None or diff_seconds > max_minutes * 60:
        return None
    if is_near_duplicate_merge and diff_seconds > 2 * 60:
        return None
    conflicts = merge_conflicting_station_reports(conn, placeholder["id"], target["id"])
    placeholder_count = conn.execute(
        "SELECT COUNT(*) AS count FROM pqr_reports WHERE event_id = ?",
        (placeholder["id"],),
    ).fetchone()["count"]
    if is_near_duplicate_merge and placeholder_count == 0:
        return None
    movable_count = max(placeholder_count - len(conflicts), 0)
    candidate_kind = "placeholder" if is_placeholder_merge else "near_duplicate"
    action_label = (
        ("Merge PQR" if not conflicts else "Review PQR Merge") + f" to {target['event_key']}"
        if candidate_kind == "near_duplicate"
        else None
    )
    return {
        "kind": candidate_kind,
        "direction": direction,
        "target_id": target["id"],
        "target_event_key": target["event_key"],
        "target_location": target["reference_location"],
        "target_region": target["region_code"],
        "target_magnitude": target["magnitude"],
        "target_datetime_pst": utc_iso_to_pst_display(target["event_datetime_utc"]),
        "time_diff_seconds": diff_seconds,
        "time_diff_label": f"{diff_seconds // 60}m {diff_seconds % 60}s",
        "conflicts": conflicts,
        "movable_count": movable_count,
        "safe": not conflicts,
        "action_label": action_label,
    }


def safe_return_url(value, fallback="/events/list"):
    text = str(value or "").strip()
    if text.startswith("/") and not text.startswith("//"):
        return text
    return fallback


def get_event_merge_context(conn, placeholder_id, target_id):
    placeholder = conn.execute(
        """
        SELECT earthquake_events.*,
               COUNT(pqr_reports.id) AS pqr_count
        FROM earthquake_events
        LEFT JOIN pqr_reports ON pqr_reports.event_id = earthquake_events.id
        WHERE earthquake_events.id = ?
        GROUP BY earthquake_events.id
        """,
        (placeholder_id,),
    ).fetchone()
    target = conn.execute(
        """
        SELECT earthquake_events.*,
               COUNT(pqr_reports.id) AS pqr_count
        FROM earthquake_events
        LEFT JOIN pqr_reports ON pqr_reports.event_id = earthquake_events.id
        WHERE earthquake_events.id = ?
        GROUP BY earthquake_events.id
        """,
        (target_id,),
    ).fetchone()
    candidate = build_event_merge_candidate(conn, placeholder, target, "selected", max_minutes=10)
    if not placeholder or not target or not candidate:
        return None
    conflict_rows = pqr_merge_conflict_rows(conn, placeholder_id, target_id)
    is_placeholder_source = candidate["kind"] == "placeholder"
    return {
        "placeholder": {
            **dict(placeholder),
            "event_datetime_pst": utc_iso_to_pst_display(placeholder["event_datetime_utc"]),
        },
        "target": {
            **dict(target),
            "event_datetime_pst": utc_iso_to_pst_display(target["event_datetime_utc"]),
        },
        "candidate": candidate,
        "conflict_rows": conflict_rows,
        "source_label": "Imported Placeholder" if is_placeholder_source else "Nearby Duplicate",
        "target_label": "PHIVOLCS Target" if is_placeholder_source else "Merge Target",
        "merge_help_text": (
            "Move PQR reports from an imported Google Sheet placeholder event into the selected PHIVOLCS event."
            if is_placeholder_source
            else "Move PQR reports from a suspected duplicate event into the selected official event."
        ),
    }


def merge_placeholder_pqr_reports(conn, placeholder_id, target_id, user_id=None, conflict_actions=None):
    context = get_event_merge_context(conn, placeholder_id, target_id)
    if not context:
        raise ValueError("This event pair is not eligible for merge.")
    conflict_actions = conflict_actions or {}
    placeholder = context["placeholder"]
    target = context["target"]
    source_kind = context["candidate"]["kind"]
    source_description = "imported Google Sheet placeholder" if source_kind == "placeholder" else "nearby duplicate event"
    source_location_label = "Google Sheet placeholder" if source_kind == "placeholder" else "nearby duplicate event"
    report_rows = conn.execute(
        """
        SELECT placeholder_reports.id,
               placeholder_reports.station_id,
               target_reports.id AS target_report_id
        FROM pqr_reports AS placeholder_reports
        LEFT JOIN pqr_reports AS target_reports
          ON target_reports.event_id = ?
         AND target_reports.station_id = placeholder_reports.station_id
        WHERE placeholder_reports.event_id = ?
        ORDER BY placeholder_reports.id
        """,
        (target_id, placeholder_id),
    ).fetchall()
    now = to_utc_iso(utc_now())
    moved = 0
    kept_target = 0
    replaced = 0
    skipped = 0
    for report in report_rows:
        if report["target_report_id"]:
            action = (conflict_actions.get(str(report["id"])) or "keep_target").strip()
            if action == "replace_with_placeholder":
                placeholder_report = conn.execute(
                    "SELECT * FROM pqr_reports WHERE id = ?",
                    (report["id"],),
                ).fetchone()
                if not placeholder_report:
                    skipped += 1
                    continue
                values = [placeholder_report[field] for field in PQR_MERGE_COPY_FIELDS]
                conn.execute(
                    """
                    UPDATE pqr_reports
                    SET officer_initials = ?,
                        p_polarity = ?,
                        p_arrival = ?,
                        s_marker = ?,
                        s_arrival = ?,
                        amplitude = ?,
                        duration = ?,
                        event_type = ?,
                        reserved_k = ?,
                        remarks = ?,
                        observed_intensities = ?,
                        instrumental_intensities = ?,
                        verified_areas_without_intensities = ?,
                        updated_at = ?,
                        updated_by = ?,
                        sheet_sync_status = 'synced',
                        sheet_synced_at = COALESCE(sheet_synced_at, ?),
                        sheet_sync_error = NULL
                    WHERE id = ?
                    """,
                    values + [now, user_id, now, report["target_report_id"]],
                )
                conn.execute(
                    """
                    INSERT INTO pqr_audit_logs (
                        report_id, updated_by, field_changed, old_value, new_value, reason, updated_at
                    )
                    VALUES (?, ?, 'event_key', ?, ?, ?, ?)
                    """,
                    (
                        report["target_report_id"],
                        user_id,
                        placeholder["event_key"],
                        target["event_key"],
                        f"Replaced target station PQR with conflicted {source_description} PQR during merge.",
                        now,
                    ),
                )
                replaced += 1
            elif action == "skip":
                skipped += 1
            else:
                kept_target += 1
            continue
        conn.execute(
            """
            UPDATE pqr_reports
            SET event_id = ?,
                updated_at = ?,
                updated_by = ?,
                sheet_sync_status = 'synced',
                sheet_synced_at = COALESCE(sheet_synced_at, ?),
                sheet_sync_error = NULL
            WHERE id = ?
            """,
            (target_id, now, user_id, now, report["id"]),
        )
        conn.execute(
            """
            INSERT INTO pqr_audit_logs (
                report_id, updated_by, field_changed, old_value, new_value, reason, updated_at
            )
            VALUES (?, ?, 'event_key', ?, ?, ?, ?)
            """,
            (
                report["id"],
                user_id,
                placeholder["event_key"],
                target["event_key"],
                f"Merged PQR from {source_description} into target event.",
                now,
            ),
        )
        moved += 1

    conn.execute(
        """
        UPDATE pqr_required_submissions
        SET event_id = ?
        WHERE event_id = ?
          AND NOT EXISTS (
            SELECT 1
            FROM pqr_required_submissions AS existing
            WHERE existing.event_id = ?
              AND existing.station_id = pqr_required_submissions.station_id
          )
        """,
        (target_id, placeholder_id, target_id),
    )
    conn.execute("DELETE FROM pqr_required_submissions WHERE event_id = ?", (placeholder_id,))
    conn.execute(
        """
        UPDATE earthquake_events
        SET status = 'closed',
            reference_location = ?,
            exclude_from_pqr_rating = 1,
            pqr_rating_exclusion_reason = ?,
            pqr_rating_excluded_at = ?,
            pqr_rating_excluded_by = ?
        WHERE id = ?
        """,
        (
            f"Merged into {target['event_key']} from {source_location_label}",
            f"Merged into {target['event_key']}; conflicted station PQR kept for audit only.",
            now,
            user_id,
            placeholder_id,
        ),
    )
    create_required_submissions(conn, target_id, target["region_code"], target["magnitude"], target["event_datetime_utc"])
    update_event_felt_status(conn, target_id)
    return {
        "moved": moved,
        "kept_target": kept_target,
        "replaced": replaced,
        "skipped": skipped,
        "target_key": target["event_key"],
        "placeholder_key": placeholder["event_key"],
        "source_key": placeholder["event_key"],
    }


def set_event_pqr_rating_exclusion(conn, event_id, excluded, reason="", user_id=None):
    event = conn.execute(
        "SELECT id, event_key FROM earthquake_events WHERE id = ?",
        (event_id,),
    ).fetchone()
    if not event:
        raise ValueError("Event not found.")
    event_key = event["event_key"]
    if excluded:
        clean_reason = (reason or "").strip() or "Unmatched imported Google Sheet placeholder"
        conn.execute(
            """
            UPDATE earthquake_events
            SET exclude_from_pqr_rating = 1,
                pqr_rating_exclusion_reason = ?,
                pqr_rating_excluded_at = ?,
                pqr_rating_excluded_by = ?
            WHERE id = ?
            """,
            (clean_reason[:500], to_utc_iso(utc_now()), user_id, event_id),
        )
        conn.execute("DELETE FROM pqr_required_submissions WHERE event_id = ?", (event_id,))
    else:
        conn.execute(
            """
            UPDATE earthquake_events
            SET exclude_from_pqr_rating = 0,
                pqr_rating_exclusion_reason = NULL,
                pqr_rating_excluded_at = NULL,
                pqr_rating_excluded_by = NULL
            WHERE id = ?
            """,
            (event_id,),
        )
        event = conn.execute(
            "SELECT id, region_code, magnitude, event_datetime_utc FROM earthquake_events WHERE id = ?",
            (event_id,),
        ).fetchone()
        create_required_submissions(conn, event["id"], event["region_code"], event["magnitude"], event["event_datetime_utc"])
    return event_key


def event_duplicate_status(conn, event):
    if getattr(event, "source_url", ""):
        row = conn.execute(
            "SELECT id, event_key FROM earthquake_events WHERE source_url = ?",
            (event.source_url,),
        ).fetchone()
        if row:
            return {"status": "Duplicate", "event_id": row["id"], "match": "source URL", "event_key": row["event_key"]}
        if phivolcs_url_indicates_felt(event.source_url):
            row = find_primary_phivolcs_base_event(conn, event.event_key)
            if row:
                return {"status": "Duplicate", "event_id": row["id"], "match": "felt bulletin base event", "event_key": row["event_key"]}
    row = conn.execute(
        """
        SELECT id, event_key
        FROM earthquake_events
        WHERE event_datetime_utc = ?
          AND reference_location = ?
          AND COALESCE(magnitude, -999) = COALESCE(?, -999)
          AND COALESCE(depth_km, -999) = COALESCE(?, -999)
          AND COALESCE(latitude, -999) = COALESCE(?, -999)
          AND COALESCE(longitude, -999) = COALESCE(?, -999)
        ORDER BY id
        LIMIT 1
        """,
        (
            event.event_datetime_utc,
            event.reference_location,
            event.magnitude,
            event.depth_km,
            event.latitude,
            event.longitude,
        ),
    ).fetchone()
    if row:
        return {"status": "Duplicate", "event_id": row["id"], "match": "event fingerprint", "event_key": row["event_key"]}
    row = conn.execute(
        "SELECT id, event_key FROM earthquake_events WHERE event_key = ?",
        (event.event_key,),
    ).fetchone()
    if row:
        return {"status": "Duplicate", "event_id": row["id"], "match": "event key", "event_key": row["event_key"]}
    return {"status": "New", "event_id": None, "match": "", "event_key": ""}


def available_phivolcs_archive_months(year):
    current_local = utc_now().astimezone(APP_LOCAL_TZ).date()
    try:
        year = int(year)
    except (TypeError, ValueError):
        year = current_local.year
    max_month = 12
    if year == current_local.year:
        max_month = current_local.month
    return [
        {"value": str(month), "label": MONTH_NAMES[month - 1]}
        for month in range(1, max_month + 1)
    ]


def selected_phivolcs_archive_months(year, month_value):
    month_options = available_phivolcs_archive_months(year)
    valid_months = [int(option["value"]) for option in month_options]
    if month_value == "all":
        return valid_months
    try:
        month = int(month_value)
    except (TypeError, ValueError):
        month = valid_months[-1] if valid_months else 1
    return [month] if month in valid_months else [valid_months[-1]]


def phivolcs_archive_import_token(event):
    source = f"{getattr(event, 'source_url', '')}|{getattr(event, 'event_key', '')}|{getattr(event, 'event_datetime_utc', '')}"
    return hashlib.sha1(source.encode("utf-8")).hexdigest()[:16]


def scan_phivolcs_archive_events(conn, year, month_value, limit_per_month=5000):
    events = []
    errors = []
    for month in selected_phivolcs_archive_months(year, month_value):
        try:
            month_events = fetch_monthly_archive_events(year, month, limit=limit_per_month)
        except Exception as error:
            errors.append(f"{MONTH_NAMES[month - 1]} {year}: {error}")
            continue
        events.extend(apply_event_region_rules(conn, event) for event in month_events)
    preview_rows = []
    for event in events:
        duplicate = event_duplicate_status(conn, event)
        preview_rows.append(
            {
                "event": event,
                "status": duplicate["status"],
                "match": duplicate["match"],
                "existing_event_key": duplicate["event_key"],
                "felt_hint": phivolcs_url_indicates_felt(event.source_url),
                "event_datetime_pst": utc_iso_to_pst_display(event.event_datetime_utc),
                "import_token": phivolcs_archive_import_token(event),
            }
        )
    return preview_rows, errors


def event_minute_key(event_datetime_utc):
    event_time = parse_utc_iso(event_datetime_utc)
    if not event_time:
        return ""
    return event_time.strftime("%Y%m%d%H%M")


def normalize_duplicate_event_keys(conn):
    rows = conn.execute(
        """
        SELECT id, event_key, event_datetime_utc, reference_location
        FROM earthquake_events
        WHERE COALESCE(exclude_from_pqr_rating, 0) = 0
        ORDER BY event_datetime_utc, id
        """
    ).fetchall()
    groups = {}
    for row in rows:
        compact = event_minute_key(row["event_datetime_utc"])
        if not compact:
            continue
        groups.setdefault(compact, []).append(dict(row))

    updates = []
    for compact, group_rows in groups.items():
        if len(group_rows) <= 1:
            row = group_rows[0]
            desired_key = compact
            if row["event_key"] != desired_key:
                updates.append((row["id"], row["event_key"], desired_key))
            continue
        suffix_counts = {}
        for row in group_rows:
            suffix = location_suffix(row["reference_location"])
            base_key = f"{compact}_{suffix}" if suffix else compact
            suffix_counts[base_key] = suffix_counts.get(base_key, 0) + 1
            desired_key = base_key
            if suffix_counts[base_key] > 1:
                desired_key = f"{base_key}_{suffix_counts[base_key]}"
            if row["event_key"] != desired_key:
                updates.append((row["id"], row["event_key"], desired_key))

    if not updates:
        return 0

    desired_keys = [new_key for _event_id, _old_key, new_key in updates]
    if desired_keys:
        placeholders = ",".join("?" for _ in desired_keys)
        active_ids = [event_id for event_id, _old_key, _new_key in updates]
        active_placeholders = ",".join("?" for _ in active_ids)
        colliding_rows = conn.execute(
            f"""
            SELECT id, event_key
            FROM earthquake_events
            WHERE event_key IN ({placeholders})
              AND id NOT IN ({active_placeholders})
            """,
            desired_keys + active_ids,
        ).fetchall()
        for row in colliding_rows:
            conn.execute(
                "UPDATE earthquake_events SET event_key = ? WHERE id = ?",
                (f"{row['event_key']}_AUDIT_{row['id']}", row["id"]),
            )

    for event_id, _old_key, _new_key in updates:
        conn.execute(
            "UPDATE earthquake_events SET event_key = ? WHERE id = ?",
            (f"__renaming_{event_id}__", event_id),
        )
    for event_id, _old_key, new_key in updates:
        conn.execute(
            "UPDATE earthquake_events SET event_key = ? WHERE id = ?",
            (new_key, event_id),
        )
    return len(updates)


def refresh_event_bulletin_intensities(conn, event_id, source_url, force=False):
    if not source_url:
        return False

    existing = conn.execute(
        """
        SELECT reported_intensities, instrumental_intensities, intensity_checked_at
        FROM earthquake_events
        WHERE id = ?
        """,
        (event_id,),
    ).fetchone()
    if not existing:
        return False

    has_intensities = bool(
        (existing["reported_intensities"] or "").strip()
        or (existing["instrumental_intensities"] or "").strip()
    )
    checked_at = parse_utc_iso(existing["intensity_checked_at"])
    recently_checked = (
        checked_at
        and (utc_now() - checked_at).total_seconds() < 30 * 60
    )
    if not force and (has_intensities or recently_checked):
        update_event_felt_status(conn, event_id)
        return has_intensities

    try:
        reported, instrumental, note = extract_bulletin_intensities(source_url)
    except Exception as error:
        reported, instrumental, note = "", "", str(error)

    conn.execute(
        """
        UPDATE earthquake_events
        SET reported_intensities = ?,
            instrumental_intensities = ?,
            intensity_note = ?,
            intensity_checked_at = ?
        WHERE id = ?
        """,
        (
            reported,
            instrumental,
            (note or "")[:500],
            to_utc_iso(utc_now()),
            event_id,
        ),
    )
    return update_event_felt_status(conn, event_id)


def create_required_submissions(conn, event_id, region_code, magnitude=None, event_datetime_utc=None):
    event = conn.execute(
        "SELECT exclude_from_pqr_rating FROM earthquake_events WHERE id = ?",
        (event_id,),
    ).fetchone()
    if event and event["exclude_from_pqr_rating"]:
        conn.execute("DELETE FROM pqr_required_submissions WHERE event_id = ?", (event_id,))
        return
    stations = stations_required_for_event(conn, region_code, magnitude, event_datetime_utc)
    conn.executemany(
        """
        INSERT OR IGNORE INTO pqr_required_submissions (event_id, station_id, status)
        VALUES (?, ?, 'pending')
        """,
        [(event_id, station["id"]) for station in stations],
    )


def run_phivolcs_sync(hours=22, limit=120):
    conn = get_db()
    started_at = to_utc_iso(utc_now())
    existing_run = conn.execute(
        """
        SELECT id, started_at
        FROM sync_runs
        WHERE source = 'PHIVOLCS'
          AND status = 'running'
        ORDER BY started_at DESC
        LIMIT 1
        """
    ).fetchone()
    if existing_run:
        started = parse_utc_iso(existing_run["started_at"])
        elapsed = (utc_now() - started).total_seconds() if started else 0
        if elapsed < 15 * 60:
            conn.close()
            return {"status": "already_running", "imported_count": 0, "error_message": ""}
        conn.execute(
            """
            UPDATE sync_runs
            SET status = 'failed',
                finished_at = ?,
                error_message = 'Sync run timed out and was marked stale.'
            WHERE id = ?
            """,
            (to_utc_iso(utc_now()), existing_run["id"]),
        )
        conn.commit()

    cursor = conn.execute(
        """
        INSERT INTO sync_runs (source, status, started_at)
        VALUES ('PHIVOLCS', 'running', ?)
        """,
        (started_at,),
    )
    sync_run_id = cursor.lastrowid
    conn.commit()
    conn.close()

    imported = 0
    conn = None
    try:
        conn = get_db()
        events = [apply_event_region_rules(conn, event) for event in fetch_recent_events(hours=hours, limit=limit)]
        close_expired_events(conn)
        for event in events:
            event_id, created = upsert_phivolcs_event(conn, event)
            refresh_event_bulletin_intensities(conn, event_id, event.source_url)
            create_required_submissions(conn, event_id, event.region_code, event.magnitude, event.event_datetime_utc)
            if created:
                imported += 1
        reconcile_phivolcs_bulletin_variant_duplicates(conn)
        normalize_duplicate_event_keys(conn)
        reconcile_required_submissions(conn)
        conn.execute(
            """
            UPDATE sync_runs
            SET status = 'success',
                finished_at = ?,
                imported_count = ?
            WHERE id = ?
            """,
            (to_utc_iso(utc_now()), imported, sync_run_id),
        )
        conn.commit()
        conn.close()
        conn = None
        return {"status": "success", "imported_count": imported, "error_message": ""}
    except Exception as error:
        if conn:
            try:
                conn.rollback()
                conn.close()
            except Exception:
                pass
        failure_conn = get_db()
        try:
            failure_conn.execute(
                """
                UPDATE sync_runs
                SET status = 'failed',
                    finished_at = ?,
                    error_message = ?
                WHERE id = ?
                """,
                (to_utc_iso(utc_now()), str(error), sync_run_id),
            )
            failure_conn.commit()
        finally:
            failure_conn.close()
        return {"status": "failed", "imported_count": imported, "error_message": str(error)}


def magnitude_triggers_all_stations(magnitude):
    try:
        return float(magnitude or 0) >= 5
    except (TypeError, ValueError):
        return False


def event_visible_for_station_code(event, station_code, station_region_code=None):
    if magnitude_triggers_all_stations(event.get("magnitude")):
        return True
    cluster_name = cluster_name_for_station(station_code, station_region_code)
    return event.get("region_code") in CLUSTER_EVENT_REGIONS.get(cluster_name, ())


def event_visible_for_station(event, station):
    station_data = dict(station or {})
    if magnitude_triggers_all_stations(event.get("magnitude")):
        return True
    cluster_name = cluster_name_for_station(station_data)
    return event.get("region_code") in CLUSTER_EVENT_REGIONS.get(cluster_name, ())


def event_visible_for_station_legacy(event, station):
    station_data = dict(station or {})
    return event_visible_for_station_code(
        event,
        station_data.get("station_code"),
        station_data.get("region_code"),
    )


def required_station_codes_for_event(region_code, magnitude=None):
    if magnitude_triggers_all_stations(magnitude):
        return PQR_STATION_CODES
    return tuple(
        station_code
        for station_code in PQR_STATION_CODES
        if region_code in CLUSTER_EVENT_REGIONS.get(STATION_CLUSTER_BY_CODE.get(station_code, ""), ())
    )


def required_station_codes_for_region(region_code):
    return required_station_codes_for_event(region_code)


def stations_required_for_event(conn, region_code, magnitude=None, event_datetime_utc=None):
    stations = conn.execute(
        """
        SELECT id, station_code, region_code, cluster_name, is_one_manned
        FROM stations
        WHERE is_active = 1
        """,
    ).fetchall()
    stations = [
        station
        for station in stations
        if event_visible_for_station(
            {"region_code": region_code, "magnitude": magnitude},
            station,
        )
    ]
    if not event_datetime_utc:
        return stations
    return [
        station
        for station in stations
        if not station_exempt_from_pqr(station, {"event_datetime_utc": event_datetime_utc})
    ]


def reconcile_required_submissions(conn):
    station_rows = conn.execute(
        """
        SELECT id, station_code, region_code, cluster_name,
               station_type, include_in_pqr_compliance, is_one_manned
        FROM stations
        WHERE is_active = 1
          AND include_in_pqr_compliance = 1
        """
    ).fetchall()
    event_rows = conn.execute(
        """
        SELECT id, region_code, event_datetime_utc, magnitude
        FROM earthquake_events
        WHERE status = 'open'
          AND COALESCE(exclude_from_pqr_rating, 0) = 0
        """
    ).fetchall()

    valid_pairs = set()
    for event in event_rows:
        if not within_update_window(event["event_datetime_utc"]):
            continue
        for station in station_rows:
            if not event_visible_for_station(dict(event), station):
                continue
            if station_exempt_from_pqr(station, dict(event)):
                continue
            valid_pairs.add((event["id"], station["id"]))

    conn.executemany(
        """
        INSERT OR IGNORE INTO pqr_required_submissions (event_id, station_id, status)
        VALUES (?, ?, 'pending')
        """,
        list(valid_pairs),
    )

    existing_rows = conn.execute(
        "SELECT id, event_id, station_id FROM pqr_required_submissions"
    ).fetchall()
    obsolete_ids = [
        row["id"]
        for row in existing_rows
        if (row["event_id"], row["station_id"]) not in valid_pairs
    ]
    if obsolete_ids:
        conn.executemany(
            "DELETE FROM pqr_required_submissions WHERE id = ?",
            [(row_id,) for row_id in obsolete_ids],
        )

    conn.execute(
        """
        UPDATE pqr_required_submissions
        SET status = 'submitted'
        WHERE EXISTS (
            SELECT 1
            FROM pqr_reports
            WHERE pqr_reports.event_id = pqr_required_submissions.event_id
              AND pqr_reports.station_id = pqr_required_submissions.station_id
        )
        """
    )
    conn.execute(
        """
        UPDATE pqr_required_submissions
        SET status = 'pending'
        WHERE NOT EXISTS (
            SELECT 1
            FROM pqr_reports
            WHERE pqr_reports.event_id = pqr_required_submissions.event_id
              AND pqr_reports.station_id = pqr_required_submissions.station_id
        )
        """
    )
    return {"valid": len(valid_pairs), "obsolete_removed": len(obsolete_ids)}


def close_expired_events(conn):
    rows = conn.execute(
        """
        SELECT id, event_datetime_utc
        FROM earthquake_events
        WHERE status = 'open'
        """
    ).fetchall()
    expired_ids = [
        row["id"]
        for row in rows
        if not within_update_window(row["event_datetime_utc"])
    ]
    if expired_ids:
        conn.executemany(
            """
            UPDATE earthquake_events
            SET status = 'closed'
            WHERE id = ?
            """,
            [(event_id,) for event_id in expired_ids],
        )
    return len(expired_ids)


def mark_sheet_sync_success(conn, report_ids):
    now = to_utc_iso(utc_now())
    conn.executemany(
        """
        UPDATE pqr_reports
        SET sheet_sync_status = 'synced',
            sheet_synced_at = ?,
            sheet_sync_error = NULL
        WHERE id = ?
        """,
        [(now, report_id) for report_id in report_ids],
    )


def mark_sheet_sync_failure(conn, report_ids, error):
    error_message = friendly_sheet_sync_error(error)
    conn.executemany(
        """
        UPDATE pqr_reports
        SET sheet_sync_status = 'failed',
            sheet_sync_error = ?
        WHERE id = ?
        """,
        [(error_message, report_id) for report_id in report_ids],
    )


def friendly_sheet_sync_error(error):
    text = str(error)
    if "RATE_LIMIT_EXCEEDED" in text or "Quota exceeded" in text or "HttpError 429" in text:
        return (
            "Google Sheets quota limit was reached. PQR data is saved locally and will sync "
            "when the next retry is allowed."
        )
    return text[:500]


def sync_google_reports_once(conn, report_ids):
    if not report_ids:
        return {"status": "empty", "appended_count": 0, "error_message": ""}
    try:
        appended_count = append_pqr_reports_to_sheet(conn, report_ids)
        mark_sheet_sync_success(conn, report_ids)
        conn.commit()
        return {"status": "success", "appended_count": appended_count, "error_message": ""}
    except Exception as error:
        mark_sheet_sync_failure(conn, report_ids, error)
        conn.commit()
        return {"status": "failed", "appended_count": 0, "error_message": friendly_sheet_sync_error(error)}


def sync_pending_google_reports(limit=300):
    conn = get_db()
    failed = conn.execute(
        """
        SELECT pqr_reports.id
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        WHERE pqr_reports.sheet_sync_status IN ('pending', 'failed')
          AND pqr_reports.created_by IS NOT NULL
          AND COALESCE(earthquake_events.exclude_from_pqr_rating, 0) = 0
        ORDER BY pqr_reports.submitted_at
        LIMIT ?
        """,
        (limit,),
    ).fetchall()
    report_ids = [row["id"] for row in failed]
    if not report_ids:
        conn.close()
        return {"status": "empty", "appended_count": 0, "error_message": ""}
    result = sync_google_reports_once(conn, report_ids)
    conn.close()
    return result


def backfill_required_submissions(conn):
    events = conn.execute(
        """
        SELECT id, region_code, event_datetime_utc, magnitude
        FROM earthquake_events
        WHERE status = 'open'
          AND COALESCE(exclude_from_pqr_rating, 0) = 0
        """
    ).fetchall()
    for event in events:
        if within_update_window(event["event_datetime_utc"]):
            create_required_submissions(
                conn,
                event["id"],
                event["region_code"],
                event["magnitude"],
                event["event_datetime_utc"],
            )


def pqr_values_from_form(form):
    return (
        form["event_id"],
        form["station_id"],
        form["officer_initials"].strip(),
        form.get("p_polarity"),
        form.get("p_arrival"),
        form.get("s_marker"),
        form.get("s_arrival"),
        optional_float(form.get("amplitude")),
        optional_float(form.get("duration")),
        form.get("event_type"),
        form.get("reserved_k"),
        form["remarks"].strip(),
        form.get("observed_intensities"),
        form.get("instrumental_intensities"),
        form.get("verified_areas_without_intensities"),
    )


PQR_ROW_FIELDS = [
    "p_polarity",
    "reserved_k",
    "p_arrival",
    "s_marker",
    "s_arrival",
    "amplitude",
    "duration",
    "event_type",
    "remarks",
    "observed_intensities",
    "instrumental_intensities",
    "verified_areas_without_intensities",
]


def extract_pqr_row(form, event_id):
    row = {
        "event_id": str(event_id),
        "station_id": form.get("station_id", ""),
        "officer_initials": form.get("officer_initials", "").strip(),
    }
    for key in PQR_ROW_FIELDS:
        row[key] = form.get(f"{key}_{event_id}", "").strip()
    return row


def submit_pqr_rows(conn, form, user):
    can_submit_outside_pending = user["role"] == "admin" or is_effective_duty_officer(user)
    event_ids = [event_id for event_id in form.getlist("event_ids") if str(event_id).strip()]
    if not event_ids:
        event_ids = [
            key.removeprefix("remarks_")
            for key in form.keys()
            if key.startswith("remarks_") and form.get(key, "").strip()
        ]
    if not event_ids:
        return ["Add Remarks to at least one event row before submitting."], 0, []
    event_ids = [
        event_id
        for event_id in event_ids
        if form.get(f"remarks_{event_id}", "").strip()
    ]
    if not event_ids:
        return ["Add Remarks to at least one selected event before submitting."], 0, []

    errors = []
    rows = []
    for event_id in event_ids:
        row = extract_pqr_row(form, event_id)
        row_errors = validate_pqr_form_strict(row)
        if row_errors:
            errors.extend([f"Event {event_id}: {error}" for error in row_errors])
            continue
        if user["role"] == "station_user" and not is_effective_duty_officer(user) and not user_can_access_station(conn, user, row["station_id"], "submit"):
            errors.append(f"Event {event_id}: Station users can submit only for assigned stations.")
            continue
        station = conn.execute(
            """
            SELECT station_code, region_code, is_one_manned
            FROM stations
            WHERE id = ?
            """,
            (row["station_id"],),
        ).fetchone()
        if not station:
            errors.append(f"Event {event_id}: Selected station was not found.")
            continue
        if can_submit_outside_pending:
            existing_report = conn.execute(
                """
                SELECT id, sheet_sync_status
                FROM pqr_reports
                WHERE event_id = ? AND station_id = ?
                """,
                (row["event_id"], row["station_id"]),
            ).fetchone()
            if existing_report and existing_report["sheet_sync_status"] == "synced":
                errors.append(
                    f"Event {event_id}: This Event Key was already submitted in the database and Google Sheet."
                )
                continue
            if existing_report:
                row["report_id"] = existing_report["id"]
            event = conn.execute(
                """
                SELECT event_datetime_utc, region_code, magnitude
                FROM earthquake_events
                WHERE id = ?
                """,
                (row["event_id"],),
            ).fetchone()
            if not event:
                errors.append(f"Event {event_id}: Event Key not found.")
                continue
            if not event_visible_for_station(dict(event), station):
                errors.append(
                    f"Event {event_id}: This Event Key is outside the selected station cluster and below magnitude 5."
                )
                continue
            if station_exempt_from_pqr(station, dict(event)):
                errors.append(f"Event {event_id}: This station is 1M-exempt for the event time.")
                continue
        else:
            required = conn.execute(
                """
                SELECT earthquake_events.event_datetime_utc
                FROM pqr_required_submissions
                JOIN earthquake_events
                    ON earthquake_events.id = pqr_required_submissions.event_id
                WHERE pqr_required_submissions.event_id = ?
                  AND pqr_required_submissions.station_id = ?
                  AND pqr_required_submissions.status = 'pending'
                  AND earthquake_events.status = 'open'
                """,
                (row["event_id"], row["station_id"]),
            ).fetchone()
            if not required:
                errors.append(f"Event {event_id}: This event is not pending for the selected station.")
                continue
            if user["role"] == "station_user" and not is_effective_duty_officer(user) and not within_update_window(required["event_datetime_utc"]):
                errors.append(f"Event {event_id}: The 22-hour submission window has expired.")
                continue
        rows.append(row)

    if errors:
        return errors, 0, []

    submitted_at = to_utc_iso(utc_now())
    report_ids = []
    for row in rows:
        if row.get("report_id"):
            conn.execute(
                """
                UPDATE pqr_reports
                SET officer_initials = ?, p_polarity = ?, p_arrival = ?,
                    s_marker = ?, s_arrival = ?, amplitude = ?, duration = ?,
                    event_type = ?, reserved_k = ?, remarks = ?,
                    observed_intensities = ?, instrumental_intensities = ?,
                    verified_areas_without_intensities = ?, updated_at = ?,
                    updated_by = ?, sheet_sync_status = 'pending',
                    sheet_sync_error = NULL
                WHERE id = ?
                  AND sheet_sync_status != 'synced'
                """,
                (
                    row["officer_initials"].strip(),
                    row.get("p_polarity"),
                    row.get("p_arrival"),
                    row.get("s_marker"),
                    row.get("s_arrival"),
                    optional_float(row.get("amplitude")),
                    optional_float(row.get("duration")),
                    row.get("event_type"),
                    row.get("reserved_k"),
                    row["remarks"].strip(),
                    row.get("observed_intensities"),
                    row.get("instrumental_intensities"),
                    row.get("verified_areas_without_intensities"),
                    submitted_at,
                    user["id"],
                    row["report_id"],
                ),
            )
            report_ids.append(row["report_id"])
        else:
            cursor = conn.execute(
                """
                INSERT INTO pqr_reports (
                    event_id, station_id, officer_initials, p_polarity, p_arrival,
                    s_marker, s_arrival, amplitude, duration, event_type, reserved_k,
                    remarks, observed_intensities, instrumental_intensities,
                    verified_areas_without_intensities, submitted_at, created_by
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                pqr_values_from_form(row) + (submitted_at, user["id"]),
            )
            report_ids.append(cursor.lastrowid)
        conn.execute(
            """
            UPDATE pqr_required_submissions
            SET status = 'submitted'
            WHERE event_id = ? AND station_id = ?
            """,
            (row["event_id"], row["station_id"]),
        )
        update_event_felt_status(conn, row["event_id"])

    return [], len(rows), report_ids


def normalize_grid_pqr_form(form):
    event_id = form.get("event_id", "")
    normalized = dict(form)
    for key in [
        "p_polarity",
        "p_arrival",
        "s_marker",
        "s_arrival",
        "amplitude",
        "duration",
        "event_type",
        "reserved_k",
        "remarks",
        "observed_intensities",
        "instrumental_intensities",
        "verified_areas_without_intensities",
    ]:
        row_key = f"{key}_{event_id}"
        if row_key in form:
            normalized[key] = form.get(row_key)
    return normalized


def pqr_update_fields_from_form(form):
    keys = [
        "officer_initials", "p_polarity", "p_arrival", "s_marker", "s_arrival",
        "amplitude", "duration", "event_type", "reserved_k", "remarks",
        "observed_intensities", "instrumental_intensities",
        "verified_areas_without_intensities",
    ]
    fields = {}
    for key in keys:
        if key in {"amplitude", "duration"}:
            fields[key] = optional_float(form.get(key))
        else:
            fields[key] = request.form.get(key)
    fields["officer_initials"] = (fields["officer_initials"] or "").strip()
    fields["remarks"] = (fields["remarks"] or "").strip()
    return fields


def fetch_report(conn, report_id):
    return conn.execute(
        """
        SELECT pqr_reports.*, earthquake_events.event_key,
               earthquake_events.event_datetime_utc, stations.station_code,
               stations.station_name
        FROM pqr_reports
        JOIN earthquake_events ON earthquake_events.id = pqr_reports.event_id
        JOIN stations ON stations.id = pqr_reports.station_id
        WHERE pqr_reports.id = ?
        """,
        (report_id,),
    ).fetchone()


def build_changes(report, fields):
    changes = []
    for field, new_value in fields.items():
        old_value = report[field]
        if str(old_value or "") != str(new_value or ""):
            changes.append((field, old_value, new_value))
    return changes


if __name__ == "__main__":
    app.run(debug=True)
